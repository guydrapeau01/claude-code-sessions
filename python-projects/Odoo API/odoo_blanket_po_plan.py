import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, timezone, date
from math import ceil

# ================================================================
# CONFIG
# ================================================================
PLAN_PRODUCTS    = cfg.ALL_PLAN_PRODUCTS
MONTHLY_PLAN     = cfg.MONTHLY_PRODUCTION_PLAN
BLANKET_START    = date(2026, 3, 1)   # Mar 2026
BLANKET_END      = date(2026, 12, 31) # Dec 2026
BLANKET_MONTHS   = 10  # Mar → Dec 2026
SAFETY_MONTHS    = 1
DELIVERY_FREQ    = 1   # monthly deliveries

# Months in blanket
def add_months(d, months):
    month = d.month - 1 + months
    year  = d.year + month // 12
    month = month % 12 + 1
    return date(year, month, 1)

MONTHS = [add_months(BLANKET_START, i) for i in range(BLANKET_MONTHS)]
MONTH_LABELS = [m.strftime('%b %Y') for m in MONTHS]

# ================================================================
# HELPERS
# ================================================================
def make_fill(hex_):   return PatternFill("solid", fgColor=hex_)
def tb():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, val, bg, fg="FFFFFF", bold=True, wrap=True, size=10, left=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(color=fg, bold=bold, size=size)
    c.fill      = make_fill(bg)
    c.alignment = Alignment(horizontal="left" if left else "center",
                            vertical="center", wrap_text=wrap)
    c.border    = tb()
    return c

def dat(ws, row, col, val, fill, left=False, bold=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill      = fill
    c.border    = tb()
    c.font      = Font(bold=bold)
    c.alignment = Alignment(horizontal="left" if left else "center", vertical="center")
    if num_fmt:
        c.number_format = num_fmt
    return c

COLORS = {
    'main':    "2E4057",
    'urgent':  "C00000",
    'warn':    "BF8F00",
    'ok':      "375623",
    'sup':     "1F4E79",
    'alt':     "F2F2F2",
    'white':   "FFFFFF",
    'red':     "FFB3B3",
    'amber':   "FFD9B3",
    'yellow':  "FFF2CC",
    'green':   "E2EFDA",
    'blue':    "DDEBF7",
}

def explode_bom(models, uid, bom_id, qty_needed, bom_cache, result, depth=0):
    if bom_id not in bom_cache:
        lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom.line', 'search_read',
            [[['bom_id', '=', bom_id]]],
            {'fields': ['product_id', 'product_qty', 'product_uom_id', 'child_bom_id']}
        )
        bom_info = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'read',
            [[bom_id]], {'fields': ['product_qty']}
        )
        bom_cache[bom_id] = {'lines': lines, 'bom_qty': bom_info[0]['product_qty'] if bom_info else 1}
    cached  = bom_cache[bom_id]
    ratio   = qty_needed / cached['bom_qty']
    for line in cached['lines']:
        comp_id   = line['product_id'][0]
        comp_name = line['product_id'][1]
        comp_qty  = line['product_qty'] * ratio
        uom       = line['product_uom_id'][1] if line.get('product_uom_id') else 'Unit'
        if line.get('child_bom_id'):
            explode_bom(models, uid, line['child_bom_id'][0], comp_qty, bom_cache, result, depth+1)
        else:
            if comp_id not in result:
                result[comp_id] = {'name': comp_name, 'qty_per_month': 0, 'uom': uom}
            result[comp_id]['qty_per_month'] += comp_qty / qty_needed  # per unit of finished

try:
    print("Connecting...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now = datetime.now(timezone.utc)

    # --- Resolve finished products & BOMs ---
    fin_prods = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', PLAN_PRODUCTS]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    prod_by_ref = {p['default_code']: p for p in fin_prods}

    bom_by_ref = {}
    for ref in PLAN_PRODUCTS:
        boms = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'search_read',
            [[['product_tmpl_id.default_code', '=', ref]]],
            {'fields': ['id', 'product_qty']}
        )
        if boms:
            bom_by_ref[ref] = boms[0]
            print(f"  BOM for {ref}: ID {boms[0]['id']}")
        else:
            print(f"  ⚠ No BOM for {ref}")

    # --- Explode BOMs — accumulate qty per month across all finished products ---
    # comp_id -> qty needed per month (sum across all finished products)
    print("\nExploding BOMs...")
    bom_cache  = {}
    # comp_id -> {name, qty_per_month (weighted by monthly plan), uom}
    comp_monthly = {}  # comp_id -> monthly qty needed

    # Also track per-finished-product BOM: comp_id -> {ref -> qty_per_unit}
    bom_by_product = {}  # ref -> {comp_id -> qty_per_unit}

    for ref, monthly_qty in MONTHLY_PLAN.items():
        if ref not in bom_by_ref:
            continue
        result = {}
        explode_bom(models, uid, bom_by_ref[ref]['id'], 1, bom_cache, result)
        bom_by_product[ref] = {comp_id: info['qty_per_month'] for comp_id, info in result.items()}
        for comp_id, info in result.items():
            if comp_id not in comp_monthly:
                comp_monthly[comp_id] = {'name': info['name'], 'monthly': 0, 'uom': info['uom']}
            comp_monthly[comp_id]['monthly'] += info['qty_per_month'] * monthly_qty

    # Also explode additional models (no monthly plan, just include components)
    extra_refs = [r for r in PLAN_PRODUCTS if r not in MONTHLY_PLAN]
    for ref in extra_refs:
        if ref not in bom_by_ref:
            continue
        result = {}
        explode_bom(models, uid, bom_by_ref[ref]['id'], 1, bom_cache, result)
        bom_by_product[ref] = {comp_id: info['qty_per_month'] for comp_id, info in result.items()}
        for comp_id, info in result.items():
            if comp_id not in comp_monthly:
                comp_monthly[comp_id] = {'name': info['name'], 'monthly': 0, 'uom': info['uom']}
            # mark as shared component even if no volume
    all_comp_ids = list(comp_monthly.keys())
    print(f"  {len(all_comp_ids)} unique components\n")

    # --- Component product details ---
    comp_prods = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['id', 'in', all_comp_ids]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    comp_info = {p['id']: p for p in comp_prods}
    comp_tmpl_ids = [p['product_tmpl_id'][0] for p in comp_prods]

    # --- Supplier info ---
    print("Fetching supplier info...")
    suppliers = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.supplierinfo', 'search_read',
        [[['product_tmpl_id', 'in', comp_tmpl_ids]]],
        {'fields': ['product_tmpl_id', 'name', 'delay', 'min_qty', 'price',
                    'product_code', 'sequence']}
    )
    sup_by_tmpl = {}
    for s in sorted(suppliers, key=lambda x: x.get('sequence', 99)):
        tmpl_id = s['product_tmpl_id'][0]
        if tmpl_id not in sup_by_tmpl:
            sup_by_tmpl[tmpl_id] = s

    # --- Current stock ---
    print("Fetching stock...")
    quants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.quant', 'search_read',
        [[['product_id', 'in', all_comp_ids],
          ['location_id.usage', '=', 'internal']]],
        {'fields': ['product_id', 'quantity', 'reserved_quantity']}
    )
    # Sum all quant rows per product first, then take max(0)
    # (individual rows can be negative due to Odoo's double-entry; net is what matters)
    _stock_raw = defaultdict(lambda: [0.0, 0.0])  # [total_qty, total_reserved]
    for q in quants:
        pid_ = q['product_id'][0]
        _stock_raw[pid_][0] += q['quantity']
        _stock_raw[pid_][1] += q['reserved_quantity']
    stock_by_comp = defaultdict(float)
    for pid_, (qty, res) in _stock_raw.items():
        stock_by_comp[pid_] = max(0, qty - res)

    # --- Open incoming moves ---
    print("Fetching open incoming...")
    moves = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.move', 'search_read',
        [[['product_id', 'in', all_comp_ids],
          ['state', 'in', ['waiting', 'confirmed', 'assigned', 'partially_available']],
          ['picking_type_id.code', '=', 'incoming']]],
        {'fields': ['product_id', 'product_qty', 'quantity_done', 'date', 'picking_id']}
    )
    # Get picking states to determine true remaining qty
    # If picking is not done, use full product_qty (ignore quantity_done which can be unreliable)
    pick_ids = list({m['picking_id'][0] for m in moves if m.get('picking_id')})
    pick_states = {}
    if pick_ids:
        picks = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.picking', 'read',
            [pick_ids], {'fields': ['id', 'state']}
        )
        pick_states = {p['id']: p['state'] for p in picks}

    incoming_by_comp = defaultdict(float)
    for m in moves:
        pick_id    = m['picking_id'][0] if m.get('picking_id') else None
        pick_state = pick_states.get(pick_id, '')
        ordered    = m.get('product_qty') or 0
        done       = m.get('quantity_done') or 0
        # If picking is not done yet, remaining = ordered - actually_received
        # quantity_done on a non-done picking is just a draft entry, not confirmed
        if pick_state != 'done':
            remaining = ordered  # full qty still expected
        else:
            remaining = ordered - done  # partial receipt
        if remaining > 0:
            incoming_by_comp[m['product_id'][0]] += remaining

    # ================================================================
    # CALCULATE BLANKET PO PLAN
    # ================================================================
    print("\nCalculating blanket PO plan...")
    plan_rows = []

    for comp_id in all_comp_ids:
        info      = comp_info.get(comp_id, {})
        tmpl_id   = info.get('product_tmpl_id', [None])[0] if info else None
        sup       = sup_by_tmpl.get(tmpl_id, {})
        monthly   = comp_monthly[comp_id]['monthly']
        uom       = comp_monthly[comp_id]['uom']

        lead_days  = sup.get('delay', 0) or 0
        min_qty    = sup.get('min_qty', 0) or 0
        price      = sup.get('price', 0) or 0
        sup_name   = sup['name'][1] if sup.get('name') else 'NO SUPPLIER'
        sup_code   = sup.get('product_code', '') or ''

        stock      = stock_by_comp.get(comp_id, 0)
        incoming   = incoming_by_comp.get(comp_id, 0)
        available  = stock + incoming

        # Gross requirement for blanket period
        gross_6m   = monthly * BLANKET_MONTHS
        safety     = monthly * SAFETY_MONTHS

        # Total target = 6 months production + 1 month safety
        target     = gross_6m + safety

        # Net to order = target - what we already have
        net_to_order = max(0, target - available)

        # Round up to min qty
        if net_to_order > 0 and min_qty > 1:
            net_to_order = ceil(net_to_order / min_qty) * min_qty

        # Months of coverage from current available (before new order)
        coverage_months = (available / monthly) if monthly > 0 else 99

        # Delivery schedule — split order into monthly deliveries
        # First delivery must arrive by BLANKET_START
        # Order must be placed lead_time days before first delivery
        from datetime import timedelta
        order_date     = BLANKET_START - timedelta(days=lead_days)
        order_date_str = order_date.strftime('%Y-%m-%d')
        days_to_order  = (order_date - now.date()).days

        # Monthly delivery qty (spread evenly)
        delivery_qty   = round(net_to_order / BLANKET_MONTHS, 1) if net_to_order > 0 else 0

        # Flags
        no_supplier  = not sup
        zero_lead    = sup and lead_days == 0
        urgent       = days_to_order <= 0 and net_to_order > 0
        order_soon   = 0 < days_to_order <= 30 and net_to_order > 0
        no_order     = net_to_order == 0

        if no_supplier:        warning = "⚠ NO SUPPLIER"
        elif zero_lead:        warning = "⚠ CHECK LEAD TIME"
        elif urgent:           warning = "🔴 ORDER NOW"
        elif order_soon:       warning = f"🟠 Order by {order_date_str}"
        elif no_order:         warning = "✅ Sufficient stock"
        else:                  warning = f"🟡 Order by {order_date_str}"

        est_value = net_to_order * price if price else 0

        plan_rows.append({
            'comp_id':        comp_id,
            'ref':            info.get('default_code', '') or '',
            'name':           comp_monthly[comp_id]['name'],
            'uom':            uom,
            'supplier':       sup_name,
            'sup_code':       sup_code,
            'lead_days':      lead_days,
            'min_qty':        min_qty,
            'price':          price,
            'stock':          round(stock, 1),
            'incoming':       round(incoming, 1),
            'available':      round(available, 1),
            'monthly_req':    round(monthly, 2),
            'coverage_months':round(coverage_months, 1) if coverage_months < 99 else '∞',
            'gross_6m':       round(gross_6m, 1),
            'safety':         round(safety, 1),
            'target':         round(target, 1),
            'net_to_order':   round(net_to_order, 1),
            'delivery_qty':   round(delivery_qty, 1),
            'order_date':     order_date_str,
            'days_to_order':  days_to_order,
            'est_value':      round(est_value, 2),
            'warning':        warning,
            'urgent':         urgent,
            'order_soon':     order_soon,
            'no_order':       no_order,
            'no_supplier':    no_supplier,
            'zero_lead':      zero_lead,
        })

    # Sort: urgent first, then by days_to_order, then by name
    plan_rows.sort(key=lambda x: (
        0 if x['urgent'] else 1 if x['order_soon'] else 2 if not x['no_order'] else 3,
        x['days_to_order'] if not x['no_order'] else 999,
        x['ref']
    ))

    urgent_rows  = [r for r in plan_rows if r['urgent']]
    soon_rows    = [r for r in plan_rows if r['order_soon']]
    future_rows  = [r for r in plan_rows if not r['urgent'] and not r['order_soon'] and not r['no_order']]
    ok_rows      = [r for r in plan_rows if r['no_order']]

    print(f"  🔴 Order NOW:    {len(urgent_rows)}")
    print(f"  🟠 Order soon:   {len(soon_rows)}")
    print(f"  🟡 Order later:  {len(future_rows)}")
    print(f"  ✅ No order:     {len(ok_rows)}")

    total_value = sum(r['est_value'] for r in plan_rows if r['est_value'])
    print(f"  💰 Est. total:   ${total_value:,.0f}")

    # ================================================================
    # BUILD EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Column definitions
    BASE_COLS = [
        ("Ref",           10), ("Component",      38), ("Supplier",       22),
        ("Sup.\nCode",    12), ("Lead\n(days)",   10), ("Min\nQty",       9),
        ("Stock\nAvail",  10), ("Open\nIncoming", 10), ("Total\nAvail.",  10),
        ("Monthly\nReq",  10), ("Coverage\n(mo)", 10), ("6mo\nGross Req", 10),
        ("Safety\nStock", 10), ("Target\nStock",  10), ("NET TO\nORDER",  11),
        ("Monthly\nDelivery",11),("Order\nDate",  12), ("Days to\nOrder", 10),
        ("Est.\nValue $", 13), ("Status",         22),
    ]
    NCOLS = len(BASE_COLS)

    def setup_sheet(ws, title, subtitle=""):
        for i, (_, w) in enumerate(BASE_COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Title row
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        ws["A1"].font      = Font(bold=True, size=15, color="FFFFFF")
        ws["A1"].fill      = make_fill(COLORS['main'])
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        # Subtitle
        sub = (f"Blanket PO Plan: {MONTH_LABELS[0]} – {MONTH_LABELS[-1]}  |  "
               f"Safety Stock: {SAFETY_MONTHS} month  |  "
               f"Plan: C-100×150, C2 clearflo×40, C2 Bird×20, C2 Yellow×20 /mo  |  Period: Mar–Dec 2026 (10 months)"
               + (f"  |  {subtitle}" if subtitle else ""))
        ws.cell(row=2, column=1, value=sub)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
        ws["A2"].font      = Font(italic=True, size=9)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 14
        # Header row
        for i, (h, _) in enumerate(BASE_COLS, 1):
            hdr(ws, 3, i, h, COLORS['main'])
        ws.row_dimensions[3].height = 36
        ws.freeze_panes = "A4"

    def row_fill(r):
        if r['no_supplier'] or r['zero_lead']:  return make_fill(COLORS['yellow'])
        if r['urgent']:                          return make_fill(COLORS['red'])
        if r['order_soon']:                      return make_fill(COLORS['amber'])
        if r['no_order']:                        return make_fill(COLORS['green'])
        return make_fill(COLORS['blue'])

    def write_row(ws, r):
        fill = row_fill(r)
        cov  = r['coverage_months']
        vals = [
            r['ref'], r['name'][:36], r['supplier'][:20],
            r['sup_code'], r['lead_days'], r['min_qty'] if r['min_qty'] else '',
            r['stock'], r['incoming'], r['available'],
            r['monthly_req'] if r['monthly_req'] else '',
            cov if isinstance(cov, str) else cov,
            r['gross_6m'], r['safety'], r['target'],
            r['net_to_order'] if r['net_to_order'] else '',
            r['delivery_qty'] if r['delivery_qty'] else '',
            r['order_date'] if not r['no_order'] else '',
            r['days_to_order'] if not r['no_order'] else '',
            r['est_value'] if r['est_value'] else '',
            r['warning'],
        ]
        ws.append(vals)
        rn = ws.max_row
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col)
            c.fill      = fill
            c.border    = tb()
            c.alignment = Alignment(
                horizontal="left" if col in [1,2,3,4,20] else "center",
                vertical="center"
            )
            if col == 19 and r['est_value']:
                c.number_format = '$#,##0.00'
        ws.row_dimensions[rn].height = 14

    def section_title(ws, text, color):
        ws.append([text])
        rn = ws.max_row
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=NCOLS)
        c = ws.cell(row=rn, column=1)
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = make_fill(color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[rn].height = 16

    # ================================================================
    # SHEET 1 — FULL BLANKET PO PLAN (all components)
    # ================================================================
    ws1 = wb.create_sheet("📋 Blanket PO Plan")
    setup_sheet(ws1, f"Blanket PO Plan — {MONTH_LABELS[0]} to {MONTH_LABELS[-1]}  ({len(plan_rows)} components)")

    section_title(ws1, f"  🔴  ORDER NOW — {len(urgent_rows)} components (lead time means you must order immediately)", COLORS['urgent'])
    for r in urgent_rows:   write_row(ws1, r)

    section_title(ws1, f"  🟠  ORDER WITHIN 30 DAYS — {len(soon_rows)} components", COLORS['warn'])
    for r in soon_rows:     write_row(ws1, r)

    section_title(ws1, f"  🟡  ORDER BEFORE START — {len(future_rows)} components", "BF8F00")
    for r in future_rows:   write_row(ws1, r)

    section_title(ws1, f"  ✅  NO ORDER NEEDED — stock covers full horizon + safety ({len(ok_rows)} components)", COLORS['ok'])
    for r in ok_rows:       write_row(ws1, r)

    # ================================================================
    # SHEET 2 — BY SUPPLIER (purchasing view)
    # ================================================================
    ws2 = wb.create_sheet("🏭 By Supplier")
    setup_sheet(ws2, "Blanket PO Plan — Grouped by Supplier")

    by_supplier = defaultdict(list)
    for r in plan_rows:
        by_supplier[r['supplier']].append(r)

    # Sort suppliers: those with urgent items first
    def sup_priority(sup_name):
        rows = by_supplier[sup_name]
        if any(r['urgent'] for r in rows):       return 0
        if any(r['order_soon'] for r in rows):   return 1
        if any(not r['no_order'] for r in rows): return 2
        return 3

    for sup_name in sorted(by_supplier.keys(), key=sup_priority):
        rows     = by_supplier[sup_name]
        sup_val  = sum(r['est_value'] for r in rows if r['est_value'])
        urg_cnt  = sum(1 for r in rows if r['urgent'])
        ord_cnt  = sum(1 for r in rows if not r['no_order'])
        sup_color = COLORS['urgent'] if urg_cnt else COLORS['warn'] if ord_cnt else COLORS['ok']
        section_title(ws2,
            f"  🏭  {sup_name}  —  {len(rows)} parts  |  {ord_cnt} to order  |  Est. ${sup_val:,.0f}",
            sup_color)
        for r in sorted(rows, key=lambda x: (0 if x['urgent'] else 1 if x['order_soon'] else 2 if not x['no_order'] else 3, x['ref'])):
            write_row(ws2, r)

    # ================================================================
    # SHEET 3 — DELIVERY SCHEDULE
    # ================================================================
    ws3 = wb.create_sheet("📅 Delivery Schedule")
    # Cols: Ref, Component, Supplier, Lead days, Order Date, then one col per month
    sched_cols = ["Ref", "Component", "Supplier", "Lead\n(days)", "Place\nOrder By",
                  "Total\nOrder"] + MONTH_LABELS + ["Notes"]
    SCOLS = len(sched_cols)
    for i in range(1, SCOLS+1):
        ws3.column_dimensions[get_column_letter(i)].width = 14
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 38
    ws3.column_dimensions["C"].width = 22

    ws3.cell(row=1, column=1, value="Blanket PO — Monthly Delivery Schedule")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=SCOLS)
    ws3["A1"].font      = Font(bold=True, size=15, color="FFFFFF")
    ws3["A1"].fill      = make_fill(COLORS['main'])
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    ws3.cell(row=2, column=1, value=f"Delivery every month from {MONTH_LABELS[0]}. Order must be placed by 'Place Order By' date.")
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SCOLS)
    ws3["A2"].font = Font(italic=True, size=9)
    ws3["A2"].alignment = Alignment(horizontal="center")

    for i, h in enumerate(sched_cols, 1):
        hdr(ws3, 3, i, h, COLORS['main'])
    ws3.row_dimensions[3].height = 36
    ws3.freeze_panes = "A4"

    order_rows_only = [r for r in plan_rows if not r['no_order']]
    for r in order_rows_only:
        fill = row_fill(r)
        # Split total order evenly across months
        monthly_del = r['delivery_qty']
        notes = r['warning'] if r['no_supplier'] or r['zero_lead'] else ''
        vals = [
            r['ref'], r['name'][:36], r['supplier'][:20],
            r['lead_days'], r['order_date'], r['net_to_order']
        ] + [monthly_del] * BLANKET_MONTHS + [notes]
        ws3.append(vals)
        rn = ws3.max_row
        for col in range(1, SCOLS+1):
            c = ws3.cell(row=rn, column=col)
            c.fill      = fill
            c.border    = tb()
            c.alignment = Alignment(
                horizontal="left" if col in [1,2,3] else "center",
                vertical="center"
            )
        ws3.row_dimensions[rn].height = 14

    # ================================================================
    # SHEET 4 — EXECUTIVE SUMMARY
    # ================================================================
    ws4 = wb.create_sheet("📊 Summary")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 18

    ws4.cell(row=1, column=1, value="Blanket PO Plan — Executive Summary")
    ws4.merge_cells("A1:E1")
    ws4["A1"].font      = Font(bold=True, size=15, color="FFFFFF")
    ws4["A1"].fill      = make_fill(COLORS['main'])
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26
    ws4.append([])

    def sum_row(ws, label, val, fill_hex, bold=False):
        ws.append([label, val])
        rn = ws.max_row
        ws.cell(rn, 1).fill = make_fill(fill_hex)
        ws.cell(rn, 1).font = Font(bold=bold)
        ws.cell(rn, 1).border = tb()
        ws.cell(rn, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(rn, 2).fill = make_fill(fill_hex)
        ws.cell(rn, 2).font = Font(bold=bold)
        ws.cell(rn, 2).border = tb()
        ws.cell(rn, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[rn].height = 16

    sum_row(ws4, "Planning period", f"{MONTH_LABELS[0]} – {MONTH_LABELS[-1]} (10 months, end of year)", "DDEBF7", True)
    sum_row(ws4, "Total components in scope", len(plan_rows), "DDEBF7")
    sum_row(ws4, "🔴 Components — ORDER NOW", len(urgent_rows), COLORS['red'])
    sum_row(ws4, "🟠 Components — Order within 30 days", len(soon_rows), COLORS['amber'])
    sum_row(ws4, "🟡 Components — Order before Mar 1 (plan ahead)", len(future_rows), COLORS['yellow'])
    sum_row(ws4, "✅ Components — No order needed", len(ok_rows), COLORS['green'])
    sum_row(ws4, "⚠ Components — No supplier configured", sum(1 for r in plan_rows if r['no_supplier']), COLORS['yellow'])
    sum_row(ws4, "⚠ Components — Lead time = 0 (check)", sum(1 for r in plan_rows if r['zero_lead'] and not r['no_supplier']), COLORS['yellow'])
    ws4.append([])
    sum_row(ws4, "💰 Estimated total blanket PO value", f"${total_value:,.0f}", "DDEBF7", True)
    ws4.append([])

    # By supplier summary
    hdr(ws4, ws4.max_row+1, 1, "Supplier", COLORS['main'])
    hdr(ws4, ws4.max_row,   2, "# Parts", COLORS['main'])
    hdr(ws4, ws4.max_row,   3, "# To Order", COLORS['main'])
    hdr(ws4, ws4.max_row,   4, "🔴 Urgent", COLORS['main'])
    hdr(ws4, ws4.max_row,   5, "Est. Value", COLORS['main'])
    ws4.row_dimensions[ws4.max_row].height = 20

    for sup_name in sorted(by_supplier.keys(), key=sup_priority):
        rows    = by_supplier[sup_name]
        sup_val = sum(r['est_value'] for r in rows if r['est_value'])
        urg     = sum(1 for r in rows if r['urgent'])
        ord_n   = sum(1 for r in rows if not r['no_order'])
        fill_h  = COLORS['red'] if urg else COLORS['amber'] if ord_n else COLORS['green']
        ws4.append([sup_name[:32], len(rows), ord_n, urg, f"${sup_val:,.0f}" if sup_val else "N/A"])
        rn = ws4.max_row
        for col in range(1, 6):
            c = ws4.cell(rn, col)
            c.fill      = make_fill(fill_h)
            c.border    = tb()
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
        ws4.row_dimensions[rn].height = 14

    # ================================================================
    # SHEET 5 — BUILDABLE UNITS
    # ================================================================
    ws5 = wb.create_sheet("🔧 Buildable Units")

    # For each finished product, calculate max buildable from current stock only
    # and from stock + incoming
    buildable_results = {}
    for ref in PLAN_PRODUCTS:
        if ref not in bom_by_product:
            continue
        bom = bom_by_product[ref]
        if not bom:
            continue

        # Find finished product info
        fp = prod_by_ref.get(ref, {})
        fp_name = fp.get('name', ref) if fp else ref

        # Calculate buildable units per component
        limiting_stock    = None  # worst case component (stock only)
        limiting_total    = None  # worst case component (stock + incoming)
        comp_details      = []

        for comp_id, qty_per_unit in bom.items():
            if qty_per_unit <= 0:
                continue
            c_info    = comp_info.get(comp_id, {})
            c_name    = comp_monthly.get(comp_id, {}).get('name', '') or c_info.get('name', '')
            c_ref     = c_info.get('default_code', '') or ''
            c_stock   = stock_by_comp.get(comp_id, 0)
            c_incom   = incoming_by_comp.get(comp_id, 0)
            c_total   = c_stock + c_incom

            buildable_from_stock = int(c_stock / qty_per_unit)
            buildable_from_total = int(c_total / qty_per_unit)

            comp_details.append({
                'comp_id':   comp_id,
                'ref':       c_ref,
                'name':      c_name,
                'qty_per_unit': qty_per_unit,
                'stock':     c_stock,
                'incoming':  c_incom,
                'total':     c_total,
                'build_stock': buildable_from_stock,
                'build_total': buildable_from_total,
            })

            if limiting_stock is None or buildable_from_stock < limiting_stock['build_stock']:
                limiting_stock = comp_details[-1]
            if limiting_total is None or buildable_from_total < limiting_total['build_total']:
                limiting_total = comp_details[-1]

        max_build_stock = limiting_stock['build_stock'] if limiting_stock else 0
        max_build_total = limiting_total['build_total'] if limiting_total else 0
        monthly_plan_qty = MONTHLY_PLAN.get(ref, 0)
        months_stock = round(max_build_stock / monthly_plan_qty, 1) if monthly_plan_qty else '∞'
        months_total = round(max_build_total / monthly_plan_qty, 1) if monthly_plan_qty else '∞'

        buildable_results[ref] = {
            'ref':            ref,
            'name':           fp_name,
            'monthly_plan':   monthly_plan_qty,
            'max_build_stock': max_build_stock,
            'max_build_total': max_build_total,
            'months_stock':   months_stock,
            'months_total':   months_total,
            'limiting_stock': limiting_stock,
            'limiting_total': limiting_total,
            'comp_details':   sorted(comp_details, key=lambda x: x['build_stock']),
        }

    # --- Build the sheet ---
    ws5.column_dimensions["A"].width = 12
    ws5.column_dimensions["B"].width = 42
    ws5.column_dimensions["C"].width = 12
    ws5.column_dimensions["D"].width = 14
    ws5.column_dimensions["E"].width = 14
    ws5.column_dimensions["F"].width = 14
    ws5.column_dimensions["G"].width = 14
    ws5.column_dimensions["H"].width = 40
    ws5.column_dimensions["I"].width = 40

    ws5.cell(row=1, column=1, value="Buildable Units Analysis — Current Stock & Incoming")
    ws5.merge_cells("A1:I1")
    ws5["A1"].font      = Font(bold=True, size=15, color="FFFFFF")
    ws5["A1"].fill      = make_fill(COLORS['main'])
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 26

    ws5.cell(row=2, column=1,
             value=f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC  |  "
                   f"Stock Only = physical stock available now  |  "
                   f"Stock + Incoming = includes all open PO receipts expected")
    ws5.merge_cells("A2:I2")
    ws5["A2"].font      = Font(italic=True, size=9)
    ws5["A2"].alignment = Alignment(horizontal="center")
    ws5.row_dimensions[2].height = 14

    # Summary table header
    sum_hdrs = ["Ref", "Finished Product", "Monthly Plan",
                "Max Build (Stock Only)", "Months Coverage",
                "Max Build (+ Incoming)", "Months Coverage",
                "Limiting Part (Stock Only)", "Limiting Part (+ Incoming)"]
    for i, h in enumerate(sum_hdrs, 1):
        hdr(ws5, 3, i, h, COLORS['main'])
    ws5.row_dimensions[3].height = 36
    ws5.freeze_panes = "A4"

    for ref, res in sorted(buildable_results.items()):
        ms = res['months_stock']
        mt = res['months_total']
        # Color by coverage
        if isinstance(ms, (int, float)) and ms < 1:   fill_s = make_fill(COLORS['red'])
        elif isinstance(ms, (int, float)) and ms < 2: fill_s = make_fill(COLORS['amber'])
        else:                                          fill_s = make_fill(COLORS['green'])
        if isinstance(mt, (int, float)) and mt < 1:   fill_t = make_fill(COLORS['red'])
        elif isinstance(mt, (int, float)) and mt < 2: fill_t = make_fill(COLORS['amber'])
        else:                                          fill_t = make_fill(COLORS['green'])

        lim_s = res['limiting_stock']
        lim_t = res['limiting_total']
        lim_s_str = f"{lim_s['ref']} — {lim_s['name'][:30]} ({lim_s['stock']:.0f} stock / {lim_s['qty_per_unit']:.2f}/unit)" if lim_s else ''
        lim_t_str = f"{lim_t['ref']} — {lim_t['name'][:30]} ({lim_t['total']:.0f} avail / {lim_t['qty_per_unit']:.2f}/unit)" if lim_t else ''

        ws5.append([
            ref, res['name'][:40], res['monthly_plan'],
            res['max_build_stock'], f"{ms} mo",
            res['max_build_total'], f"{mt} mo",
            lim_s_str, lim_t_str
        ])
        rn = ws5.max_row
        for col in range(1, 10):
            c = ws5.cell(rn, col)
            c.fill   = fill_s if col in [4, 5, 8] else fill_t if col in [6, 7, 9] else make_fill(COLORS['blue'])
            c.border = tb()
            c.alignment = Alignment(
                horizontal="left" if col in [2, 8, 9] else "center",
                vertical="center", wrap_text=col in [8, 9]
            )
        ws5.row_dimensions[rn].height = 28

    ws5.append([])

    # --- Detail section: bottom 10 limiting components per product ---
    for ref, res in sorted(buildable_results.items()):
        # Section header
        fp_name = res['name'][:40]
        monthly = res['monthly_plan']
        ws5.append([f"  {ref} — {fp_name}  |  Monthly plan: {monthly}  |  "
                    f"Max buildable (stock): {res['max_build_stock']}  |  "
                    f"Max buildable (+ incoming): {res['max_build_total']}"])
        rn = ws5.max_row
        ws5.merge_cells(f"A{rn}:I{rn}")
        ws5.cell(rn, 1).font      = Font(bold=True, color="FFFFFF", size=11)
        ws5.cell(rn, 1).fill      = make_fill(COLORS['main'])
        ws5.cell(rn, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws5.row_dimensions[rn].height = 16

        # Sub-header
        det_hdrs = ["Ref", "Component", "Qty/Unit", "Stock", "Incoming", "Total Avail", "Build (Stock)", "Build (Total)", ""]
        for i, h in enumerate(det_hdrs, 1):
            hdr(ws5, ws5.max_row+1, i, h, COLORS['sup'], size=9)
        ws5.row_dimensions[ws5.max_row].height = 18

        # Show bottom 15 (most limiting) components
        limiting_comps = res['comp_details'][:15]
        for cd in limiting_comps:
            bs = cd['build_stock']
            bt = cd['build_total']
            if bs == 0:       rf = make_fill(COLORS['red'])
            elif bs < monthly: rf = make_fill(COLORS['amber'])
            else:              rf = make_fill(COLORS['green'])
            ws5.append([
                cd['ref'], cd['name'][:40],
                round(cd['qty_per_unit'], 4),
                round(cd['stock'], 1), round(cd['incoming'], 1), round(cd['total'], 1),
                bs, bt, ''
            ])
            rn = ws5.max_row
            for col in range(1, 9):
                c = ws5.cell(rn, col)
                c.fill   = rf
                c.border = tb()
                c.alignment = Alignment(
                    horizontal="left" if col in [1, 2] else "center",
                    vertical="center"
                )
            ws5.row_dimensions[rn].height = 14
        ws5.append([])

    # Move Summary sheet to front
    wb.move_sheet("📊 Summary", offset=-wb.index(wb["📊 Summary"]))

    output_file = f"blanket_po_plan_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\n✅ Excel saved: {output_file}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
