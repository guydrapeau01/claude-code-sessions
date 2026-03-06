import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

TOP_LEVEL_REF = "101336"

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    print("Fetching internal locations...")
    internal_locations = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.location', 'search_read',
        [[['usage', '=', 'internal'], ['active', '=', True]]],
        {'fields': ['id']}
    )
    internal_location_ids = {loc['id'] for loc in internal_locations}
    print(f"Found {len(internal_location_ids)} internal locations.\n")

    def explode_bom(product_ref, level=0, parent_ref=None, visited=None):
        if visited is None:
            visited = set()
        product = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
            [[['default_code', '=', product_ref]]],
            {'fields': ['id', 'name', 'default_code', 'product_tmpl_id'], 'limit': 1}
        )
        if not product:
            print(f"  {'  '*level}⚠ Part {product_ref} not found")
            return []
        product_id   = product[0]['id']
        tmpl_id      = product[0]['product_tmpl_id'][0]
        product_name = product[0]['name']
        boms = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'search_read',
            [[['product_tmpl_id', '=', tmpl_id], ['type', 'in', ['normal', 'phantom']]]],
            {'fields': ['id'], 'limit': 1}
        )
        components = []
        if not boms:
            if product_ref not in visited:
                visited.add(product_ref)
                print(f"  {'  '*level}→ [{product_ref}] {product_name}  (raw part)")
                components.append({'level': level, 'product_id': product_id, 'default_code': product_ref, 'name': product_name, 'parent_ref': parent_ref, 'has_bom': False})
            return components
        bom_id = boms[0]['id']
        if product_ref not in visited:
            visited.add(product_ref)
            label = "TOP LEVEL" if level == 0 else "sub-assembly"
            print(f"  {'  '*level}▶ [{product_ref}] {product_name}  ({label})")
            if level > 0:
                components.append({'level': level, 'product_id': product_id, 'default_code': product_ref, 'name': product_name, 'parent_ref': parent_ref, 'has_bom': True})
        bom_lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom.line', 'search_read',
            [[['bom_id', '=', bom_id]]],
            {'fields': ['product_id', 'product_qty']}
        )
        for line in bom_lines:
            child_id   = line['product_id'][0]
            child_name = line['product_id'][1]
            child_data = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'read',
                [[child_id]], {'fields': ['default_code']}
            )
            child_ref = child_data[0].get('default_code', '') if child_data else ''
            if not child_ref:
                print(f"  {'  '*(level+1)}⚠ No internal ref: {child_name}")
                continue
            if child_ref in visited:
                print(f"  {'  '*(level+1)}↩ [{child_ref}] already processed")
                continue
            components.extend(explode_bom(child_ref, level + 1, product_ref, visited))
        return components

    print(f"Exploding BOM for: {TOP_LEVEL_REF}\n")
    all_components = explode_bom(TOP_LEVEL_REF)
    if not all_components:
        print(f"No BOM components found for {TOP_LEVEL_REF}")
        input("\nPress Enter to close...")
        exit()
    print(f"\nTotal unique components: {len(all_components)}\n")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "BOM Stock Report"

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    subassy_fill = PatternFill("solid", fgColor="D9E1F2")
    subassy_font = Font(bold=True)
    alt_fill     = PatternFill("solid", fgColor="EBF3FB")
    no_stock_fill = PatternFill("solid", fgColor="FFE0E0")
    no_stock_font = Font(color="CC0000", italic=True)
    center = Alignment(horizontal="center")

    ws_out.append([f"BOM Stock Report - Top Level: {TOP_LEVEL_REF}"])
    ws_out["A1"].font = Font(bold=True, size=14)
    ws_out.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_out.append([])

    headers = ["Level", "Internal Ref", "Product Name", "Parent Assy", "Lot / Serial", "Location", "On-Hand Qty", "Reserved", "Available"]
    ws_out.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws_out.cell(row=4, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    row_num = 5
    total_skipped = 0

    for idx, comp in enumerate(all_components):
        part_ref     = comp['default_code']
        product_id   = comp['product_id']
        product_name = comp['name']
        level        = comp['level']
        parent_ref   = comp['parent_ref'] or TOP_LEVEL_REF
        has_bom      = comp['has_bom']
        indent       = "  " * level
        print(f"Querying stock for: {part_ref}...")
        quants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.quant', 'search_read',
            [[['product_id', '=', product_id]]],
            {'fields': ['lot_id', 'location_id', 'quantity', 'reserved_quantity']}
        )
        rows_added = 0
        for q in quants:
            loc_id = q['location_id'][0] if q['location_id'] else None
            qty    = q['quantity']
            if loc_id not in internal_location_ids or qty <= 0:
                total_skipped += 1
                continue
            lot_name      = q['lot_id'][1] if q['lot_id'] else 'No Lot'
            location_name = q['location_id'][1]
            available     = qty - q['reserved_quantity']
            ws_out.append([indent + str(level), part_ref, product_name, parent_ref, lot_name, location_name, qty, q['reserved_quantity'], available])
            if has_bom:
                for col in range(1, 10):
                    ws_out.cell(row=row_num, column=col).fill = subassy_fill
                    ws_out.cell(row=row_num, column=col).font = subassy_font
            elif idx % 2 == 0:
                for col in range(1, 10):
                    ws_out.cell(row=row_num, column=col).fill = alt_fill
            row_num += 1
            rows_added += 1
        if rows_added == 0:
            ws_out.append([indent + str(level), part_ref, product_name, parent_ref, "⚠ No stock", "", 0, 0, 0])
            for col in range(1, 10):
                ws_out.cell(row=row_num, column=col).fill = no_stock_fill
                ws_out.cell(row=row_num, column=col).font = no_stock_font
            row_num += 1

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in ws_out.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                except: pass
        ws_out.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    output_file = f"bom_stock_{TOP_LEVEL_REF}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb_out.save(output_file)
    print(f"\nExcel report saved: {output_file}")
    print(f"Rows skipped (virtual/zero/negative): {total_skipped}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
