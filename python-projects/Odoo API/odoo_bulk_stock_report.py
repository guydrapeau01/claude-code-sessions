import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

INPUT_FILE = "part_list.xlsx"
SKIP_HEADER = True

try:
    print(f"Reading part numbers from {INPUT_FILE}...")
    wb_in = openpyxl.load_workbook(INPUT_FILE)
    ws_in = wb_in.active
    part_numbers = []
    for i, row in enumerate(ws_in.iter_rows(min_col=1, max_col=1, values_only=True)):
        if i == 0 and SKIP_HEADER:
            continue
        if row[0]:
            part_numbers.append(str(row[0]).strip())
    print(f"Found {len(part_numbers)} part numbers: {part_numbers}\n")
    if not part_numbers:
        print("No part numbers found.")
        input("\nPress Enter to close...")
        exit()

    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    internal_locations = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.location', 'search_read',
        [[['usage', '=', 'internal'], ['active', '=', True]]],
        {'fields': ['id']}
    )
    internal_location_ids = {loc['id'] for loc in internal_locations}

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Stock Report"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    not_found_font = Font(color="FF0000", italic=True)
    center = Alignment(horizontal="center")

    ws_out.append([f"Stock Report by Part Number"])
    ws_out["A1"].font = Font(bold=True, size=14)
    ws_out.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws_out.append([])

    headers = ["Internal Ref", "Product Name", "Lot / Serial", "Location", "Location Type", "On-Hand Qty", "Reserved", "Available"]
    ws_out.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws_out.cell(row=4, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    row_num = 5
    total_skipped = 0

    for idx, part_ref in enumerate(part_numbers):
        print(f"Querying part: {part_ref}...")
        product = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
            [[['default_code', '=', part_ref]]],
            {'fields': ['id', 'name', 'default_code'], 'limit': 1}
        )
        if not product:
            ws_out.append([part_ref, "NOT FOUND IN ODOO", "", "", "", "", "", ""])
            for col in range(1, 9):
                ws_out.cell(row=row_num, column=col).font = not_found_font
            row_num += 1
            continue

        product_id   = product[0]['id']
        product_name = product[0]['name']

        quants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.quant', 'search_read',
            [[['product_id', '=', product_id]]],
            {'fields': ['lot_id', 'location_id', 'quantity', 'reserved_quantity']}
        )

        fill = alt_fill if idx % 2 == 0 else None
        rows_added = 0
        for q in quants:
            loc_id = q['location_id'][0] if q['location_id'] else None
            qty    = q['quantity']
            if loc_id not in internal_location_ids or qty <= 0:
                total_skipped += 1
                continue
            lot_name      = q['lot_id'][1] if q['lot_id'] else 'No Lot'
            location_name = q['location_id'][1]
            available     = qty - q['reserved_quantity']
            ws_out.append([part_ref, product_name, lot_name, location_name, "Internal", qty, q['reserved_quantity'], available])
            if fill:
                for col in range(1, 9):
                    ws_out.cell(row=row_num, column=col).fill = fill
            row_num += 1
            rows_added += 1

        if rows_added == 0:
            ws_out.append([part_ref, product_name, "No stock in internal locations", "", "", 0, 0, 0])
            row_num += 1

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in ws_out.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except: pass
        ws_out.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    output_file = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb_out.save(output_file)
    print(f"\nExcel report saved: {output_file}")
    print(f"Rows skipped (virtual/zero/negative): {total_skipped}")
    print("=== Done ===")

except FileNotFoundError:
    print(f"\nERROR: Could not find '{INPUT_FILE}'. Make sure it is in the same folder.")
except Exception as e:
    print(f"\nERROR: {e}")

input("\nPress Enter to close...")
