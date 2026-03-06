import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timezone

def make_fill(hex_): return PatternFill("solid", fgColor=hex_)
def tb():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

try:
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")

    TEST_LOAD_REF = '102069'
    SO_REFS       = ['S04521', 'S04416']
    now           = datetime.now(timezone.utc)

    prod = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', '=', TEST_LOAD_REF]]], {'fields': ['id', 'name']}
    )
    pid       = prod[0]['id']
    prod_name = prod[0]['name']

    rows = []
    for so_ref in SO_REFS:
        so = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'sale.order', 'search_read',
            [[['name', '=', so_ref]]], {'fields': ['id', 'name', 'partner_id', 'picking_ids']}
        )
        if not so:
            print(f"  {so_ref}: NOT FOUND")
            continue

        s        = so[0]
        customer = s['partner_id'][1] if s.get('partner_id') else ''
        pick_ids = s['picking_ids']
        picks    = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.picking', 'read',
            [pick_ids], {'fields': ['name', 'state', 'date_done']}
        )
        pick_map = {p['id']: p for p in picks}

        mls = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.move.line', 'search_read',
            [[['picking_id', 'in', pick_ids], ['product_id', '=', pid]]],
            {'fields': ['lot_id', 'qty_done', 'picking_id']}
        )
        lot_ids = [ml['lot_id'][0] for ml in mls if ml.get('lot_id')]
        lot_ml  = {ml['lot_id'][0]: ml for ml in mls if ml.get('lot_id')}

        if lot_ids:
            lots = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.production.lot', 'read',
                [lot_ids],
                {'fields': ['name', 'cal_id', 'expiration_date']}
            )
            for lot in lots:
                ml       = lot_ml.get(lot['id'], {})
                pick_id  = ml['picking_id'][0] if ml.get('picking_id') else None
                pick     = pick_map.get(pick_id, {})
                rows.append({
                    'so':       so_ref,
                    'customer': customer,
                    'delivery': pick.get('name', ''),
                    'state':    pick.get('state', ''),
                    'date':     str(pick.get('date_done', '') or '')[:10],
                    'serial':   lot['name'],
                    'cal_id':   lot.get('cal_id', '') or 'N/A',
                    'expiry':   str(lot.get('expiration_date', '') or '')[:10],
                })

    print(f"Found {len(rows)} test load records\n")

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Load CAL IDs"

    MAIN = "2E4057"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 14

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Test Load CAL IDs — {', '.join(SO_REFS)}"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = make_fill(MAIN)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Product: {prod_name}  |  Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC"
    ws["A2"].font      = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 13

    # Headers
    headers = ["Sales Order", "Customer", "Delivery", "State", "Ship Date", "Serial #", "CAL ID", "Expiry Date"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = make_fill(MAIN)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = tb()
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    # Data rows
    for r in rows:
        fill = make_fill("DDEBF7")
        ws.append([r['so'], r['customer'], r['delivery'], r['state'],
                   r['date'], r['serial'], r['cal_id'], r['expiry']])
        rn = ws.max_row
        for col in range(1, 9):
            c = ws.cell(rn, col)
            c.fill      = fill
            c.border    = tb()
            c.alignment = Alignment(
                horizontal="left" if col in [2, 3] else "center",
                vertical="center"
            )
            # Highlight CAL ID column
            if col == 7:
                c.font = Font(bold=True)
                c.fill = make_fill("E2EFDA")
        ws.row_dimensions[rn].height = 15

    output = f"test_load_cal_ids_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output)
    print(f"✅ Saved: {output}")

except Exception as e:
    import traceback
    traceback.print_exc()

input("\nPress Enter to close...")
