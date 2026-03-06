import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

# Divisions and device lines loaded from config

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now = datetime.now(timezone.utc)

    # Resolve C2 product IDs
    c2_products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', cfg.C2_PRODUCT_REFS]]],
        {'fields': ['id']}
    )
    c2_product_ids = {p['id'] for p in c2_products}

    all_teams = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.team', 'search_read',
        [[]], {'fields': ['id', 'name']}
    )
    team_map = {t['name']: t['id'] for t in all_teams}

    all_stages = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.stage', 'search_read',
        [[]], {'fields': ['id', 'name', 'sequence']}
    )
    stage_order = {s['name']: s['sequence'] for s in all_stages}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    division_colors = {
        "Support - Americas": "1F4E79",
        "Support - EMEA":     "375623",
        "Support - APAC":     "7B2C2C",
    }

    stage_fills = {
        "Initial Contact": make_fill("FFF2CC"),
        "In Progress":     make_fill("DDEBF7"),
        "Solved":          make_fill("E2EFDA"),
        "Cancelled":       make_fill("F2F2F2"),
    }

    header_font = Font(color="FFFFFF", bold=True, size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    priority_map = {'0': 'Normal', '1': 'Low', '2': 'High', '3': 'Urgent'}

    FIELDS = [
        'id', 'name', 'stage_id', 'team_id', 'partner_id', 'partner_email',
        'ticket_type_id', 'x_studio_customer_type', 'tag_ids_char',
        'product_id', 'lot_id', 'x_studio_age_of_device',
        'x_studio_other_related_products', 'create_date', 'create_uid',
        'user_id', 'division_id', 'priority', 'x_studio_under_warranty',
        'date_last_stage_update', 'product_id'
    ]

    # ---- SUMMARY SHEET ----
    ws_sum = wb.create_sheet("Summary")
    ws_sum.freeze_panes = "A5"

    ws_sum.merge_cells("A1:H1")
    ws_sum["A1"] = "Helpdesk Ticket Report — All Divisions"
    ws_sum["A1"].font = Font(bold=True, size=16)
    ws_sum["A1"].alignment = center
    ws_sum.merge_cells("A2:H2")
    ws_sum["A2"] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC"
    ws_sum["A2"].alignment = center
    ws_sum["A2"].font = Font(italic=True)
    ws_sum.append([])

    sum_headers = ["Division", "Stage", "# Tickets", "Avg Days Open", "Min Days Open", "Max Days Open", "Oldest Ticket", "# Unassigned"]
    ws_sum.append(sum_headers)
    for col, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=4, column=col)
        c.font = header_font
        c.fill = make_fill("2E4057")
        c.alignment = center
        c.border = thin_border()

    sum_row = 5

    # ---- ONE SHEET PER DIVISION ----
    for division in cfg.DIVISIONS:
        team_id = team_map.get(division)
        if not team_id:
            print(f"  ⚠ Team not found: {division}")
            continue

        print(f"Fetching tickets for: {division}...")
        tickets = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.ticket', 'search_read',
            [[['team_id', '=', team_id]]],
            {'fields': FIELDS}
        )
        print(f"  → {len(tickets)} tickets found")

        color      = division_colors.get(division, "2E4057")
        short_name = division.replace("Support - ", "")
        ws = wb.create_sheet(short_name)
        ws.freeze_panes = "A5"

        # Title
        ws.merge_cells("A1:U1")
        ws["A1"] = f"{division} — Helpdesk Tickets"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = make_fill(color)
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 22

        ws.merge_cells("A2:U2")
        ws["A2"] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   Total Tickets: {len(tickets)}"
        ws["A2"].alignment = center
        ws["A2"].font = Font(italic=True)
        ws.append([])

        col_headers = [
            "Ticket #", "Subject", "Stage", "Days Open",
            "Customer", "Customer Email", "Ticket Type", "Customer Type",
            "Tags", "Product", "Device Line", "Lot/Serial", "Age of Device",
            "Created On", "Created By", "Assigned To", "Division",
            "Priority", "Under Warranty", "Other Related Products", "Last Stage Update"
        ]
        ws.append(col_headers)
        for col, h in enumerate(col_headers, 1):
            c = ws.cell(row=4, column=col)
            c.font = header_font
            c.fill = make_fill(color)
            c.alignment = center
            c.border = thin_border()

        col_widths = [10, 35, 16, 11, 25, 28, 18, 16, 25, 35, 12, 15, 13, 13, 20, 20, 18, 10, 13, 35, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Group by stage
        by_stage = {}
        for t in tickets:
            sname = t['stage_id'][1] if t['stage_id'] else 'Unknown'
            by_stage.setdefault(sname, []).append(t)

        row_num = 5

        for stage_name in sorted(by_stage.keys(), key=lambda s: stage_order.get(s, 99)):
            stage_tickets = by_stage[stage_name]
            sfill = stage_fills.get(stage_name, make_fill("FFFFFF"))

            # Stage group header
            ws.merge_cells(f"A{row_num}:U{row_num}")
            ws.cell(row=row_num, column=1).value = f"  ▶  {stage_name}  ({len(stage_tickets)} tickets)"
            ws.cell(row=row_num, column=1).font  = Font(bold=True, size=11, color="FFFFFF")
            ws.cell(row=row_num, column=1).fill  = make_fill(color)
            ws.cell(row=row_num, column=1).alignment = left
            ws.row_dimensions[row_num].height = 18
            row_num += 1

            days_list  = []
            unassigned = 0

            for t in sorted(stage_tickets, key=lambda x: x.get('create_date') or ''):
                created_str = t.get('create_date', '')
                if created_str:
                    created_dt = datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
                    days_open  = (now - created_dt).days
                    created_fmt = created_dt.strftime('%Y-%m-%d')
                else:
                    days_open   = 0
                    created_fmt = ''
                days_list.append(days_open)

                last_update = t.get('date_last_stage_update', '')
                last_update_fmt = datetime.strptime(last_update, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d') if last_update else ''

                assigned = t['user_id'][1] if t['user_id'] else 'Unassigned'
                if assigned == 'Unassigned':
                    unassigned += 1

                # Other related products
                other_prods = ''
                if t.get('x_studio_other_related_products'):
                    prod_ids = [p[0] if isinstance(p, list) else p for p in t['x_studio_other_related_products']]
                    if prod_ids:
                        prods = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'read',
                            [prod_ids], {'fields': ['default_code', 'name']}
                        )
                        other_prods = ', '.join([f"[{p.get('default_code','')}] {p['name']}" for p in prods])

                row_data = [
                    f"#{t['id']}",
                    t.get('name', ''),
                    stage_name,
                    days_open,
                    t['partner_id'][1]      if t['partner_id']      else '',
                    t.get('partner_email', '')                       or '',
                    t['ticket_type_id'][1]  if t['ticket_type_id']  else '',
                    t.get('x_studio_customer_type', '')              or '',
                    t.get('tag_ids_char', '')                        or '',
                    t['product_id'][1]      if t['product_id']      else '',
                    'C2' if (t['product_id'] and t['product_id'][0] in c2_product_ids) else 'C-100',
                    t['lot_id'][1]          if t['lot_id']          else '',
                    t.get('x_studio_age_of_device', '')              or '',
                    created_fmt,
                    t['create_uid'][1]      if t['create_uid']      else '',
                    assigned,
                    t['division_id'][1]     if t['division_id']     else '',
                    priority_map.get(str(t.get('priority', '0')), 'Normal'),
                    'Yes' if t.get('x_studio_under_warranty') else 'No',
                    other_prods,
                    last_update_fmt,
                ]
                ws.append(row_data)

                # Highlight aging open tickets
                row_fill = sfill
                if stage_name not in ('Solved', 'Cancelled'):
                    if days_open > 60:
                        row_fill = make_fill("FFB3B3")
                    elif days_open > 30:
                        row_fill = make_fill("FFD9B3")

                for col in range(1, len(col_headers) + 1):
                    c = ws.cell(row=row_num, column=col)
                    c.fill   = row_fill
                    c.border = thin_border()
                    c.alignment = left
                ws.row_dimensions[row_num].height = 15
                row_num += 1

            # Summary stats for this stage
            if days_list:
                avg_d = round(sum(days_list) / len(days_list), 1)
                min_d = min(days_list)
                max_d = max(days_list)
                oldest = min((t.get('create_date','') for t in stage_tickets if t.get('create_date')), default='')
                oldest_fmt = datetime.strptime(oldest, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d') if oldest else ''
            else:
                avg_d = min_d = max_d = 0
                oldest_fmt = ''

            ws_sum.append([division, stage_name, len(stage_tickets), avg_d, min_d, max_d, oldest_fmt, unassigned])
            for col in range(1, 9):
                c = ws_sum.cell(row=sum_row, column=col)
                c.fill      = stage_fills.get(stage_name, make_fill("FFFFFF"))
                c.border    = thin_border()
                c.alignment = center
            sum_row += 1

            ws.append([])
            row_num += 1

        ws_sum.append([])
        sum_row += 1

    sum_widths = [22, 18, 12, 15, 15, 15, 20, 14]
    for i, w in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    output_file = f"helpdesk_report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\nExcel report saved: {output_file}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
