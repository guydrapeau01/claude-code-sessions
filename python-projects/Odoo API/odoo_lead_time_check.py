import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

def make_fill(hex):    return PatternFill("solid", fgColor=hex)
def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now = datetime.now(timezone.utc)

    # --- Resolve BOM components for all planning products ---
    PLAN_PRODUCTS = ['101336', '101769', '101711', '102237',
                     '101490', '101759', '101760', '102240']

    print(f"Resolving BOM components for: {PLAN_PRODUCTS}")
    fin_products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', PLAN_PRODUCTS]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    print(f"  Found {len(fin_products)} finished products")

    # Find BOMs by template default_code
    bom_ids = []
    for p in fin_products:
        boms = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'search_read',
            [[['product_tmpl_id.default_code', '=', p['default_code']]]],
            {'fields': ['id']}
        )
        if boms:
            bom_ids.append(boms[0]['id'])
            print(f"  BOM found for {p['default_code']}: ID {boms[0]['id']}")
        else:
            print(f"  ⚠ No BOM for {p['default_code']}")

    # Explode all BOMs to get component product IDs
    def get_bom_component_ids(models, uid, bom_id, visited=None):
        if visited is None:
            visited = set()
        if bom_id in visited:
            return set()
        visited.add(bom_id)
        lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom.line', 'search_read',
            [[['bom_id', '=', bom_id]]],
            {'fields': ['product_id', 'child_bom_id']}
        )
        comp_ids = set()
        for l in lines:
            comp_ids.add(l['product_id'][0])
            if l.get('child_bom_id'):
                comp_ids |= get_bom_component_ids(models, uid, l['child_bom_id'][0], visited)
        return comp_ids

    all_comp_ids = set()
    for bom_id in bom_ids:
        ids = get_bom_component_ids(models, uid, bom_id)
        all_comp_ids |= ids
    # Also include the finished products themselves (they may be purchased)
    all_comp_ids |= {p['id'] for p in fin_products}
    print(f"  Total unique components to check: {len(all_comp_ids)}\n")

    # --- Get all supplier info (configured lead times) ---
    print("Fetching supplier lead times...")
    sup_info = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.supplierinfo', 'search_read',
        [[]], {'fields': ['product_tmpl_id', 'product_id', 'name', 'delay',
                          'min_qty', 'price', 'product_code']}
    )
    # Keep best (first) supplier per template
    sup_by_tmpl = {}
    for s in sup_info:
        tmpl_id = s['product_tmpl_id'][0]
        if tmpl_id not in sup_by_tmpl:
            sup_by_tmpl[tmpl_id] = s

    # --- Get all done PO lines with receipt date ---
    print("Fetching completed PO receipts...")
    # Get done stock pickings (receipts)
    pickings = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.picking', 'search_read',
        [[['picking_type_id.code', '=', 'incoming'],
          ['state', '=', 'done']]],
        {'fields': ['id', 'name', 'date_done', 'origin', 'purchase_id']}
    )
    picking_map = {p['id']: p for p in pickings}
    print(f"  → {len(pickings)} completed receipts found")

    # Get done stock moves with purchase line ref
    print("Fetching done purchase move lines...")
    done_moves = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.move', 'search_read',
        [[['picking_type_id.code', '=', 'incoming'],
          ['state', '=', 'done'],
          ['purchase_line_id', '!=', False],
          ['product_id', 'in', list(all_comp_ids)]]],
        {'fields': ['product_id', 'product_qty', 'quantity_done',
                    'picking_id', 'purchase_line_id', 'date']}
    )
    print(f"  → {len(done_moves)} done purchase moves found")

    # Get PO confirmation dates
    po_line_ids = list({m['purchase_line_id'][0] for m in done_moves if m.get('purchase_line_id')})
    print(f"  → Fetching {len(po_line_ids)} PO lines...")

    # Batch fetch PO lines
    po_lines = {}
    batch_size = 200
    for i in range(0, len(po_line_ids), batch_size):
        batch = po_line_ids[i:i+batch_size]
        lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order.line', 'read',
            [batch], {'fields': ['id', 'order_id', 'product_id', 'product_qty']}
        )
        for l in lines:
            po_lines[l['id']] = l

    # Get PO confirmation dates
    po_ids = list({l['order_id'][0] for l in po_lines.values() if l.get('order_id')})
    print(f"  → Fetching {len(po_ids)} POs...")
    po_map = {}
    for i in range(0, len(po_ids), batch_size):
        batch = po_ids[i:i+batch_size]
        pos = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order', 'read',
            [batch], {'fields': ['id', 'name', 'date_approve', 'partner_id']}
        )
        for p in pos:
            po_map[p['id']] = p

    # --- Calculate actual lead times per product ---
    print("\nCalculating actual lead times...")
    from collections import defaultdict
    # product_id -> list of actual lead days
    actual_leads = defaultdict(list)
    actual_details = defaultdict(list)  # product_id -> list of detail dicts

    for move in done_moves:
        if not move.get('purchase_line_id') or not move.get('picking_id'):
            continue
        pl_id   = move['purchase_line_id'][0]
        pick_id = move['picking_id'][0]
        pl      = po_lines.get(pl_id)
        pick    = picking_map.get(pick_id)
        if not pl or not pick:
            continue
        po = po_map.get(pl['order_id'][0]) if pl.get('order_id') else None
        if not po or not po.get('date_approve') or not pick.get('date_done'):
            continue

        try:
            date_confirmed = datetime.strptime(po['date_approve'][:19], '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
            date_received  = datetime.strptime(pick['date_done'][:19], '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc)
            actual_days    = (date_received - date_confirmed).days
            if actual_days < 0 or actual_days > 730:  # skip anomalies
                continue
            prod_id = move['product_id'][0]
            actual_leads[prod_id].append(actual_days)
            actual_details[prod_id].append({
                'po':       po['name'],
                'supplier': po['partner_id'][1] if po.get('partner_id') else '',
                'confirmed': date_confirmed.strftime('%Y-%m-%d'),
                'received':  date_received.strftime('%Y-%m-%d'),
                'actual_days': actual_days,
                'qty':       move['quantity_done'],
            })
        except Exception:
            continue

    print(f"  → Actual lead times calculated for {len(actual_leads)} products")

    # --- Get product details ---
    all_prod_ids = list(actual_leads.keys())
    prod_details = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['id', 'in', all_prod_ids]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    prod_map = {p['id']: p for p in prod_details}

    # --- Build rows ---
    rows = []
    for prod_id, lead_list in actual_leads.items():
        prod     = prod_map.get(prod_id, {})
        tmpl_id  = prod.get('product_tmpl_id', [None])[0] if prod else None
        sup      = sup_by_tmpl.get(tmpl_id, {})
        configured_lead = sup.get('delay', 0) or 0
        sup_name = sup['name'][1] if sup.get('name') else 'No supplier'

        avg_actual  = round(sum(lead_list) / len(lead_list))
        min_actual  = min(lead_list)
        max_actual  = max(lead_list)
        last_detail = sorted(actual_details[prod_id], key=lambda x: x['received'], reverse=True)[0]
        last_actual = last_detail['actual_days']
        num_orders  = len(lead_list)

        diff = last_actual - configured_lead
        if abs(diff) <= 5:       status = "✅ OK"
        elif diff > 30:          status = "🔴 Much longer"
        elif diff > 5:           status = "🟠 Longer"
        elif diff < -5:          status = "🟡 Shorter"
        else:                    status = "✅ OK"

        rows.append({
            'ref':              prod.get('default_code', '') or '',
            'name':             prod.get('name', '') or '',
            'supplier':         sup_name,
            'configured_lead':  configured_lead,
            'last_actual':      last_actual,
            'avg_actual':       avg_actual,
            'min_actual':       min_actual,
            'max_actual':       max_actual,
            'num_orders':       num_orders,
            'diff':             diff,
            'status':           status,
            'last_po':          last_detail['po'],
            'last_confirmed':   last_detail['confirmed'],
            'last_received':    last_detail['received'],
            'last_supplier':    last_detail['supplier'],
        })

    # Sort by biggest discrepancy first
    rows.sort(key=lambda x: abs(x['diff']), reverse=True)

    # ================================================================
    # BUILD EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Time Sanity Check"

    MAIN_COLOR = "2E4057"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 14
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 25

    ws.append(["Lead Time Sanity Check — Configured vs Actual"])
    ws.merge_cells("A1:O1")
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = make_fill(MAIN_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:O2")
    ws["A2"] = (f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   "
                f"{len(rows)} products with purchase history   |   "
                f"Based on last receipt vs PO confirmation date")
    ws["A2"].font      = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    headers = ["Ref", "Product Name", "Supplier", "Configured\nLead (days)",
               "Last Actual\n(days)", "Avg Actual\n(days)", "Min\n(days)", "Max\n(days)",
               "# Orders", "Diff\n(last vs config)", "Status",
               "Last PO", "PO Confirmed", "Receipt Date", "Last Supplier"]
    ws.append(headers)
    r = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.fill      = make_fill(MAIN_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[r].height = 35
    ws.freeze_panes = "A4"

    STATUS_COLORS = {
        "✅ OK":          "E2EFDA",
        "🔴 Much longer": "FFB3B3",
        "🟠 Longer":      "FFD9B3",
        "🟡 Shorter":     "FFF2CC",
    }

    for row in rows:
        fill_color = STATUS_COLORS.get(row['status'], "FFFFFF")
        diff_str = f"+{row['diff']}" if row['diff'] > 0 else str(row['diff'])
        ws.append([
            row['ref'], row['name'][:38], row['supplier'][:23],
            row['configured_lead'], row['last_actual'], row['avg_actual'],
            row['min_actual'], row['max_actual'], row['num_orders'],
            diff_str, row['status'],
            row['last_po'], row['last_confirmed'], row['last_received'],
            row['last_supplier'][:23]
        ])
        r = ws.max_row
        row_fill = make_fill(fill_color)
        for col in range(1, 16):
            c = ws.cell(row=r, column=col)
            c.fill      = row_fill
            c.border    = thin_border()
            c.alignment = Alignment(
                horizontal="left" if col in [1, 2, 3, 15] else "center",
                vertical="center"
            )
        ws.row_dimensions[r].height = 15

    # Summary at bottom
    ws.append([])
    ws.append(["SUMMARY"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ok_cnt    = sum(1 for r in rows if r['status'] == "✅ OK")
    long_cnt  = sum(1 for r in rows if "longer" in r['status'].lower() or "Much" in r['status'])
    short_cnt = sum(1 for r in rows if r['status'] == "🟡 Shorter")
    ws.append([f"✅ OK (within 5 days): {ok_cnt}  |  "
               f"🔴🟠 Lead time longer than configured: {long_cnt}  |  "
               f"🟡 Shorter: {short_cnt}"])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True)

    output_file = f"lead_time_check_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\nExcel saved: {output_file}")
    print(f"\n  ✅ OK:            {ok_cnt}")
    print(f"  🔴🟠 Longer:      {long_cnt}")
    print(f"  🟡 Shorter:       {short_cnt}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
