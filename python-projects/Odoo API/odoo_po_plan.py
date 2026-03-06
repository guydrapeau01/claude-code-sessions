import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, timezone, date, timedelta
from math import ceil

# ================================================================
# CONFIG
# ================================================================
PLAN_PRODUCTS           = cfg.ALL_PLAN_PRODUCTS
KANBAN_SUPPLIERS        = cfg.KANBAN_SUPPLIERS

MONTHLY_PLAN  = cfg.MONTHLY_PRODUCTION_PLAN  # 101336:150, 101711:40, 101769:20, 102237:20

COVERAGE_TARGET  = 7   # months to have on hand after ordering (6mo + 1mo safety)
REORDER_POINT    = 3   # order when coverage drops below this (months)
ROUND_TO         = 50  # round order qty up to nearest N

# ================================================================
# HELPERS
# ================================================================
def make_fill(hex_): return PatternFill("solid", fgColor=hex_)
def tb():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def add_months(d, months):
    month = d.month - 1 + months
    year  = d.year + month // 12
    month = month % 12 + 1
    return date(year, month, 1)

def cell_style(c, fill_hex, bold=False, left=False, size=10, color="000000", wrap=False):
    c.fill      = make_fill(fill_hex)
    c.font      = Font(bold=bold, size=size, color=color)
    c.border    = tb()
    c.alignment = Alignment(horizontal="left" if left else "center",
                            vertical="center", wrap_text=wrap)

COLORS = {
    'main':   "2E4057", 'red':    "FFB3B3", 'amber':  "FFD9B3",
    'yellow': "FFF2CC", 'green':  "E2EFDA", 'blue':   "DDEBF7",
    'urgent': "C00000", 'warn':   "BF8F00", 'ok':     "375623",
    'sup':    "1F4E79", 'grey':   "F2F2F2",
}

def round_up_50(n):
    if n <= 0: return 0
    return ceil(n / ROUND_TO) * ROUND_TO

def explode_bom(models, uid, bom_id, bom_cache, result, sup_prod_ids=None):
    """Explode BOM. Stops at sub-assemblies that have a real supplier (order at that level)."""
    if bom_id not in bom_cache:
        lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom.line', 'search_read',
            [[['bom_id', '=', bom_id]]],
            {'fields': ['product_id', 'product_qty', 'product_uom_id', 'child_bom_id']}
        )
        bom_info = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'read',
            [[bom_id]], {'fields': ['product_qty']}
        )
        bom_cache[bom_id] = {'lines': lines, 'bom_qty': bom_info[0]['product_qty'] if bom_info else 1}
    cached = bom_cache[bom_id]
    ratio  = 1.0 / cached['bom_qty']
    for line in cached['lines']:
        comp_id   = line['product_id'][0]
        comp_name = line['product_id'][1]
        comp_qty  = line['product_qty'] * ratio
        uom       = line['product_uom_id'][1] if line.get('product_uom_id') else 'Unit'
        if line.get('child_bom_id'):
            # If this sub-assembly has a supplier, buy it — don't explode its children
            if sup_prod_ids and comp_id in sup_prod_ids:
                if comp_id not in result:
                    result[comp_id] = {'name': comp_name, 'qty': 0.0, 'uom': uom}
                result[comp_id]['qty'] += comp_qty
            else:
                sub = {}
                explode_bom(models, uid, line['child_bom_id'][0], bom_cache, sub, sup_prod_ids)
                for sid, sinfo in sub.items():
                    if sid not in result:
                        result[sid] = {'name': sinfo['name'], 'qty': 0.0, 'uom': sinfo['uom']}
                    result[sid]['qty'] += sinfo['qty'] * comp_qty * cached['bom_qty']
        else:
            if comp_id not in result:
                result[comp_id] = {'name': comp_name, 'qty': 0.0, 'uom': uom}
            result[comp_id]['qty'] += comp_qty

try:
    print("Connecting...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")
    now = datetime.now(timezone.utc)
    today = now.date()

    # --- Resolve finished products & BOMs ---
    fin_prods = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', PLAN_PRODUCTS]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    prod_by_ref = {p['default_code']: p for p in fin_prods}

    bom_by_ref = {}
    for ref in PLAN_PRODUCTS:
        boms = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'search_read',
            [[['product_tmpl_id.default_code', '=', ref]]],
            {'fields': ['id', 'product_qty']}
        )
        if boms:
            bom_by_ref[ref] = boms[0]
            print(f"  BOM {ref}: ID {boms[0]['id']}")
        else:
            print(f"  ⚠ No BOM: {ref}")

    # --- Identify which plan products are manufactured devices (have a BOM) ---
    # These should NOT have their stock counted as available — they are built, not bought
    manufactured_device_ids = set()
    for ref in PLAN_PRODUCTS:
        fp = prod_by_ref.get(ref, {})
        if fp and ref in bom_by_ref:
            manufactured_device_ids.add(fp['id'])
    print(f"  Manufactured devices (stock excluded from availability): {len(manufactured_device_ids)}")

    # --- Fetch ALL supplier info upfront (needed before BOM explosion to stop at purchased sub-assemblies) ---
    print("Fetching all supplier info...")
    all_suppliers_raw = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.supplierinfo', 'search_read',
        [[]], {'fields': ['product_tmpl_id', 'name', 'delay', 'min_qty', 'price', 'product_code', 'sequence']}
    )
    # Rule: stop BOM explosion at a sub-assembly if it has an EXTERNAL supplier
    # If supplier is THORASYS (internal) or no supplier → keep exploding
    INTERNAL_SUPPLIER_KEYWORDS = ['THORASYS', 'Thorasys', 'thorasys']

    # Build dict: tmpl_id -> supplier name for first/primary supplier
    sup_by_tmpl_all = {}
    for s in sorted(all_suppliers_raw, key=lambda x: x.get('sequence', 99)):
        tid = s['product_tmpl_id'][0]
        if tid not in sup_by_tmpl_all:
            sup_by_tmpl_all[tid] = s['name'][1] if s.get('name') else ''

    # Templates with an EXTERNAL supplier = stop explosion here
    external_sup_tmpl_ids = [
        tid for tid, sname in sup_by_tmpl_all.items()
        if not any(kw in sname for kw in INTERNAL_SUPPLIER_KEYWORDS)
    ]

    all_sup_prod_ids = set()
    if external_sup_tmpl_ids:
        sup_variants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
            [[['product_tmpl_id', 'in', external_sup_tmpl_ids]]],
            {'fields': ['id']}
        )
        all_sup_prod_ids = {v['id'] for v in sup_variants}

    internal_count = len(sup_by_tmpl_all) - len(external_sup_tmpl_ids)
    print(f"  {len(all_sup_prod_ids)} variants with external supplier → explosion stops (order from supplier)")
    print(f"  {internal_count} variants with THORASYS/internal supplier → will be exploded further")

    # --- Explode all BOMs (stops at sub-assemblies that have a supplier) ---
    print("\nExploding BOMs...")
    bom_cache   = {}
    comp_monthly = {}
    bom_per_prod = {}

    # Track products with no BOM — treated as direct purchased items
    direct_purchase_refs = set()

    for ref, monthly_qty in MONTHLY_PLAN.items():
        fp = prod_by_ref.get(ref, {})
        if not fp:
            print(f"  ⚠ Product not found: {ref}")
            continue
        if ref not in bom_by_ref:
            # No BOM — this is a purchased finished good, plan it directly
            comp_id = fp['id']
            direct_purchase_refs.add(ref)
            if comp_id not in comp_monthly:
                comp_monthly[comp_id] = {'name': fp['name'], 'monthly': 0.0, 'uom': 'Unit'}
            comp_monthly[comp_id]['monthly'] += monthly_qty
            bom_per_prod[ref] = {comp_id: 1.0}  # 1 unit of itself
            print(f"  📦 {ref} has no BOM — treated as direct purchase ({monthly_qty}/mo)")
            continue
        result = {}
        explode_bom(models, uid, bom_by_ref[ref]['id'], bom_cache, result, all_sup_prod_ids)
        bom_per_prod[ref] = {cid: info['qty'] for cid, info in result.items()}
        for comp_id, info in result.items():
            if comp_id not in comp_monthly:
                comp_monthly[comp_id] = {'name': info['name'], 'monthly': 0.0, 'uom': info['uom']}
            comp_monthly[comp_id]['monthly'] += info['qty'] * monthly_qty

    for ref in [r for r in PLAN_PRODUCTS if r not in MONTHLY_PLAN]:
        fp = prod_by_ref.get(ref, {})
        if not fp:
            continue
        if ref not in bom_by_ref:
            # No BOM, not in monthly plan — still include for buildable units check
            comp_id = fp['id']
            direct_purchase_refs.add(ref)
            if comp_id not in comp_monthly:
                comp_monthly[comp_id] = {'name': fp['name'], 'monthly': 0.0, 'uom': 'Unit'}
            bom_per_prod[ref] = {comp_id: 1.0}
            continue
        result = {}
        explode_bom(models, uid, bom_by_ref[ref]['id'], bom_cache, result, all_sup_prod_ids)
        bom_per_prod[ref] = {cid: info['qty'] for cid, info in result.items()}
        for comp_id, info in result.items():
            if comp_id not in comp_monthly:
                comp_monthly[comp_id] = {'name': info['name'], 'monthly': 0.0, 'uom': info['uom']}

    if direct_purchase_refs:
        print(f"  📦 Direct purchase products (no BOM): {', '.join(direct_purchase_refs)}")

    all_comp_ids = list(comp_monthly.keys())
    print(f"  {len(all_comp_ids)} unique components\n")

    # --- Component details ---
    comp_prods = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['id', 'in', all_comp_ids]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    comp_info   = {p['id']: p for p in comp_prods}
    tmpl_ids    = [p['product_tmpl_id'][0] for p in comp_prods]

    # --- Supplier info (filtered to our components) ---
    print("Fetching supplier info for components...")
    suppliers = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.supplierinfo', 'search_read',
        [[['product_tmpl_id', 'in', tmpl_ids]]],
        {'fields': ['product_tmpl_id', 'name', 'delay', 'min_qty', 'price',
                    'product_code', 'sequence']}
    )
    sup_by_tmpl = {}
    for s in sorted(suppliers, key=lambda x: x.get('sequence', 99)):
        tid = s['product_tmpl_id'][0]
        if tid not in sup_by_tmpl:
            sup_by_tmpl[tid] = s

    # --- Identify kanban components (from kanban suppliers or 0 stock + 0 incoming) ---
    kanban_comp_ids = set()
    for p in comp_prods:
        tid = p['product_tmpl_id'][0]
        sup = sup_by_tmpl.get(tid, {})
        sup_name = sup['name'][1] if sup.get('name') else 'NO SUPPLIER'
        if sup_name in KANBAN_SUPPLIERS:
            kanban_comp_ids.add(p['id'])
    print(f"  Kanban components identified: {len(kanban_comp_ids)}")

    # --- Last PO price per component ---
    print("Fetching last PO prices and currencies...")
    done_po_lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order.line', 'search_read',
        [[['product_id', 'in', all_comp_ids],
          ['order_id.state', 'in', ['purchase', 'done']]]],
        {'fields': ['product_id', 'price_unit', 'currency_id', 'order_id']}
    )
    po_ids_price = list({l['order_id'][0] for l in done_po_lines if l.get('order_id')})
    po_dates = {}
    if po_ids_price:
        pos = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order', 'read',
            [po_ids_price], {'fields': ['id', 'name', 'date_approve', 'currency_id']}
        )
        po_dates = {p['id']: p for p in pos}

    last_po_price = {}
    for l in done_po_lines:
        if not l.get('order_id') or not l.get('price_unit'):
            continue
        po       = po_dates.get(l['order_id'][0], {})
        po_date  = po.get('date_approve', '') or ''
        comp_id  = l['product_id'][0]
        currency = po.get('currency_id') or l.get('currency_id')
        if comp_id not in last_po_price or po_date > last_po_price[comp_id]['date']:
            last_po_price[comp_id] = {
                'price':         l['price_unit'],
                'currency_name': currency[1] if currency else 'USD',
                'date':          po_date,
                'po_name':       po.get('name', ''),
            }
    print(f"  Last PO prices: {len(last_po_price)} components")

    # --- Currency rates (convert everything to USD) ---
    print("Fetching currency rates...")
    currencies = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'res.currency', 'search_read',
        [[['active', '=', True]]], {'fields': ['id', 'name', 'rate']}
    )
    rate_map = {c['name']: c['rate'] for c in currencies}
    usd_rate  = rate_map.get('USD', 1.0)

    def to_usd(amount, currency_name):
        if not amount: return 0.0
        if currency_name == 'USD': return amount
        curr_rate = rate_map.get(currency_name, 1.0)
        if curr_rate == 0: return amount
        return round(amount * (usd_rate / curr_rate), 4)

    cad_sample = to_usd(1, 'CAD')
    print(f"  Rates loaded: {len(rate_map)} currencies | 1 CAD = {cad_sample:.4f} USD")

    # --- Stock ---
    print("Fetching stock...")
    quants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.quant', 'search_read',
        [[['product_id', 'in', all_comp_ids], ['location_id.usage', '=', 'internal']]],
        {'fields': ['product_id', 'quantity', 'reserved_quantity']}
    )
    _raw = defaultdict(lambda: [0.0, 0.0])
    for q in quants:
        _raw[q['product_id'][0]][0] += q['quantity']
        _raw[q['product_id'][0]][1] += q['reserved_quantity']
    stock_by_comp = {}
    for pid, (qty, res) in _raw.items():
        if pid in manufactured_device_ids:
            stock_by_comp[pid] = 0.0  # manufactured — ignore stock, plan full demand
        else:
            stock_by_comp[pid] = max(0, qty - res)

    # --- Open incoming moves ---
    print("Fetching open incoming...")
    moves = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.move', 'search_read',
        [[['product_id', 'in', all_comp_ids],
          ['state', 'in', ['waiting', 'confirmed', 'assigned', 'partially_available']],
          ['picking_type_id.code', '=', 'incoming']]],
        {'fields': ['product_id', 'product_qty', 'quantity_done', 'picking_id']}
    )
    pick_ids = list({m['picking_id'][0] for m in moves if m.get('picking_id')})
    pick_states = {}
    if pick_ids:
        picks = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.picking', 'read',
            [pick_ids], {'fields': ['id', 'state']}
        )
        pick_states = {p['id']: p['state'] for p in picks}
    incoming_by_comp = defaultdict(float)
    for m in moves:
        comp_id_m  = m['product_id'][0]
        if comp_id_m in manufactured_device_ids:
            continue  # manufactured device — skip incoming stock moves (those are MOs not POs)
        pick_state = pick_states.get(m['picking_id'][0] if m.get('picking_id') else None, '')
        ordered    = m.get('product_qty') or 0
        done       = m.get('quantity_done') or 0
        remaining  = ordered if pick_state != 'done' else max(0, ordered - done)
        if remaining > 0:
            incoming_by_comp[comp_id_m] += remaining

    # Add zero stock + zero incoming parts with no supplier to kanban set
    for comp_id in all_comp_ids:
        if comp_id in kanban_comp_ids:
            continue
        info    = comp_info.get(comp_id, {})
        tmpl_id = info.get('product_tmpl_id', [None])[0] if info else None
        sup     = sup_by_tmpl.get(tmpl_id, {})
        sup_name = sup['name'][1] if sup.get('name') else 'NO SUPPLIER'
        if (sup_name == 'NO SUPPLIER'
                and stock_by_comp.get(comp_id, 0) == 0
                and incoming_by_comp.get(comp_id, 0) == 0):
            kanban_comp_ids.add(comp_id)
    print(f"  Total kanban/untracked: {len(kanban_comp_ids)} (excluded from PO plan & buildable units)")

    # ================================================================
    # CALCULATE PO PLAN
    # ================================================================
    print("Calculating PO plan...")
    plan_rows = []

    for comp_id in all_comp_ids:
        if comp_id in kanban_comp_ids:
            continue  # kanban — skip from PO plan
        info     = comp_info.get(comp_id, {})
        tmpl_id  = info.get('product_tmpl_id', [None])[0] if info else None
        sup      = sup_by_tmpl.get(tmpl_id, {})

        monthly      = comp_monthly[comp_id]['monthly']
        uom          = comp_monthly[comp_id]['uom']
        lead_days    = int(sup.get('delay', 0) or 0)
        min_qty      = sup.get('min_qty', 0) or 0
        sup_name     = sup['name'][1] if sup.get('name') else 'NO SUPPLIER'
        sup_code     = sup.get('product_code', '') or ''
        lpp          = last_po_price.get(comp_id, {})
        raw_price    = lpp.get('price', sup.get('price', 0) or 0)
        price_curr   = lpp.get('currency_name', 'USD')
        price_usd    = to_usd(raw_price, price_curr)
        price_source = (lpp['po_name'][:14] + ' (' + price_curr + ')') if lpp else 'pricelist'
        price        = price_usd

        stock        = stock_by_comp.get(comp_id, 0)
        incoming     = incoming_by_comp.get(comp_id, 0)
        available    = stock + incoming

        # Current coverage in months
        if monthly > 0:
            coverage_now = available / monthly
        else:
            coverage_now = 99.0  # not consumed

        # Target qty on hand = COVERAGE_TARGET months
        target_qty   = monthly * COVERAGE_TARGET

        # Net to order = target - available, rounded up to nearest 50
        # (only if coverage < reorder point OR coverage < target)
        raw_order    = max(0, target_qty - available)
        if raw_order > 0:
            order_qty = round_up_50(max(raw_order, min_qty))
        else:
            order_qty = 0

        # Coverage AFTER order
        coverage_after = (available + order_qty) / monthly if monthly > 0 else 99.0

        # When does stock hit reorder point (3 months)?
        # reorder_point_qty = monthly * REORDER_POINT
        # If available > reorder_point_qty, we have time before we need to order
        reorder_qty  = monthly * REORDER_POINT
        if available > reorder_qty and monthly > 0:
            # months until we hit reorder point
            months_until_reorder = (available - reorder_qty) / monthly
            reorder_date = add_months(today, int(months_until_reorder))
        else:
            reorder_date = today  # already at or below reorder point

        # Latest SAFE order date = today + buffer days (when buffer hits 0 = must order)
        # Buffer = days of stock remaining - lead time
        # Latest safe order date = today + buffer (if buffer > 0), else today
        _buffer_preview = (available / monthly * 30.44) - lead_days if monthly > 0 else 9999
        if _buffer_preview > 0:
            order_by_date = today + timedelta(days=int(_buffer_preview))
        else:
            order_by_date = today  # already critical
        days_to_order = (order_by_date - today).days  # kept for internal sort only

        # Flags
        no_sup    = not sup or sup_name == 'NO SUPPLIER'
        zero_lead = bool(sup) and lead_days == 0
        no_order  = order_qty == 0

        # Real risk: will stock run out before order arrives if placed TODAY?
        if monthly > 0 and not no_order:
            days_of_stock = coverage_now * 30.44 if coverage_now < 99 else 9999
            days_buffer   = days_of_stock - lead_days
            will_stockout = days_buffer < 0
        else:
            days_of_stock = 9999
            days_buffer   = 9999
            will_stockout = False

        urgent = not no_order and will_stockout
        soon   = not no_order and not will_stockout and days_buffer <= 14

        if no_order:
            status     = "✅ Sufficient"
            status_col = COLORS['green']
        elif will_stockout:
            status     = f"🔴 CRITICAL — stock out in {int(days_of_stock)}d, lead {lead_days}d — ORDER TODAY"
            status_col = COLORS['urgent']
        elif days_buffer <= 14:
            status     = f"🟠 Order NOW — only {int(days_buffer)}d buffer"
            status_col = "CC5500"
        elif days_to_order < 0:
            # Missed ideal order date but still safe — show real buffer clearly
            if days_buffer <= 30:
                status     = f"🟠 Order THIS WEEK — only {int(days_buffer)}d buffer remaining"
                status_col = "CC5500"
            else:
                status     = f"🟡 Order now — {int(days_buffer)}d until stockout risk (lead {lead_days}d)"
                status_col = COLORS['warn']
        elif days_to_order <= 45:
            status     = f"🟡 Order by {order_by_date} — {int(days_buffer)}d buffer (lead {lead_days}d)"
            status_col = COLORS['warn']
        else:
            status     = f"🟢 Order by {order_by_date} — {int(days_buffer)}d buffer (lead {lead_days}d)"
            status_col = COLORS['ok']

        if no_sup:    status = "⚠ NO SUPPLIER — " + status
        if zero_lead: status = "⚠ CHECK LEAD TIME — " + status

        est_value       = order_qty * price if price else 0
        days_buffer_val = int(days_buffer) if days_buffer < 9999 else 999

        plan_rows.append({
            'comp_id':        comp_id,
            'ref':            info.get('default_code', '') or '',
            'name':           comp_monthly[comp_id]['name'],
            'uom':            uom,
            'supplier':       sup_name,
            'sup_code':       sup_code,
            'lead_days':      lead_days,
            'min_qty':        int(min_qty) if min_qty else 0,
            'price':          price,
            'stock':          round(stock, 1),
            'incoming':       round(incoming, 1),
            'available':      round(available, 1),
            'monthly_req':    round(monthly, 2),
            'coverage_now':   round(coverage_now, 1) if coverage_now < 99 else '∞',
            'reorder_point':  round(reorder_qty, 1),
            'target_qty':     round(target_qty, 1),
            'order_qty':      order_qty,
            'coverage_after': round(coverage_after, 1) if coverage_after < 99 else '∞',
            'order_by_date':  order_by_date.strftime('%Y-%m-%d') if not no_order else '',
            'days_to_order':  days_to_order if not no_order else 999,
            'reorder_date':   reorder_date.strftime('%Y-%m-%d') if not no_order else '',
            'est_value':      round(est_value, 2),
            'price_usd':      round(price_usd, 4),
            'price_curr':     price_curr,
            'price_source':   price_source,
            'status':         status,
            'status_col':     status_col,
            'no_order':       no_order,
            'no_sup':         no_sup,
            'zero_lead':      zero_lead,
            'urgent':         urgent,
            'soon':           soon,
            'days_buffer':    days_buffer_val,
        })

    # Sort: urgent → soon → by days_to_order → sufficient
    plan_rows.sort(key=lambda x: (
        0 if x['urgent'] else 1 if x['soon'] else 2 if not x['no_order'] else 3,
        x['days_to_order'],
        x['ref']
    ))

    cnt_urgent  = sum(1 for r in plan_rows if r['urgent'])
    cnt_soon    = sum(1 for r in plan_rows if r['soon'])
    cnt_plan    = sum(1 for r in plan_rows if not r['no_order'] and not r['urgent'] and not r['soon'])
    cnt_ok      = sum(1 for r in plan_rows if r['no_order'])
    total_value = sum(r['est_value'] for r in plan_rows)

    print(f"  🔴 ORDER NOW:   {cnt_urgent}")
    print(f"  🟠 Next 14d:    {cnt_soon}")
    print(f"  🟡 Plan ahead:  {cnt_plan}")
    print(f"  ✅ Sufficient:  {cnt_ok}")
    print(f"  💰 Est. value:  ${total_value:,.0f}\n")

    # ================================================================
    # EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    COLS = [
        ("Ref",              11), ("Component Name",   40), ("Supplier",         22),
        ("Sup Code",         12), ("Lead\nDays",        9), ("Min\nQty",          9),
        ("Stock\nAvail",     10), ("Open\nIncoming",   10), ("Total\nAvail",      10),
        ("Monthly\nReq",     10), ("Coverage\nNow(mo)",11), ("Reorder\nPoint",   10),
        ("Target\nQty(7mo)", 11), ("ORDER\nQTY",       11), ("Coverage\nAfter",  11),
        ("Latest Safe\nOrder Date", 13), ("Buffer\n(days)",   10),
        ("Price Source",     16), ("Est Value\n(USD)",  13),
        ("Status",           32),
    ]
    NCOLS = len(COLS)

    def make_sheet(ws, title, subtitle=""):
        for i, (_, w) in enumerate(COLS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        c = ws["A1"]
        c.font      = Font(bold=True, size=14, color="FFFFFF")
        c.fill      = make_fill(COLORS['main'])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        plan_str = '  '.join(f"{ref}×{qty}" for ref, qty in MONTHLY_PLAN.items())
        sub = (f"Rate: {plan_str} /mo  |  "
               f"Target: {COVERAGE_TARGET}mo  |  Reorder at: {REORDER_POINT}mo  |  "
               f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC"
               + (f"  |  {subtitle}" if subtitle else ""))
        ws.cell(row=2, column=1, value=sub)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
        ws["A2"].font      = Font(italic=True, size=9)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 13
        for i, (h, _) in enumerate(COLS, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font      = Font(color="FFFFFF", bold=True, size=9)
            c.fill      = make_fill(COLORS['main'])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = tb()
        ws.row_dimensions[3].height = 32
        ws.freeze_panes = "A4"

    def row_color(r):
        if r['no_sup'] or r['zero_lead']: return COLORS['yellow']
        if r['urgent']:   return COLORS['red']
        if r['soon']:     return "FFD0A0"
        if r['no_order']: return COLORS['green']
        # days_to_order drives shade
        d = r['days_to_order']
        if d <= 45:  return COLORS['amber']
        return COLORS['blue']

    def write_row(ws, r):
        fill = make_fill(row_color(r))
        cov  = r['coverage_now']
        ca   = r['coverage_after']
        vals = [
            r['ref'], r['name'][:38], r['supplier'][:20], r['sup_code'],
            r['lead_days'], r['min_qty'] or '',
            r['stock'], r['incoming'], r['available'],
            r['monthly_req'] if r['monthly_req'] else '',
            cov, r['reorder_point'], r['target_qty'],
            r['order_qty'] if r['order_qty'] else '',
            ca if not r['no_order'] else '',
            r['order_by_date'], r['days_buffer'] if not r['no_order'] else '',
            r.get('price_source', ''),
            r['est_value'] if r['est_value'] else '',
            r['status'],
        ]
        ws.append(vals)
        rn = ws.max_row
        for col, v in enumerate(vals, 1):
            c = ws.cell(rn, col)
            c.fill      = fill
            c.border    = tb()
            c.alignment = Alignment(
                horizontal="left" if col in [1,2,3,4,19] else "center",
                vertical="center"
            )
            if col == 14 and r['order_qty']:  # ORDER QTY bold
                c.font = Font(bold=True)
            if col == 18 and r['est_value']:
                c.number_format = '$#,##0.00'
        ws.row_dimensions[rn].height = 14

    def section_hdr(ws, text, color):
        ws.append([text])
        rn = ws.max_row
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=NCOLS)
        c = ws.cell(rn, 1)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = make_fill(color)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[rn].height = 15

    # ================================================================
    # SHEET 1 — FULL PO PLAN
    # ================================================================
    ws1 = wb.create_sheet("📋 PO Plan")
    make_sheet(ws1, f"Component PO Plan — 6-Month Rolling Coverage ({len(plan_rows)} components)")

    urgent_rows = [r for r in plan_rows if r['urgent']]
    soon_rows   = [r for r in plan_rows if r['soon']]
    plan_ahead  = [r for r in plan_rows if not r['no_order'] and not r['urgent'] and not r['soon']]
    ok_rows     = [r for r in plan_rows if r['no_order']]

    if urgent_rows:
        section_hdr(ws1, f"  🔴  ORDER NOW — {len(urgent_rows)} components (lead time already exceeded)", COLORS['urgent'])
        for r in urgent_rows: write_row(ws1, r)
    if soon_rows:
        section_hdr(ws1, f"  🟠  ORDER WITHIN 14 DAYS — {len(soon_rows)} components", "CC5500")
        for r in soon_rows: write_row(ws1, r)
    if plan_ahead:
        section_hdr(ws1, f"  🟡  PLAN AHEAD — {len(plan_ahead)} components (order date upcoming)", COLORS['warn'])
        for r in plan_ahead: write_row(ws1, r)
    if ok_rows:
        section_hdr(ws1, f"  ✅  SUFFICIENT STOCK — {len(ok_rows)} components (coverage > {COVERAGE_TARGET} months)", COLORS['ok'])
        for r in ok_rows: write_row(ws1, r)

    # ================================================================
    # SHEET 2 — BY SUPPLIER
    # ================================================================
    ws2 = wb.create_sheet("🏭 By Supplier")
    make_sheet(ws2, "Component PO Plan — Grouped by Supplier")

    by_sup = defaultdict(list)
    for r in plan_rows:
        by_sup[r['supplier']].append(r)

    def sup_sort_key(s):
        rows = by_sup[s]
        if any(r['urgent'] for r in rows): return 0
        if any(r['soon'] for r in rows):   return 1
        if any(not r['no_order'] for r in rows): return 2
        return 3

    for sup_name in sorted(by_sup.keys(), key=sup_sort_key):
        rows    = by_sup[sup_name]
        to_ord  = [r for r in rows if not r['no_order']]
        val     = sum(r['est_value'] for r in rows)
        urg     = sum(1 for r in rows if r['urgent'])
        col     = COLORS['urgent'] if urg else "CC5500" if any(r['soon'] for r in rows) \
                  else COLORS['warn'] if to_ord else COLORS['ok']
        section_hdr(ws2,
            f"  🏭  {sup_name}  —  {len(rows)} parts  |  {len(to_ord)} to order  |  Est. ${val:,.0f}",
            col)
        for r in sorted(rows, key=lambda x: x['days_to_order']):
            write_row(ws2, r)

    # ================================================================
    # SHEET 3 — BUILDABLE UNITS
    # ================================================================
    ws3 = wb.create_sheet("🔧 Buildable Units")
    BU_COLS = [
        ("Ref", 12), ("Finished Product", 42), ("Monthly Plan", 13),
        ("Max Build\n(Stock Only)", 14), ("Months\nCoverage", 12),
        ("Max Build\n(+Incoming)", 14), ("Months\nCoverage", 12),
        ("Limiting Component\n(Stock Only)", 42),
        ("Limiting Component\n(+Incoming)", 42),
    ]
    for i, (_, w) in enumerate(BU_COLS, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.cell(row=1, column=1, value="Buildable Units — How Many Devices Can We Build Right Now?")
    ws3.merge_cells(f"A1:{get_column_letter(len(BU_COLS))}1")
    ws3["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill      = make_fill(COLORS['main'])
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    ws3.cell(row=2, column=1,
             value="Stock Only = physical stock now  |  +Incoming = stock + all open PO receipts expected")
    ws3.merge_cells(f"A2:{get_column_letter(len(BU_COLS))}2")
    ws3["A2"].font      = Font(italic=True, size=9)
    ws3["A2"].alignment = Alignment(horizontal="center")

    for i, (h, _) in enumerate(BU_COLS, 1):
        c = ws3.cell(row=3, column=i, value=h)
        c.font      = Font(color="FFFFFF", bold=True, size=9)
        c.fill      = make_fill(COLORS['main'])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = tb()
    ws3.row_dimensions[3].height = 32
    ws3.freeze_panes = "A4"

    for ref in PLAN_PRODUCTS:
        if ref not in bom_per_prod:
            continue
        bom       = bom_per_prod[ref]
        fp        = prod_by_ref.get(ref, {})
        fp_name   = fp.get('name', ref) if fp else ref
        monthly   = MONTHLY_PLAN.get(ref, 0)
        if not bom: continue

        lim_s = lim_t = None
        for comp_id, qty_per_unit in bom.items():
            if qty_per_unit <= 0: continue
            if comp_id in kanban_comp_ids: continue  # kanban — always available
            s  = stock_by_comp.get(comp_id, 0)
            t  = s + incoming_by_comp.get(comp_id, 0)
            bs = int(s / qty_per_unit)
            bt = int(t / qty_per_unit)
            if lim_s is None or bs < lim_s['bs']:
                lim_s = {'id': comp_id, 'bs': bs, 'stock': s, 'qpu': qty_per_unit}
            if lim_t is None or bt < lim_t['bt']:
                lim_t = {'id': comp_id, 'bt': bt, 'total': t, 'qpu': qty_per_unit}

        max_s  = lim_s['bs'] if lim_s else 0
        max_t  = lim_t['bt'] if lim_t else 0
        mo_s   = round(max_s / monthly, 1) if monthly else '∞'
        mo_t   = round(max_t / monthly, 1) if monthly else '∞'

        def lim_str(lim, key, val_key):
            if not lim: return ''
            ci = comp_info.get(lim['id'], {})
            return (f"{ci.get('default_code','')[:10]} — "
                    f"{comp_monthly.get(lim['id'],{}).get('name','')[:32]}  "
                    f"({lim[val_key]:.0f} avail / {lim['qpu']:.3f}/unit = {lim[key]} units)")

        if isinstance(mo_s, float) and mo_s < 1:   fcolor = COLORS['red']
        elif isinstance(mo_s, float) and mo_s < 3: fcolor = COLORS['amber']
        else:                                        fcolor = COLORS['green']

        ws3.append([ref, fp_name[:40], monthly,
                    max_s, f"{mo_s} mo", max_t, f"{mo_t} mo",
                    lim_str(lim_s, 'bs', 'stock'),
                    lim_str(lim_t, 'bt', 'total')])
        rn = ws3.max_row
        for col in range(1, len(BU_COLS)+1):
            c = ws3.cell(rn, col)
            c.fill      = make_fill(fcolor)
            c.border    = tb()
            c.alignment = Alignment(
                horizontal="left" if col in [2,8,9] else "center",
                vertical="center", wrap_text=col in [8,9]
            )
        ws3.row_dimensions[rn].height = 28

    ws3.append([])
    # Detail: bottom 15 limiting components per product
    for ref in PLAN_PRODUCTS:
        if ref not in bom_per_prod: continue
        bom     = bom_per_prod[ref]
        fp      = prod_by_ref.get(ref, {})
        monthly = MONTHLY_PLAN.get(ref, 0)
        if not bom: continue

        details = []
        for comp_id, qpu in bom.items():
            if qpu <= 0: continue
            if comp_id in kanban_comp_ids: continue  # kanban — skip
            s  = stock_by_comp.get(comp_id, 0)
            t  = s + incoming_by_comp.get(comp_id, 0)
            details.append({
                'ref':  comp_info.get(comp_id,{}).get('default_code',''),
                'name': comp_monthly.get(comp_id,{}).get('name',''),
                'qpu':  qpu, 'stock': s, 'incoming': incoming_by_comp.get(comp_id,0),
                'total': t, 'bs': int(s/qpu), 'bt': int(t/qpu)
            })
        details.sort(key=lambda x: x['bs'])

        fp_name = fp.get('name', ref) if fp else ref
        ws3.append([f"{ref} — {fp_name[:38]}  |  Monthly: {monthly}  |  "
                    f"Max buildable (stock): {details[0]['bs'] if details else 0}  |  "
                    f"Max buildable (+incoming): {details[0]['bt'] if details else 0}"])
        rn = ws3.max_row
        ws3.merge_cells(f"A{rn}:{get_column_letter(len(BU_COLS))}{rn}")
        c = ws3.cell(rn,1)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = make_fill(COLORS['main'])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.row_dimensions[rn].height = 15

        sub_h = ["Ref","Component","Qty/Unit","Stock","Incoming","Total","Build(Stock)","Build(+Inc)",""]
        for i, h in enumerate(sub_h, 1):
            c = ws3.cell(ws3.max_row+1 if i==1 else ws3.max_row, column=i, value=h)
            c.font      = Font(color="FFFFFF", bold=True, size=9)
            c.fill      = make_fill(COLORS['sup'])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = tb()
        ws3.row_dimensions[ws3.max_row].height = 16

        for d in details[:15]:
            if d['bs'] == 0:         rf = make_fill(COLORS['red'])
            elif d['bs'] < monthly:  rf = make_fill(COLORS['amber'])
            else:                    rf = make_fill(COLORS['green'])
            ws3.append([d['ref'], d['name'][:38], round(d['qpu'],4),
                        round(d['stock'],1), round(d['incoming'],1), round(d['total'],1),
                        d['bs'], d['bt'], ''])
            rn = ws3.max_row
            for col in range(1,9):
                c = ws3.cell(rn,col)
                c.fill   = rf
                c.border = tb()
                c.alignment = Alignment(
                    horizontal="left" if col in [1,2] else "center",
                    vertical="center")
            ws3.row_dimensions[rn].height = 14
        ws3.append([])

    # ================================================================
    # SHEET 4 — SUMMARY (move to front)
    # ================================================================
    ws4 = wb.create_sheet("📊 Summary")
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 18
    ws4.column_dimensions["E"].width = 18

    ws4.cell(row=1, column=1, value="Component PO Plan — Executive Summary")
    ws4.merge_cells("A1:E1")
    ws4["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill      = make_fill(COLORS['main'])
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24
    ws4.append([])

    def srow(ws, label, val, fhex, bold=False, fmt=None):
        ws.append([label, val])
        rn = ws.max_row
        for col in [1,2]:
            c = ws.cell(rn, col)
            c.fill      = make_fill(fhex)
            c.font      = Font(bold=bold)
            c.border    = tb()
            c.alignment = Alignment(
                horizontal="left" if col==1 else "center",
                vertical="center", indent=1 if col==1 else 0)
            if fmt and col==2: c.number_format = fmt
        ws.row_dimensions[rn].height = 16

    plan_summary = '   '.join(f"{ref}: {qty}/mo" for ref, qty in MONTHLY_PLAN.items())
    srow(ws4, "Production rate (monthly)", plan_summary, COLORS['blue'], True)
    srow(ws4, "Coverage target", f"{COVERAGE_TARGET} months", COLORS['blue'])
    srow(ws4, "Reorder point", f"{REORDER_POINT} months coverage remaining", COLORS['blue'])
    srow(ws4, "Order qty rounding", f"Round up to nearest {ROUND_TO}", COLORS['blue'])
    srow(ws4, "Generated", now.strftime('%Y-%m-%d %H:%M UTC'), COLORS['blue'])
    ws4.append([])
    srow(ws4, "Total components in scope", len(plan_rows), COLORS['blue'], True)
    srow(ws4, "🔴 ORDER NOW (overdue)", cnt_urgent, COLORS['red'])
    srow(ws4, "🟠 Order within 14 days", cnt_soon, COLORS['amber'])
    srow(ws4, "🟡 Plan ahead (order date upcoming)", cnt_plan, COLORS['yellow'])
    srow(ws4, "✅ Sufficient stock", cnt_ok, COLORS['green'])
    srow(ws4, "⚠ No supplier configured", sum(1 for r in plan_rows if r['no_sup']), COLORS['yellow'])
    srow(ws4, "⚠ Lead time = 0 (needs review)", sum(1 for r in plan_rows if r['zero_lead'] and not r['no_sup']), COLORS['yellow'])
    ws4.append([])
    srow(ws4, "💰 Estimated total PO value", f"${total_value:,.0f}", COLORS['blue'], True)
    ws4.append([])

    # Supplier summary
    sup_hdr_row = ws4.max_row + 1
    for i, h in enumerate(["Supplier","Parts","To Order","Urgent","Est. Value"], 1):
        c = ws4.cell(sup_hdr_row, i, value=h)
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.fill      = make_fill(COLORS['main'])
        c.alignment = Alignment(horizontal="left" if i==1 else "center", vertical="center")
        c.border    = tb()
    ws4.row_dimensions[sup_hdr_row].height = 18

    for sname in sorted(by_sup.keys(), key=sup_sort_key):
        rows  = by_sup[sname]
        val   = sum(r['est_value'] for r in rows)
        urg   = sum(1 for r in rows if r['urgent'])
        ordn  = sum(1 for r in rows if not r['no_order'])
        fhex  = COLORS['red'] if urg else COLORS['amber'] if ordn else COLORS['green']
        ws4.append([sname[:35], len(rows), ordn, urg, f"${val:,.0f}" if val else "N/A"])
        rn = ws4.max_row
        for col in range(1,6):
            c = ws4.cell(rn,col)
            c.fill      = make_fill(fhex)
            c.border    = tb()
            c.alignment = Alignment(horizontal="left" if col==1 else "center", vertical="center")
        ws4.row_dimensions[rn].height = 14

    wb.move_sheet("📊 Summary", offset=-wb.index(wb["📊 Summary"]))

    output_file = f"po_plan_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"✅ Saved: {output_file}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
