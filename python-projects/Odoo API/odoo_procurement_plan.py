import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from math import ceil

# ================================================================
# PRODUCTION PLAN (units per month)
# ================================================================
MONTHLY_PLAN = {
    '101336': 150,   # C-100
    '101769': 20,    # C2 Bird Purple
    '101711': 40,    # C2 clearflo
    '102237': 20,    # C2 Yellow VTG
}
PLANNING_MONTHS = 3

# ================================================================
# HELPERS
# ================================================================
def make_fill(hex):    return PatternFill("solid", fgColor=hex)
def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(ws, row_num, num_cols, fill, bold=False, left_cols=None):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill      = fill
        c.border    = thin_border()
        c.font      = Font(bold=bold)
        c.alignment = Alignment(
            horizontal="left" if left_cols and col in left_cols else "center",
            vertical="center", wrap_text=True
        )

def write_section_title(ws, title, color, num_cols):
    ws.append([title])
    r = ws.max_row
    ws.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")
    ws.cell(row=r, column=1).font      = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=r, column=1).fill      = make_fill(color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

def write_headers(ws, headers, color):
    ws.append(headers)
    r = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.fill      = make_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[r].height = 35

MAIN_COLOR = "2E4057"

def get_month_label(base_now, offset):
    y = base_now.year + (base_now.month - 1 + offset) // 12
    m = (base_now.month - 1 + offset) % 12 + 1
    return f"{y}-{m:02d}"

def explode_bom(models, uid, bom_id, qty_needed, bom_cache, component_totals, depth=0):
    """Recursively explode BOM and accumulate component quantities."""
    if bom_id not in bom_cache:
        lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom.line', 'search_read',
            [[['bom_id', '=', bom_id]]],
            {'fields': ['product_id', 'product_qty', 'product_uom_id', 'child_bom_id']}
        )
        # Get parent BOM qty
        bom_info = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'read',
            [[bom_id]], {'fields': ['product_qty']}
        )
        bom_cache[bom_id] = {'lines': lines, 'bom_qty': bom_info[0]['product_qty'] if bom_info else 1}
    
    cached    = bom_cache[bom_id]
    bom_qty   = cached['bom_qty']
    lines     = cached['lines']
    ratio     = qty_needed / bom_qty

    for line in lines:
        comp_id   = line['product_id'][0]
        comp_name = line['product_id'][1]
        comp_qty  = line['product_qty'] * ratio
        uom       = line['product_uom_id'][1] if line.get('product_uom_id') else 'Unit'

        if line.get('child_bom_id'):
            # Sub-assembly — recurse
            explode_bom(models, uid, line['child_bom_id'][0], comp_qty,
                       bom_cache, component_totals, depth+1)
        else:
            # Raw component — accumulate
            if comp_id not in component_totals:
                component_totals[comp_id] = {'name': comp_name, 'qty': 0, 'uom': uom}
            component_totals[comp_id]['qty'] += comp_qty

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now = datetime.now(timezone.utc)
    months = [get_month_label(now, i) for i in range(PLANNING_MONTHS)]
    print(f"Planning months: {months}\n")

    # --- Resolve finished products & BOMs ---
    products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', list(MONTHLY_PLAN.keys())]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    prod_by_ref = {p['default_code']: p for p in products}

    # Find BOMs by template default_code
    bom_by_ref = {}
    for ref in MONTHLY_PLAN:
        boms = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.bom', 'search_read',
            [[['product_tmpl_id.default_code', '=', ref]]],
            {'fields': ['id', 'product_qty', 'product_tmpl_id']}
        )
        if boms:
            bom_by_ref[ref] = boms[0]
            print(f"  BOM found for {ref}: ID {boms[0]['id']} (qty {boms[0]['product_qty']})")
        else:
            print(f"  ⚠ No BOM found for {ref}")

    # ================================================================
    # EXPLODE BOMs — per month, per finished product
    # ================================================================
    print("\nExploding BOMs...")
    bom_cache = {}

    # Per-month component requirements: month -> {comp_id -> qty}
    monthly_req = {}
    for m in months:
        monthly_req[m] = defaultdict(float)

    # Also track which finished product each component comes from
    comp_by_finished = defaultdict(lambda: defaultdict(float))  # comp_id -> ref -> qty_per_unit

    for ref, plan_qty in MONTHLY_PLAN.items():
        if ref not in bom_by_ref:
            continue
        bom_id = bom_by_ref[ref]['id']
        comp_totals = {}
        explode_bom(models, uid, bom_id, plan_qty, bom_cache, comp_totals)
        print(f"  {ref} ({plan_qty}/mo): {len(comp_totals)} components")
        for comp_id, info in comp_totals.items():
            for m in months:
                monthly_req[m][comp_id] += info['qty']
            comp_by_finished[comp_id][ref] = info['qty'] / plan_qty  # qty per unit

    # All component IDs
    all_comp_ids = set()
    for m_req in monthly_req.values():
        all_comp_ids.update(m_req.keys())
    all_comp_ids = list(all_comp_ids)
    print(f"\n  Total unique components: {len(all_comp_ids)}")

    # --- Fetch component product details ---
    print("Fetching component details...")
    comp_products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['id', 'in', all_comp_ids]]],
        {'fields': ['id', 'name', 'default_code', 'product_tmpl_id']}
    )
    comp_info = {p['id']: p for p in comp_products}
    comp_tmpl_ids = [p['product_tmpl_id'][0] for p in comp_products]

    # --- Fetch supplier info (lead times) ---
    print("Fetching supplier lead times...")
    suppliers = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.supplierinfo', 'search_read',
        [[['product_tmpl_id', 'in', comp_tmpl_ids]]],
        {'fields': ['product_tmpl_id', 'product_id', 'name', 'delay', 'min_qty', 'price', 'product_code']}
    )
    # Best supplier per product template (lowest sequence = preferred)
    sup_by_tmpl = {}
    for s in suppliers:
        tmpl_id = s['product_tmpl_id'][0]
        if tmpl_id not in sup_by_tmpl:
            sup_by_tmpl[tmpl_id] = s
    print(f"  Supplier info found for {len(sup_by_tmpl)} components")

    # --- Fetch current stock ---
    print("Fetching current stock...")
    quants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.quant', 'search_read',
        [[['product_id', 'in', all_comp_ids],
          ['location_id.usage', '=', 'internal']]],
        {'fields': ['product_id', 'quantity', 'reserved_quantity']}
    )
    stock_by_comp = defaultdict(float)
    for q in quants:
        stock_by_comp[q['product_id'][0]] += max(0, q['quantity'] - q['reserved_quantity'])
    print(f"  Stock found for {len(stock_by_comp)} components")

    # --- Fetch open POs for components ---
    print("Fetching open PO commitments...")
    po_lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order.line', 'search_read',
        [[['order_id.state', 'in', ['purchase', 'draft', 'done']],
          ['product_id', 'in', all_comp_ids]]],
        {'fields': ['product_id', 'product_qty', 'qty_received', 'qty_invoiced',
                    'date_planned', 'order_id']}
    )
    po_by_comp = defaultdict(float)
    po_detail_by_comp = defaultdict(list)  # for debugging
    for l in po_lines:
        ordered  = l.get('product_qty', 0) or 0
        received = l.get('qty_received', 0) or 0
        remaining = ordered - received
        if remaining > 0:
            po_by_comp[l['product_id'][0]] += remaining
            po_detail_by_comp[l['product_id'][0]].append({
                'po': l['order_id'][1] if l.get('order_id') else '?',
                'ordered': ordered, 'received': received, 'remaining': remaining,
                'date': l.get('date_planned', '')
            })
    print(f"  Open PO commitments for {len(po_by_comp)} components")
    print(f"  Total PO lines fetched: {len(po_lines)}")

    # Debug: check part 100318 specifically
    debug_prod = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', '=', '100318']]], {'fields': ['id', 'name', 'default_code']}
    )
    if debug_prod:
        debug_id = debug_prod[0]['id']
        debug_po = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order.line', 'search_read',
            [[['product_id', '=', debug_id]]],
            {'fields': ['product_id', 'product_qty', 'qty_received', 'order_id', 'date_planned'],
             'limit': 10}
        )
        print(f"\n  DEBUG 100318 (id={debug_id}) — all PO lines ({len(debug_po)} found):")
        for l in debug_po:
            remaining = (l.get('product_qty') or 0) - (l.get('qty_received') or 0)
            print(f"    PO: {l['order_id'][1] if l.get('order_id') else '?'} | "
                  f"Ordered: {l['product_qty']} | Received: {l['qty_received']} | "
                  f"Remaining: {remaining} | State check needed")
        # Check PO states
        po_ids = [l['order_id'][0] for l in debug_po if l.get('order_id')]
        if po_ids:
            pos = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order', 'read',
                [list(set(po_ids))], {'fields': ['name', 'state']}
            )
            print(f"    PO states: {[(p['name'], p['state']) for p in pos]}")

    # ================================================================
    # BUILD PROCUREMENT PLAN
    # ================================================================
    # For each component:
    #   gross_req_m1 = monthly_req[months[0]][comp_id]
    #   available    = stock + open_POs
    #   net_req_m1   = max(0, gross_req_m1 - available)
    #   order_by     = start of month1 - lead_time_days  (order date)
    #   carry_forward stock after m1 for m2, etc.

    plan_rows = []
    for comp_id in sorted(all_comp_ids, key=lambda x: comp_info.get(x, {}).get('default_code', '') or ''):
        info     = comp_info.get(comp_id, {})
        tmpl_id  = info.get('product_tmpl_id', [None])[0] if info else None
        sup      = sup_by_tmpl.get(tmpl_id, {})
        lead_days = sup.get('delay', 0) or 0
        min_qty   = sup.get('min_qty', 0) or 0
        sup_name  = sup['name'][1] if sup.get('name') else 'No supplier'
        sup_code  = sup.get('product_code', '') or ''
        price     = sup.get('price', 0) or 0

        stock     = stock_by_comp.get(comp_id, 0)
        open_po   = po_by_comp.get(comp_id, 0)
        available = stock + open_po

        # Calculate net requirements per month with rolling stock
        rolling_stock = available
        month_data = []
        for m in months:
            gross = monthly_req[m].get(comp_id, 0)
            net   = max(0, gross - rolling_stock)
            # Round up to min order qty
            order_qty = max(net, min_qty) if net > 0 else 0
            if order_qty > 0 and min_qty > 0:
                order_qty = ceil(order_qty / min_qty) * min_qty
            # Order must be placed lead_time days before month start
            m_year, m_month = int(m[:4]), int(m[5:])
            month_start = datetime(m_year, m_month, 1, tzinfo=timezone.utc)
            order_by    = month_start - timedelta(days=lead_days)
            order_urgent = order_by <= now

            month_data.append({
                'month':      m,
                'gross':      gross,
                'net':        net,
                'order_qty':  order_qty,
                'order_by':   order_by.strftime('%Y-%m-%d'),
                'urgent':     order_urgent,
            })
            # Update rolling stock for next month
            rolling_stock = max(0, rolling_stock - gross) + 0  # no replenishment assumed yet

        # Which finished products use this component
        used_by = ', '.join(
            f"{ref}×{qty:.2f}".rstrip('0').rstrip('.')
            for ref, qty in comp_by_finished.get(comp_id, {}).items()
        )

        plan_rows.append({
            'comp_id':    comp_id,
            'ref':        info.get('default_code', '') or '',
            'name':       info.get('name', '') or '',
            'supplier':   sup_name,
            'sup_code':   sup_code,
            'lead_days':  lead_days,
            'min_qty':    min_qty,
            'price':      price,
            'stock':      stock,
            'open_po':    open_po,
            'available':  available,
            'used_by':    used_by,
            'months':     month_data,
            'needs_order': any(d['order_qty'] > 0 for d in month_data),
            'urgent':     any(d['urgent'] and d['order_qty'] > 0 for d in month_data),
        })

    urgent_rows  = [r for r in plan_rows if r['urgent']]
    order_rows   = [r for r in plan_rows if r['needs_order'] and not r['urgent']]
    ok_rows      = [r for r in plan_rows if not r['needs_order']]
    print(f"\n  🔴 Urgent orders needed: {len(urgent_rows)}")
    print(f"  🟡 Orders needed soon:   {len(order_rows)}")
    print(f"  🟢 Stock sufficient:     {len(ok_rows)}")

    # ================================================================
    # BUILD EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    NUM_COLS = 11 + len(months)

    def set_col_widths(ws):
        ws.column_dimensions["A"].width = 12  # ref
        ws.column_dimensions["B"].width = 40  # name
        ws.column_dimensions["C"].width = 22  # supplier
        ws.column_dimensions["D"].width = 14  # sup code
        ws.column_dimensions["E"].width = 10  # lead days
        ws.column_dimensions["F"].width = 10  # min qty
        ws.column_dimensions["G"].width = 10  # stock
        ws.column_dimensions["H"].width = 10  # open PO
        ws.column_dimensions["I"].width = 10  # available
        for i, m in enumerate(months):
            ws.column_dimensions[get_column_letter(10 + i*3)].width = 10   # gross
            ws.column_dimensions[get_column_letter(11 + i*3)].width = 10   # net
            ws.column_dimensions[get_column_letter(12 + i*3)].width = 12   # order qty
        ws.column_dimensions[get_column_letter(10 + len(months)*3)].width = 18  # used by

    def add_header_row(ws, title):
        # Row 1 - title (only merge up to actual used cols to avoid bleed)
        ws.cell(row=1, column=1, value=title)
        try:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        except Exception:
            pass
        ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
        ws["A1"].fill      = make_fill(MAIN_COLOR)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Row 2 - subtitle
        ws.cell(row=2, column=1,
                value=(f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   "
                       f"Plan: C-100×150, C2 clearflo×40, C2 Bird×20, C2 Yellow×20 per month   |   "
                       f"Horizon: {months[0]} → {months[-1]}"))
        try:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
        except Exception:
            pass
        ws["A2"].font      = Font(italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Row 3 - month group headers (write directly by row/col, no append)
        r = 3
        base_headers = ["Ref", "Component Name", "Supplier", "Sup. Code",
                         "Lead (days)", "Min Qty", "Stock", "Open PO", "Available"]
        for i, h in enumerate(base_headers, 1):
            ws.cell(row=r, column=i, value=h).font      = Font(color="FFFFFF", bold=True)
            ws.cell(row=r, column=i).fill      = make_fill(MAIN_COLOR)
            ws.cell(row=r, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=r, column=i).border    = thin_border()

        for i, m in enumerate(months):
            col_start = 10 + i * 3
            # Merge month label across 3 cols
            ws.merge_cells(f"{get_column_letter(col_start)}{r}:{get_column_letter(col_start+2)}{r}")
            c = ws.cell(row=r, column=col_start)
            m_color = ["1F4E79", "375623", "7B2C2C"][i % 3]
            c.value     = m
            c.font      = Font(color="FFFFFF", bold=True)
            c.fill      = make_fill(m_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()

        last_col = 10 + len(months) * 3
        ws.cell(row=r, column=last_col, value="Used By (qty/unit)").font = Font(color="FFFFFF", bold=True)
        ws.cell(row=r, column=last_col).fill      = make_fill(MAIN_COLOR)
        ws.cell(row=r, column=last_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=r, column=last_col).border    = thin_border()
        ws.row_dimensions[r].height = 30

        # Row 4 - sub-headers
        r2 = 4
        for col in range(1, 10):
            ws.cell(row=r2, column=col).fill   = make_fill(MAIN_COLOR)
            ws.cell(row=r2, column=col).border = thin_border()
        for i, m in enumerate(months):
            col_start = 10 + i * 3
            m_color = ["1F4E79", "375623", "7B2C2C"][i % 3]
            for j, sub in enumerate(["Gross Req", "Net Req", "Order Qty"]):
                col = col_start + j
                ws.cell(row=r2, column=col, value=sub).font      = Font(color="FFFFFF", bold=True, size=9)
                ws.cell(row=r2, column=col).fill      = make_fill(m_color)
                ws.cell(row=r2, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(row=r2, column=col).border    = thin_border()
        ws.cell(row=r2, column=10 + len(months) * 3).fill   = make_fill(MAIN_COLOR)
        ws.cell(row=r2, column=10 + len(months) * 3).border = thin_border()
        ws.row_dimensions[r2].height = 28
        ws.freeze_panes = "A5"

    def add_data_row(ws, row, row_num):
        fill_color = "FFB3B3" if row['urgent'] else "FFF2CC" if row['needs_order'] else "E2EFDA"
        row_fill = make_fill(fill_color)
        data = [
            row['ref'], row['name'][:38], row['supplier'][:20],
            row['sup_code'], row['lead_days'], row['min_qty'],
            round(row['stock'], 2), round(row['open_po'], 2), round(row['available'], 2)
        ]
        for md in row['months']:
            data += [round(md['gross'], 2), round(md['net'], 2),
                     round(md['order_qty'], 2) if md['order_qty'] else '']
        data.append(row['used_by'])

        ws.append(data)
        r = ws.max_row
        for col in range(1, len(data) + 1):
            c = ws.cell(row=r, column=col)
            c.fill      = row_fill
            c.border    = thin_border()
            c.alignment = Alignment(
                horizontal="left" if col in [1, 2, 3, 4] else "center",
                vertical="center", wrap_text=False
            )
        ws.row_dimensions[r].height = 15

    # ================================================================
    # SHEET 1 — URGENT (order now)
    # ================================================================
    ws1 = wb.create_sheet("🔴 Order Now")
    set_col_widths(ws1)
    add_header_row(ws1, f"🔴 Components — ORDER NOW ({len(urgent_rows)} items)")
    ws1["A1"].fill = make_fill("C00000")
    for row in sorted(urgent_rows, key=lambda x: x['lead_days'], reverse=True):
        add_data_row(ws1, row, ws1.max_row + 1)
    ws1.append([])
    ws1.append(["", f"⚠ These {len(urgent_rows)} components need to be ordered immediately based on lead time and monthly requirements."])
    ws1.cell(row=ws1.max_row, column=2).font = Font(bold=True, color="C00000")

    # ================================================================
    # SHEET 2 — PLAN AHEAD
    # ================================================================
    ws2 = wb.create_sheet("🟡 Plan Ahead")
    set_col_widths(ws2)
    add_header_row(ws2, f"🟡 Components — Order Required Within Planning Horizon ({len(order_rows)} items)")
    ws2["A1"].fill = make_fill("B8860B")
    for row in sorted(order_rows, key=lambda x: x['months'][0]['order_by']):
        add_data_row(ws2, row, ws2.max_row + 1)

    # ================================================================
    # SHEET 3 — FULL PLAN
    # ================================================================
    ws3 = wb.create_sheet("📋 Full Plan")
    set_col_widths(ws3)
    add_header_row(ws3, f"📋 Full Component Procurement Plan — {len(plan_rows)} components")
    # Section: urgent
    if urgent_rows:
        write_section_title(ws3, f"  🔴  ORDER NOW — {len(urgent_rows)} components", "C00000", NUM_COLS)
        for row in sorted(urgent_rows, key=lambda x: x['lead_days'], reverse=True):
            add_data_row(ws3, row, ws3.max_row + 1)
    # Section: order soon
    if order_rows:
        write_section_title(ws3, f"  🟡  ORDER WITHIN HORIZON — {len(order_rows)} components", "B8860B", NUM_COLS)
        for row in sorted(order_rows, key=lambda x: x['months'][0]['order_by']):
            add_data_row(ws3, row, ws3.max_row + 1)
    # Section: ok
    if ok_rows:
        write_section_title(ws3, f"  🟢  STOCK SUFFICIENT — {len(ok_rows)} components", "375623", NUM_COLS)
        for row in sorted(ok_rows, key=lambda x: x['ref']):
            add_data_row(ws3, row, ws3.max_row + 1)

    # ================================================================
    # SHEET 4 — SUMMARY BY SUPPLIER
    # ================================================================
    ws4 = wb.create_sheet("By Supplier")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 12
    for i in range(4, 8):
        ws4.column_dimensions[get_column_letter(i)].width = 14

    ws4.append(["Procurement Plan — Summary by Supplier"])
    ws4.merge_cells(f"A1:G1")
    ws4["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws4["A1"].fill      = make_fill(MAIN_COLOR)
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24
    ws4.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC"])
    ws4.merge_cells("A2:G2")
    ws4["A2"].font = Font(italic=True)
    ws4["A2"].alignment = Alignment(horizontal="center")
    ws4.append([])

    # Group by supplier
    by_supplier = defaultdict(list)
    for row in plan_rows:
        by_supplier[row['supplier']].append(row)

    write_headers(ws4, ["Supplier", "# Parts", "Urgent Items",
                         f"{months[0]} Order", f"{months[1]} Order" if len(months) > 1 else "",
                         f"{months[2]} Order" if len(months) > 2 else "",
                         "Est. Total Value"], MAIN_COLOR)

    for sup_name in sorted(by_supplier.keys()):
        rows = by_supplier[sup_name]
        urgent_cnt = sum(1 for r in rows if r['urgent'])
        month_orders = []
        for mi in range(len(months)):
            total_order = sum(r['months'][mi]['order_qty'] for r in rows)
            month_orders.append(round(total_order, 0))
        est_value = sum(
            r['months'][mi]['order_qty'] * r['price']
            for r in rows for mi in range(len(months))
            if r['price'] and r['months'][mi]['order_qty']
        )
        fill_color = "FFB3B3" if urgent_cnt > 0 else "DDEBF7"
        ws4.append([sup_name, len(rows), urgent_cnt] + month_orders[:3] + [f"${est_value:,.2f}" if est_value else "N/A"])
        style_row(ws4, ws4.max_row, 7, make_fill(fill_color), left_cols=[1])

    # Save
    output_file = f"procurement_plan_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\nExcel saved: {output_file}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
