import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from calendar import monthrange

PLANNING_MONTHS = 3   # current month + next 2
PLAN_PRODUCTS   = cfg.C2_PRODUCT_REFS + ['101336']

# State labels
MO_STATE_MAP = {
    'draft':    'Draft',
    'confirmed':'Confirmed',
    'progress': 'In Progress',
    'done':     'Done',
    'cancel':   'Cancelled',
}
ACTIVE_STATES = ['draft', 'confirmed', 'progress']

# ================================================================
# HELPERS
# ================================================================
def make_fill(hex):    return PatternFill("solid", fgColor=hex)
def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(ws, row_num, num_cols, fill, bold=False, left_cols=None):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill   = fill
        c.border = thin_border()
        c.font   = Font(bold=bold)
        c.alignment = Alignment(
            horizontal="left" if left_cols and col in left_cols else "center",
            vertical="center", wrap_text=True
        )

def write_section_title(ws, title, color, num_cols):
    ws.append([title])
    r = ws.max_row
    ws.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")
    ws.cell(row=r, column=1).font      = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=r, column=1).fill      = make_fill(color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

def write_headers(ws, headers, color):
    ws.append(headers)
    r = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.fill      = make_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[r].height = 30

MAIN_COLOR   = "2E4057"
OVERDUE_COLOR = "C00000"
WARN_COLOR    = "FF9900"

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now = datetime.now(timezone.utc)

    # --- Build month buckets ---
    months = []
    for i in range(PLANNING_MONTHS):
        y = now.year + (now.month - 1 + i) // 12
        m = (now.month - 1 + i) % 12 + 1
        months.append((y, m))
    month_labels = [f"{y}-{m:02d}" for y, m in months]
    print(f"Planning months: {month_labels}\n")

    # --- Resolve product IDs ---
    products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', PLAN_PRODUCTS]]],
        {'fields': ['id', 'name', 'default_code']}
    )
    prod_map   = {p['id']: p for p in products}  # id -> product info
    prod_ids   = list(prod_map.keys())
    prod_by_ref = {p['default_code']: p for p in products}
    print(f"Products found: {[(p['default_code'], p['name']) for p in products]}\n")

    # Device line helper
    c2_ids = {p['id'] for p in products if p['default_code'] in cfg.C2_PRODUCT_REFS}
    def device_line(prod_id):
        return 'C2' if prod_id in c2_ids else 'C-100'

    # --- Fetch ALL open MOs ---
    print("Fetching open manufacturing orders...")
    mos = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'mrp.production', 'search_read',
        [[['product_id', 'in', prod_ids],
          ['state', 'in', ACTIVE_STATES]]],
        {'fields': ['name', 'product_id', 'product_qty', 'date_planned_start',
                    'state', 'origin']}
    )
    print(f"  → {len(mos)} open MOs found\n")

    # Annotate MOs
    for mo in mos:
        dp = mo.get('date_planned_start', '')
        mo['_dt']     = datetime.strptime(dp[:19], '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc) if dp else None
        mo['_month']  = f"{mo['_dt'].year}-{mo['_dt'].month:02d}" if mo['_dt'] else 'Unknown'
        mo['_device'] = device_line(mo['product_id'][0])
        mo['_state']  = MO_STATE_MAP.get(mo['state'], mo['state'])
        mo['_overdue'] = mo['_dt'] and mo['_dt'] < now and mo['state'] in ['draft', 'confirmed']
        mo['_age_days'] = (now - mo['_dt']).days if mo['_dt'] else 0

    # Categorize
    overdue_mos  = [m for m in mos if m['_overdue']]
    future_mos   = [m for m in mos if not m['_overdue']]

    # --- Fetch open POs ---
    print("Fetching open purchase orders...")
    po_lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'purchase.order.line', 'search_read',
        [[['order_id.state', 'in', ['purchase', 'draft']],
          ['product_id', 'in', prod_ids],
          ['qty_received', '<', 1]]],   # only lines with outstanding qty
        {'fields': ['product_id', 'product_qty', 'qty_received',
                    'date_planned', 'order_id', 'price_unit']}
    )
    # Filter to actually outstanding
    po_lines = [l for l in po_lines if l['product_qty'] - l['qty_received'] > 0]
    print(f"  → {len(po_lines)} outstanding PO lines\n")

    for l in po_lines:
        dp = l.get('date_planned', '')
        l['_dt']      = datetime.strptime(dp[:19], '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc) if dp else None
        l['_month']   = f"{l['_dt'].year}-{l['_dt'].month:02d}" if l['_dt'] else 'Unknown'
        l['_remaining'] = l['product_qty'] - l['qty_received']
        l['_device']  = device_line(l['product_id'][0])
        l['_overdue'] = l['_dt'] and l['_dt'] < now

    # --- Fetch current stock (all internal locations) ---
    print("Fetching current stock...")
    quants = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'stock.quant', 'search_read',
        [[['product_id', 'in', prod_ids],
          ['location_id.usage', '=', 'internal']]],
        {'fields': ['product_id', 'quantity', 'reserved_quantity', 'location_id']}
    )
    stock_by_prod = defaultdict(float)
    for q in quants:
        avail = q['quantity'] - q['reserved_quantity']
        if avail > 0:
            stock_by_prod[q['product_id'][0]] += avail
    print(f"  → Stock found for {len(stock_by_prod)} products\n")

    # --- Fetch open SOs (demand) ---
    print("Fetching open sales orders...")
    so_lines = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'sale.order.line', 'search_read',
        [[['order_id.state', 'in', ['sale']],
          ['product_id', 'in', prod_ids],
          ['qty_delivered', '<', 1]]],
        {'fields': ['product_id', 'product_uom_qty', 'qty_delivered',
                    'order_id', 'customer_lead']}
    )
    so_lines = [l for l in so_lines if l['product_uom_qty'] - l['qty_delivered'] > 0]
    print(f"  → {len(so_lines)} open SO lines\n")

    so_by_prod = defaultdict(float)
    for l in so_lines:
        so_by_prod[l['product_id'][0]] += l['product_uom_qty'] - l['qty_delivered']

    # ================================================================
    # BUILD EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ================================================================
    # SHEET 1 — EXECUTIVE DASHBOARD
    # ================================================================
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 45
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.append(["Production Planning Dashboard — Montreal"])
    ws.merge_cells("A1:K1")
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = make_fill(MAIN_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    ws["A2"] = (f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   "
                f"Planning Horizon: {month_labels[0]} → {month_labels[-1]}   |   "
                f"Open MOs: {len(mos)}   |   Overdue MOs: {len(overdue_mos)}")
    ws["A2"].font      = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])

    # --- TABLE 1: Summary by Product ---
    write_section_title(ws, "  📦  Production Summary by Product", MAIN_COLOR, 8)
    write_headers(ws, ["Product", "Ref", "Device", "Stock Available",
                        "Open SO Demand", "Open MOs", "Overdue MOs", "Net Position"], MAIN_COLOR)
    for p in sorted(products, key=lambda x: x['default_code']):
        pid    = p['id']
        stock  = stock_by_prod.get(pid, 0)
        demand = so_by_prod.get(pid, 0)
        open_mo_cnt   = sum(1 for m in mos if m['product_id'][0] == pid)
        overdue_cnt   = sum(1 for m in overdue_mos if m['product_id'][0] == pid)
        net    = stock + open_mo_cnt - demand
        short_name = p['name'][:42]

        if net < 0:         row_fill = make_fill("FFB3B3")
        elif overdue_cnt > 0: row_fill = make_fill("FFD9B3")
        else:               row_fill = make_fill("E2EFDA")

        ws.append([short_name, p['default_code'], device_line(pid),
                   stock, demand, open_mo_cnt, overdue_cnt,
                   f"{'+' if net >= 0 else ''}{net:.0f}"])
        style_row(ws, ws.max_row, 8, row_fill, left_cols=[1, 2, 3])
    ws.append([])

    # --- TABLE 2: MO Pipeline by Month ---
    write_section_title(ws, "  🗓  MO Pipeline — Next 3 Months", MAIN_COLOR, 8)
    # Build columns: Product | Ref | Device | Month1 | Month2 | Month3 | Overdue | Total
    write_headers(ws, ["Product", "Ref", "Device"] + month_labels + ["Overdue", "Total Open"], MAIN_COLOR)
    for p in sorted(products, key=lambda x: x['default_code']):
        pid = p['id']
        month_counts = []
        for ym in month_labels:
            cnt = sum(1 for m in mos if m['product_id'][0] == pid and m['_month'] == ym)
            month_counts.append(cnt)
        overdue_cnt = sum(1 for m in overdue_mos if m['product_id'][0] == pid)
        total       = sum(1 for m in mos if m['product_id'][0] == pid)
        row_fill = make_fill("FFD9B3") if overdue_cnt > 0 else make_fill("DDEBF7")
        ws.append([p['name'][:42], p['default_code'], device_line(pid)]
                  + month_counts + [overdue_cnt, total])
        style_row(ws, ws.max_row, 8, row_fill, left_cols=[1, 2, 3])
    ws.append([])

    # --- TABLE 3: Outstanding PO Commitments ---
    write_section_title(ws, "  🚚  Outstanding Purchase Order Commitments", "375623", 7)
    write_headers(ws, ["Product", "Ref", "Device", "Remaining Qty", "Expected Date", "Overdue?", "PO #"], "375623")
    po_sorted = sorted(po_lines, key=lambda x: x['_dt'] or datetime.max.replace(tzinfo=timezone.utc))
    for l in po_sorted:
        overdue_str = "⚠ OVERDUE" if l['_overdue'] else "On Track"
        row_fill = make_fill("FFB3B3") if l['_overdue'] else make_fill("E2EFDA")
        date_str = l['_dt'].strftime('%Y-%m-%d') if l['_dt'] else 'Unknown'
        ws.append([
            l['product_id'][1][:42],
            prod_map.get(l['product_id'][0], {}).get('default_code', ''),
            l['_device'],
            l['_remaining'],
            date_str,
            overdue_str,
            l['order_id'][1] if l.get('order_id') else ''
        ])
        style_row(ws, ws.max_row, 7, row_fill, left_cols=[1, 2, 3, 5, 6, 7])
    ws.append([])

    # ================================================================
    # SHEET 2 — OVERDUE MOs
    # ================================================================
    ws2 = wb.create_sheet("⚠ Overdue MOs")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 45
    for i in range(3, 8):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    ws2.append([f"Overdue Manufacturing Orders — {len(overdue_mos)} total"])
    ws2.merge_cells("A1:G1")
    ws2["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill      = make_fill(OVERDUE_COLOR)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22
    ws2.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   Draft/Confirmed MOs with planned date in the past"])
    ws2.merge_cells("A2:G2")
    ws2["A2"].font = Font(italic=True)
    ws2["A2"].alignment = Alignment(horizontal="center")
    ws2.append([])

    write_headers(ws2, ["MO #", "Product", "Device", "State", "Planned Date", "Days Overdue", "Origin"], OVERDUE_COLOR)
    ws2.freeze_panes = "A4"
    for mo in sorted(overdue_mos, key=lambda x: x['_age_days'], reverse=True):
        days = mo['_age_days']
        row_fill = make_fill("FFB3B3") if days > 60 else make_fill("FFD9B3") if days > 30 else make_fill("FFF2CC")
        ws2.append([
            mo['name'],
            mo['product_id'][1][:42],
            mo['_device'],
            mo['_state'],
            mo['_dt'].strftime('%Y-%m-%d') if mo['_dt'] else '',
            days,
            mo.get('origin', '') or ''
        ])
        style_row(ws2, ws2.max_row, 7, row_fill, left_cols=[1, 2, 3, 4, 7])

    # ================================================================
    # SHEET 3 — FULL MO LIST
    # ================================================================
    ws3 = wb.create_sheet("All Open MOs")
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 45
    for i in range(3, 9):
        ws3.column_dimensions[get_column_letter(i)].width = 16

    ws3.append(["All Open Manufacturing Orders"])
    ws3.merge_cells("A1:H1")
    ws3["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill      = make_fill(MAIN_COLOR)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22
    ws3.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   {len(mos)} open MOs"])
    ws3.merge_cells("A2:H2")
    ws3["A2"].font = Font(italic=True)
    ws3["A2"].alignment = Alignment(horizontal="center")
    ws3.append([])

    write_headers(ws3, ["MO #", "Product", "Device", "Ref", "State", "Planned Date", "Month", "Origin"], MAIN_COLOR)
    ws3.freeze_panes = "A4"

    for mo in sorted(mos, key=lambda x: x['_dt'] or datetime.max.replace(tzinfo=timezone.utc)):
        days = mo['_age_days']
        if mo['_overdue']:
            row_fill = make_fill("FFD9B3")
        elif mo['_month'] == month_labels[0]:
            row_fill = make_fill("E2EFDA")
        elif mo['_month'] == month_labels[1] if len(month_labels) > 1 else '':
            row_fill = make_fill("DDEBF7")
        else:
            row_fill = make_fill("F2F2F2")

        ref = prod_map.get(mo['product_id'][0], {}).get('default_code', '')
        ws3.append([
            mo['name'],
            mo['product_id'][1][:42],
            mo['_device'],
            ref,
            mo['_state'],
            mo['_dt'].strftime('%Y-%m-%d') if mo['_dt'] else '',
            mo['_month'],
            mo.get('origin', '') or ''
        ])
        style_row(ws3, ws3.max_row, 8, row_fill, left_cols=[1, 2, 3, 4, 5, 7, 8])

    # Save
    output_file = f"production_plan_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\nExcel report saved: {output_file}")

    # Summary print
    print("\n=== SUMMARY ===")
    for p in sorted(products, key=lambda x: x['default_code']):
        pid = p['id']
        stock = stock_by_prod.get(pid, 0)
        demand = so_by_prod.get(pid, 0)
        mo_cnt = sum(1 for m in mos if m['product_id'][0] == pid)
        ov_cnt = sum(1 for m in overdue_mos if m['product_id'][0] == pid)
        print(f"  {p['default_code']:<10} | Stock: {stock:>4.0f} | Demand: {demand:>4.0f} | "
              f"Open MOs: {mo_cnt:>3} | Overdue: {ov_cnt:>3}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
