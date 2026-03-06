import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from collections import Counter
from datetime import datetime, timezone, timedelta

# ================================================================
# STATE MAPPING — only these states are included
# ================================================================
STATE_MAP = {
    'draft':        'RMA Created',
    'confirmed':    'Device Received',
    'under_repair': 'Under Repair',
}
STATE_ORDER  = ['RMA Created', 'Device Received', 'Under Repair']
STATE_COLORS = {
    'RMA Created':    'FFF2CC',
    'Device Received':'DDEBF7',
    'Under Repair':   'FFD9B3',
}

TOP_N     = 5
MIN_TREND = 3

# ================================================================
# HELPERS
# ================================================================
def make_fill(hex):    return PatternFill("solid", fgColor=hex)
def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(ws, row_num, num_cols, fill, left_cols=None):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill   = fill
        c.border = thin_border()
        c.alignment = Alignment(
            horizontal="left" if left_cols and col in left_cols else "center",
            vertical="center"
        )

def write_section_title(ws, title, color, num_cols):
    ws.append([title])
    r = ws.max_row
    ws.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")
    ws.cell(row=r, column=1).font      = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=r, column=1).fill      = make_fill(color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

def write_headers(ws, headers, color):
    ws.append(headers)
    r = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.fill      = make_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

def make_bar(pct, width=20):
    filled = round(pct / 100 * width)
    return "█" * filled + "░" * (width - filled)

def count_tags(repair_list):
    c = Counter()
    for r in repair_list:
        for tid in r.get('_tag_names', []):
            c[tid] += 1
    return c

def sl(repairs, division=None, device=None, state=None):
    r = repairs
    if division: r = [x for x in r if x['_division'] == division]
    if device:   r = [x for x in r if x['_device']   == device]
    if state:    r = [x for x in r if x['_state_label'] == state]
    return r

rank_fills = ["FFD700", "C0C0C0", "CD7F32", "DDEBF7", "EBF3FB"]

# Track next available column in hidden data sheet
_chart_col = [1]

def add_pie_chart(ws_data, wb, title, rows, anchor, slice_colors=None):
    """Add a pie chart using separate columns per chart to avoid overlap."""
    col = _chart_col[0]
    _chart_col[0] += 2  # use 2 columns per chart (label + value), skip 1 for spacing

    # Write header + data into this column pair
    ws_data.cell(row=1, column=col,   value="Label")
    ws_data.cell(row=1, column=col+1, value="Count")
    for i, (label, val) in enumerate(rows, start=2):
        ws_data.cell(row=i, column=col,   value=label)
        ws_data.cell(row=i, column=col+1, value=val)

    num_rows = len(rows)

    chart = PieChart()
    chart.title  = title
    chart.style  = 10
    chart.width  = 16
    chart.height = 12

    data_ref   = Reference(ws_data, min_col=col+1, min_row=1, max_row=num_rows+1)
    labels_ref = Reference(ws_data, min_col=col,   min_row=2, max_row=num_rows+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)

    chart.dataLabels                 = DataLabelList()
    chart.dataLabels.showPercent     = True
    chart.dataLabels.showCatName     = True
    chart.dataLabels.showVal         = True
    chart.dataLabels.showSerName     = False
    chart.dataLabels.showLeaderLines = True

    colors = slice_colors or ["1F4E79", "375623", "7B2C2C", "FFC000", "5B9BD5", "ED7D31"]
    for i, color in enumerate(colors[:len(rows)]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        chart.series[0].dPt.append(pt)

    return chart, anchor
bucket_colors = {
    '0-30 days':  'E2EFDA',
    '31-60 days': 'FFF2CC',
    '61-90 days': 'FFD9B3',
    '90+ days':   'FFB3B3'
}
bucket_order = ['0-30 days', '31-60 days', '61-90 days', '90+ days']

division_colors = {
    "THORASYS-Americas": "1F4E79",
    "THORASYS-EMEA":     "375623",
    "THORASYS-APAC":     "7B2C2C",
}
device_colors = {
    "C2":    "4A235A",
    "C-100": "1A5276",
}
MAIN_COLOR = "2E4057"

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now = datetime.now(timezone.utc)
    d30 = now - timedelta(days=30)
    d60 = now - timedelta(days=60)
    d90 = now - timedelta(days=90)

    # --- Teams ---
    all_teams = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.team', 'search_read',
        [[]], {'fields': ['id', 'name']}
    )
    team_map = {t['name']: t['id'] for t in all_teams}

    # --- C2 product IDs ---
    c2_products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', cfg.C2_PRODUCT_REFS]]],
        {'fields': ['id']}
    )
    c2_product_ids = {p['id'] for p in c2_products}

    # --- Repair tags ---
    repair_tags = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'repair.tags', 'search_read',
        [[]], {'fields': ['id', 'name']}
    )
    repair_tag_map = {t['id']: t['name'] for t in repair_tags}

    # Also get x_studio_repair_action_tags model name
    action_tags = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'repair.order', 'search_read',
        [[['state', 'in', list(STATE_MAP.keys())]]], 
        {'fields': ['id', 'x_studio_repair_action_tags'], 'limit': 5}
    )
    print(f"Sample action tags field: {[r['x_studio_repair_action_tags'] for r in action_tags[:2]]}")

    # --- Resolve excluded customer ID ---
    # Search each excluded customer name with ilike for flexible matching
    excluded_customer_ids = []
    for cust_name in cfg.REPAIR_EXCLUDED_CUSTOMERS:
        matches = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'res.partner', 'search_read',
            [[['name', '=', cust_name]]],
            {'fields': ['id', 'name']}
        )
        for m in matches:
            print(f"  Excluding customer: {m['name']} (ID: {m['id']})")
            excluded_customer_ids.append(m['id'])
    excluded_partner_ids = excluded_customer_ids
    print(f"  Total excluded partner IDs: {excluded_partner_ids}")

    # --- Fetch active repairs (only included states, excluding ERT) ---
    print("Fetching repairs...")
    repairs_raw = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'repair.order', 'search_read',
        [[['state', 'in', list(STATE_MAP.keys())],
          ['partner_id', 'not in', excluded_partner_ids]]],
        {'fields': [
            'id', 'name', 'state', 'product_id', 'lot_id', 'partner_id',
            'create_date', 'user_id', 'division_id', 'ticket_id',
            'tag_ids', 'x_studio_repair_action_tags', 'x_studio_repair_action_tags_char',
            'x_studio_repair_tags_char', 'x_studio_reason_for_return',
            'x_studio_issue_reproduced', 'x_studio_under_warranty_ts_case',
            'guarantee_limit', 'x_studio_incoming_tracking_',
            'x_studio_outgoing_tracking', 'location_id'
        ]}
    )
    print(f"  → {len(repairs_raw)} active repairs found\n")

    # --- Resolve excluded tag IDs ---
    excluded_tags = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'repair.tags', 'search_read',
        [[['name', 'ilike', 'Refurbishment']]],
        {'fields': ['id', 'name']}
    )
    excluded_tag_ids = {t['id'] for t in excluded_tags}
    print(f"Excluding tag(s): {[t['name'] for t in excluded_tags]}")

    # Remove repairs that have the excluded tag
    repairs_raw = [r for r in repairs_raw
                   if not any(tid in excluded_tag_ids for tid in r.get('tag_ids', []))]
    print(f"  → {len(repairs_raw)} repairs after tag exclusion\n")

    # --- Annotate repairs ---
    for r in repairs_raw:
        cd = r.get('create_date', '')
        r['_created']     = datetime.strptime(cd, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc) if cd else None
        r['_state_label'] = STATE_MAP.get(r['state'], r['state'])
        r['_division']    = r['division_id'][1] if r.get('division_id') else 'No Division'
        prod_id           = r['product_id'][0] if r.get('product_id') else None
        r['_device']      = 'C2' if prod_id in c2_product_ids else 'C-100'

        # Combine both tag fields
        tag_names = []
        for tid in r.get('tag_ids', []):
            name = repair_tag_map.get(tid, f'Tag({tid})')
            tag_names.append(name)
        # Also parse action tags char field as fallback
        action_char = r.get('x_studio_repair_action_tags_char', '') or ''
        if action_char and not tag_names:
            tag_names = [t.strip() for t in action_char.split(',') if t.strip()]
        r['_tag_names'] = tag_names

        # Age bucket
        if r['_created']:
            age = (now - r['_created']).days
            r['_age'] = age
            if age <= 30:   r['_bucket'] = '0-30 days'
            elif age <= 60: r['_bucket'] = '31-60 days'
            elif age <= 90: r['_bucket'] = '61-90 days'
            else:           r['_bucket'] = '90+ days'
        else:
            r['_age']    = 0
            r['_bucket'] = 'Unknown'

    total_active = len(repairs_raw)
    repairs_30   = [r for r in repairs_raw if r['_created'] and r['_created'] >= d30]
    repairs_90   = [r for r in repairs_raw if r['_created'] and r['_created'] >= d90]
    repairs_p30  = [r for r in repairs_raw if r['_created'] and d60 <= r['_created'] < d30]

    # --- Trending ---
    tags_30d    = count_tags(repairs_30)
    tags_prev30 = count_tags(repairs_p30)
    trending = {}
    for tag, cnt_now in tags_30d.items():
        if cnt_now < MIN_TREND: continue
        cnt_prev = tags_prev30.get(tag, 0)
        if cnt_prev == 0: continue
        pct = round((cnt_now - cnt_prev) / cnt_prev * 100, 1)
        if pct > 0:
            trending[tag] = {'now': cnt_now, 'prev': cnt_prev, 'change': cnt_now - cnt_prev, 'pct_change': pct}
    top_trending = sorted(trending.items(), key=lambda x: x[1]['pct_change'], reverse=True)[:TOP_N]

    # ================================================================
    # BUILD EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ================================================================
    # SHEET 1 — EXECUTIVE SUMMARY
    # ================================================================
    ws = wb.create_sheet("Executive Summary")
    ws.column_dimensions["A"].width = 38
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = 15

    ws.append(["Repair / RMA Report — Executive Summary"])
    ws.merge_cells("A1:H1")
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = make_fill(MAIN_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   "
                f"Active RMAs: {total_active}   |   "
                f"Last 90d: {len(repairs_90)}   |   Last 30d: {len(repairs_30)}")
    ws["A2"].font      = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])

    # --- TABLE 1: Volume by State ---
    write_section_title(ws, "  📊  Active RMAs by State", MAIN_COLOR, 5)
    write_headers(ws, ["State", "Count", "% of Active", "C2", "C-100"], MAIN_COLOR)
    for state in STATE_ORDER:
        st = sl(repairs_raw, state=state)
        cnt = len(st)
        pct = round(cnt / total_active * 100, 1) if total_active else 0
        c2_cnt  = len(sl(st, device='C2'))
        c100_cnt = len(sl(st, device='C-100'))
        ws.append([state, cnt, f"{pct}%", c2_cnt, c100_cnt])
        style_row(ws, ws.max_row, 5, make_fill(STATE_COLORS[state]), left_cols=[1])
    ws.append([])

    # --- TABLE 2: Volume by Division x Device ---
    write_section_title(ws, "  🌍  Volume by Division × Device", MAIN_COLOR, 8)
    write_headers(ws, ["Division", "Device", "Active", "Last 90d", "Last 30d", "RMA Created", "Device Received", "Under Repair"], MAIN_COLOR)
    for div in cfg.REPAIR_DIVISIONS:
        short = cfg.REPAIR_DIVISION_LABELS.get(div, div)
        for device in cfg.DEVICE_LINES:
            d_active = sl(repairs_raw, division=div, device=device)
            d_90     = sl(repairs_90,  division=div, device=device)
            d_30     = sl(repairs_30,  division=div, device=device)
            row_fill = make_fill("EBF3FB") if device == "C2" else make_fill("FFF9E6")
            ws.append([
                short, device,
                len(d_active), len(d_90), len(d_30),
                len(sl(d_active, state='RMA Created')),
                len(sl(d_active, state='Device Received')),
                len(sl(d_active, state='Under Repair')),
            ])
            style_row(ws, ws.max_row, 8, row_fill, left_cols=[1, 2])
    ws.append([])

    # --- TABLE 3: Age Distribution ---
    write_section_title(ws, "  ⏱  Age Distribution — Active RMAs", "4A4A4A", 8)
    write_headers(ws, ["Age Bracket", "Total", "% of Active", "Americas", "EMEA", "APAC", "C2", "C-100"], "4A4A4A")
    for bucket in bucket_order:
        bt   = [r for r in repairs_raw if r['_bucket'] == bucket]
        cnt  = len(bt)
        pct  = round(cnt / total_active * 100, 1) if total_active else 0
        divs = [len(sl(bt, division=d)) for d in cfg.REPAIR_DIVISIONS]
        devs = [len(sl(bt, device=dv)) for dv in cfg.DEVICE_LINES]
        ws.append([bucket, cnt, f"{pct}%"] + divs + devs)
        style_row(ws, ws.max_row, 8, make_fill(bucket_colors[bucket]), left_cols=[1])
    ws.append([])

    # --- TABLE 4: Top 5 Tags All Time ---
    tags_all = count_tags(repairs_raw)
    write_section_title(ws, "  🏷  Top 5 Repair Tags — Active RMAs", MAIN_COLOR, 5)
    write_headers(ws, ["Rank", "Tag", "# Repairs", "% of Active", "Visual Bar"], MAIN_COLOR)
    if tags_all:
        for rank, (tag, cnt) in enumerate(tags_all.most_common(TOP_N), 1):
            pct = round(cnt / total_active * 100, 1) if total_active else 0
            ws.append([rank, tag, cnt, f"{pct}%", make_bar(pct)])
            style_row(ws, ws.max_row, 5, make_fill(rank_fills[rank-1]), left_cols=[2])
    else:
        ws.append(["", "No tagged repairs found", "", "", ""])
    ws.append([])

    # --- TABLE 5: Top 5 Tags Last 30d ---
    write_section_title(ws, "  🏷  Top 5 Repair Tags — Last 30 Days", "7B2C2C", 5)
    write_headers(ws, ["Rank", "Tag", "# Repairs (30d)", "% of 30d", "Visual Bar"], "7B2C2C")
    total_30 = len(repairs_30)
    if tags_30d:
        for rank, (tag, cnt) in enumerate(tags_30d.most_common(TOP_N), 1):
            pct = round(cnt / total_30 * 100, 1) if total_30 else 0
            ws.append([rank, tag, cnt, f"{pct}%", make_bar(pct)])
            style_row(ws, ws.max_row, 5, make_fill(rank_fills[rank-1]), left_cols=[2])
    else:
        ws.append(["", "No tagged repairs in last 30 days", "", "", ""])
    ws.append([])

    # --- TABLE 6: Trending Tags ---
    write_section_title(ws, f"  📈  Trending Repair Tags — Last 30d vs Prior 30d (min {MIN_TREND}, rising only)", "5C4033", 6)
    write_headers(ws, ["Rank", "Tag", "Last 30d", "Prior 30d", "Change", "% Increase"], "5C4033")
    if top_trending:
        for rank, (tag, data) in enumerate(top_trending, 1):
            ws.append([rank, tag, data['now'], data['prev'], f"+{data['change']}", f"+{data['pct_change']}%"])
            style_row(ws, ws.max_row, 6, make_fill(rank_fills[rank-1]), left_cols=[2])
    else:
        ws.append(["", "No qualifying trending tags found", "", "", "", ""])
        style_row(ws, ws.max_row, 6, make_fill("F2F2F2"), left_cols=[2])
    ws.append([])

    # ================================================================
    # SHEET 2 — DETAIL BY DIVISION × DEVICE
    # ================================================================
    ws2 = wb.create_sheet("By Division × Device")
    ws2.column_dimensions["A"].width = 38
    for i in range(2, 8):
        ws2.column_dimensions[get_column_letter(i)].width = 15

    ws2.append(["Repair Detail — By Division × Device"])
    ws2.merge_cells("A1:G1")
    ws2["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws2["A1"].fill      = make_fill(MAIN_COLOR)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    ws2.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC"])
    ws2.merge_cells("A2:G2")
    ws2["A2"].font      = Font(italic=True)
    ws2["A2"].alignment = Alignment(horizontal="center")
    ws2.append([])

    for div in cfg.REPAIR_DIVISIONS:
        short     = cfg.REPAIR_DIVISION_LABELS.get(div, div)
        div_color = division_colors.get(div, MAIN_COLOR)
        for device in cfg.DEVICE_LINES:
            dev_color = device_colors.get(device, MAIN_COLOR)
            seg       = sl(repairs_raw, division=div, device=device)
            seg_30    = sl(repairs_30,  division=div, device=device)
            seg_90    = sl(repairs_90,  division=div, device=device)

            write_section_title(ws2, f"  {short}  ▸  {device}  ({len(seg)} active RMAs)", dev_color, 6)

            # State breakdown
            write_section_title(ws2, "    By State", div_color, 4)
            write_headers(ws2, ["State", "Count", "% of Segment", "Visual Bar"], div_color)
            for state in STATE_ORDER:
                st_seg = sl(seg, state=state)
                cnt    = len(st_seg)
                pct    = round(cnt / len(seg) * 100, 1) if seg else 0
                ws2.append([state, cnt, f"{pct}%", make_bar(pct)])
                style_row(ws2, ws2.max_row, 4, make_fill(STATE_COLORS[state]), left_cols=[1])
            ws2.append([])

            # Age distribution
            write_section_title(ws2, "    Age Distribution", div_color, 4)
            write_headers(ws2, ["Age Bracket", "Count", "% of Segment", "Visual Bar"], div_color)
            for bucket in bucket_order:
                bt  = [r for r in seg if r['_bucket'] == bucket]
                cnt = len(bt)
                pct = round(cnt / len(seg) * 100, 1) if seg else 0
                ws2.append([bucket, cnt, f"{pct}%", make_bar(pct)])
                style_row(ws2, ws2.max_row, 4, make_fill(bucket_colors[bucket]), left_cols=[1])
            ws2.append([])

            # Top tags
            seg_tags = count_tags(seg)
            write_section_title(ws2, f"    Top {TOP_N} Tags", div_color, 5)
            write_headers(ws2, ["Rank", "Tag", "# Repairs", "% of Segment", "Visual Bar"], div_color)
            if seg_tags:
                for rank, (tag, cnt) in enumerate(seg_tags.most_common(TOP_N), 1):
                    pct = round(cnt / len(seg) * 100, 1) if seg else 0
                    ws2.append([rank, tag, cnt, f"{pct}%", make_bar(pct)])
                    style_row(ws2, ws2.max_row, 5, make_fill(rank_fills[rank-1]), left_cols=[2])
            else:
                ws2.append(["", "No tagged repairs in this segment", "", "", ""])
            ws2.append([])

    # ================================================================
    # SHEET 3 — REPAIR DETAIL LIST
    # ================================================================
    ws3 = wb.create_sheet("Repair List")
    ws3.freeze_panes = "A5"
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 35
    col_w = [14, 35, 16, 12, 14, 20, 25, 15, 12, 15, 20, 20, 20, 18, 18, 20]
    for i, w in enumerate(col_w, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.append(["Repair / RMA — Full Detail List (Active Only)"])
    ws3.merge_cells("A1:P1")
    ws3["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill      = make_fill(MAIN_COLOR)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22
    ws3.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   Active RMAs: {total_active}"])
    ws3.merge_cells("A2:P2")
    ws3["A2"].font      = Font(italic=True)
    ws3["A2"].alignment = Alignment(horizontal="center")
    ws3.append([])

    col_headers = [
        "RMA #", "Product", "Device Line", "Lot/Serial", "State",
        "Customer", "Division", "Responsible", "Age (Days)",
        "Under Warranty", "Issue Reproduced", "Reason for Return",
        "Linked Ticket", "Incoming Tracking", "Outgoing Tracking",
        "Created On"
    ]
    write_headers(ws3, col_headers, MAIN_COLOR)
    ws3.row_dimensions[ws3.max_row].height = 30

    # Group by division then state
    row_num = 5
    for div in cfg.REPAIR_DIVISIONS:
        short     = cfg.REPAIR_DIVISION_LABELS.get(div, div)
        div_color = division_colors.get(div, MAIN_COLOR)
        div_reps  = sl(repairs_raw, division=div)
        if not div_reps:
            continue

        # Division header
        ws3.merge_cells(f"A{row_num}:P{row_num}")
        ws3.cell(row=row_num, column=1).value     = f"  ▶  {short}  ({len(div_reps)} active RMAs)"
        ws3.cell(row=row_num, column=1).font      = Font(bold=True, size=12, color="FFFFFF")
        ws3.cell(row=row_num, column=1).fill      = make_fill(div_color)
        ws3.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[row_num].height = 18
        row_num += 1

        for state in STATE_ORDER:
            state_reps = sl(div_reps, state=state)
            if not state_reps:
                continue

            # State sub-header
            ws3.merge_cells(f"A{row_num}:P{row_num}")
            ws3.cell(row=row_num, column=1).value     = f"    {state}  ({len(state_reps)} repairs)"
            ws3.cell(row=row_num, column=1).font      = Font(bold=True, size=11)
            ws3.cell(row=row_num, column=1).fill      = make_fill(STATE_COLORS[state])
            ws3.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
            row_num += 1

            for idx, r in enumerate(sorted(state_reps, key=lambda x: x.get('_age', 0), reverse=True)):
                row_fill = make_fill("EBF3FB") if idx % 2 == 0 else make_fill("FFFFFF")

                # Flag old open repairs
                if r['_age'] > 60:   row_fill = make_fill("FFB3B3")
                elif r['_age'] > 30: row_fill = make_fill("FFD9B3")

                ticket_name = r['ticket_id'][1] if r.get('ticket_id') else '—'
                row_data = [
                    r.get('name', ''),
                    r['product_id'][1] if r.get('product_id') else '',
                    r['_device'],
                    r['lot_id'][1]     if r.get('lot_id')     else '',
                    r['_state_label'],
                    r['partner_id'][1] if r.get('partner_id') else '',
                    r['division_id'][1] if r.get('division_id') else '',
                    r['user_id'][1]    if r.get('user_id')    else 'Unassigned',
                    r['_age'],
                    'Yes' if r.get('x_studio_under_warranty_ts_case') else 'No',
                    r.get('x_studio_issue_reproduced', '') or '',
                    r.get('x_studio_reason_for_return', '') or '',
                    ticket_name,
                    r.get('x_studio_incoming_tracking_', '') or '',
                    r.get('x_studio_outgoing_tracking', '') or '',
                    r['_created'].strftime('%Y-%m-%d') if r['_created'] else '',
                ]
                ws3.append(row_data)
                for col in range(1, len(col_headers) + 1):
                    c = ws3.cell(row=row_num, column=col)
                    c.fill      = row_fill
                    c.border    = thin_border()
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws3.row_dimensions[row_num].height = 15
                row_num += 1

        ws3.append([])
        row_num += 1

    # ================================================================
    # SHEET 4 — PIE CHARTS
    # ================================================================
    ws_charts = wb.create_sheet("Charts")
    ws_charts.column_dimensions["A"].width = 3

    ws_charts.append(["Repair / RMA — Charts"])
    ws_charts.merge_cells("A1:P1")
    ws_charts["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws_charts["A1"].fill      = make_fill(MAIN_COLOR)
    ws_charts["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_charts.row_dimensions[1].height = 26
    ws_charts.merge_cells("A2:P2")
    ws_charts["A2"] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   Active RMAs: {total_active}"
    ws_charts["A2"].font      = Font(italic=True)
    ws_charts["A2"].alignment = Alignment(horizontal="center")

    # Hidden data sheet for chart data
    ws_data = wb.create_sheet("_chart_data")
    ws_data.sheet_state = 'hidden'

    # --- Chart 1: Repairs by Division ---
    print("\n=== Chart division debug ===")
    print(f"Total repairs_raw at chart time: {len(repairs_raw)}")
    div_sample = set(r['_division'] for r in repairs_raw)
    print(f"Distinct _division values: {div_sample}")
    div_rows = []
    for div in cfg.REPAIR_DIVISIONS:
        short = cfg.REPAIR_DIVISION_LABELS.get(div, div)
        cnt   = len(sl(repairs_raw, division=div))
        print(f"  {div} -> {cnt} repairs")
        if cnt > 0:
            div_rows.append((short, cnt))
    if div_rows:
        chart1, _ = add_pie_chart(ws_data, wb,
            f"Active RMAs by Division ({total_active} total)",
            div_rows, "B4",
            slice_colors=["1F4E79", "375623", "7B2C2C"]
        )
        ws_charts.add_chart(chart1, "B4")

    # --- Chart 2: Repairs by Device (C2 vs C-100) ---
    dev_rows = [(dv, len(sl(repairs_raw, device=dv))) for dv in cfg.DEVICE_LINES if len(sl(repairs_raw, device=dv)) > 0]
    if dev_rows:
        chart2, _ = add_pie_chart(ws_data, wb,
            f"Active RMAs by Device Line",
            dev_rows, "K4",
            slice_colors=["4A235A", "1A5276"]
        )
        ws_charts.add_chart(chart2, "K4")

    # --- Chart 3: Repairs by State ---
    state_rows = [(s, len(sl(repairs_raw, state=s))) for s in STATE_ORDER if len(sl(repairs_raw, state=s)) > 0]
    if state_rows:
        chart3, _ = add_pie_chart(ws_data, wb,
            f"Active RMAs by State",
            state_rows, "B30",
            slice_colors=["FFC000", "4472C4", "ED7D31"]
        )
        ws_charts.add_chart(chart3, "B30")

    # --- Chart 4: Repairs by Division × Device (C2) ---
    c2_div_rows = []
    for div in cfg.REPAIR_DIVISIONS:
        short = cfg.REPAIR_DIVISION_LABELS.get(div, div)
        cnt   = len(sl(repairs_raw, division=div, device='C2'))
        if cnt > 0:
            c2_div_rows.append((short, cnt))
    if c2_div_rows:
        chart4, _ = add_pie_chart(ws_data, wb,
            "C2 RMAs by Division",
            c2_div_rows, "K30",
            slice_colors=["6C3483", "A569BD", "D2B4DE"]
        )
        ws_charts.add_chart(chart4, "K30")

    # --- Chart 5: Repairs by Division × Device (C-100) ---
    c100_div_rows = []
    for div in cfg.REPAIR_DIVISIONS:
        short = cfg.REPAIR_DIVISION_LABELS.get(div, div)
        cnt   = len(sl(repairs_raw, division=div, device='C-100'))
        if cnt > 0:
            c100_div_rows.append((short, cnt))
    if c100_div_rows:
        chart5, _ = add_pie_chart(ws_data, wb,
            "C-100 RMAs by Division",
            c100_div_rows, "B56",
            slice_colors=["1A5276", "2E86C1", "AED6F1"]
        )
        ws_charts.add_chart(chart5, "B56")

    # Move Charts sheet to front
    wb.move_sheet("Charts", offset=-wb.index(wb["Charts"]))

    # Save
    output_file = f"repair_report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\nExcel report saved: {output_file}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
