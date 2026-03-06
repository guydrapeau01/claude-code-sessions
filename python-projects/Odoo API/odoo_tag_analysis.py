import xmlrpc.client
import openpyxl
import odoo_config as cfg
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from collections import Counter
from datetime import datetime, timezone, timedelta

TOP_N     = 5
MIN_TREND = 3  # minimum tickets in last 30d to qualify as trending

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def style_row(ws, row_num, num_cols, fill, left_cols=None):
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.fill   = fill
        c.border = thin_border()
        c.alignment = Alignment(
            horizontal="left" if left_cols and col in left_cols else "center",
            vertical="center"
        )

def write_section_title(ws, title, color, num_cols):
    ws.append([title])
    r = ws.max_row
    ws.merge_cells(f"A{r}:{get_column_letter(num_cols)}{r}")
    ws.cell(row=r, column=1).font      = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=r, column=1).fill      = make_fill(color)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

def write_headers(ws, headers, color):
    ws.append(headers)
    r = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.font      = Font(color="FFFFFF", bold=True, size=10)
        c.fill      = make_fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

def make_bar(pct, width=20):
    filled = round(pct / 100 * width)
    return "█" * filled + "░" * (width - filled)

def add_pie_chart(wb, title, tags_counter, total, period_label, color_hex):
    """Create a hidden data sheet and pie chart for top 5 tags"""
    # Create data sheet
    sheet_name = f"_data_{period_label}"[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_data = wb.create_sheet(sheet_name)
    ws_data.sheet_state = 'hidden'

    ws_data.append(["Tag", "Count"])
    top5 = tags_counter.most_common(5)
    for tag, cnt in top5:
        ws_data.append([tag, cnt])

    # Create pie chart
    chart = PieChart()
    chart.title = title
    chart.style = 10

    data_ref   = Reference(ws_data, min_col=2, min_row=1, max_row=len(top5) + 1)
    labels_ref = Reference(ws_data, min_col=1, min_row=2, max_row=len(top5) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels_ref)
    chart.dataLabels         = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showPercent    = True
    chart.dataLabels.showCatName    = True
    chart.dataLabels.showVal        = False
    chart.dataLabels.showSerName    = False
    chart.dataLabels.showLeaderLines = True

    # Color slices
    slice_colors = ["4472C4", "ED7D31", "A9D18E", "FFC000", "5B9BD5"]
    for i, color in enumerate(slice_colors[:len(top5)]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        chart.series[0].dPt.append(pt)

    chart.width  = 18
    chart.height = 14

    return chart

def count_tags(ticket_list):
    c = Counter()
    for t in ticket_list:
        for tid in t.get('tag_ids', []):
            c[tag_name_map.get(tid, f'Unknown({tid})')] += 1
    return c

def top_tags_table(ws, ticket_list, color, period_label, num_cols=5):
    total = len(ticket_list)
    tags  = count_tags(ticket_list)
    write_headers(ws, ["Rank", "Tag", f"# Tickets ({period_label})", f"% of {period_label}", "Visual Bar"], color)
    if tags:
        for rank, (tag, cnt) in enumerate(tags.most_common(TOP_N), 1):
            pct = round(cnt / total * 100, 1) if total else 0
            ws.append([rank, tag, cnt, f"{pct}%", make_bar(pct)])
            style_row(ws, ws.max_row, num_cols, make_fill(rank_fills[rank-1]), left_cols=[2])
    else:
        ws.append(["", "No tagged tickets found", "", "", ""])
        style_row(ws, ws.max_row, num_cols, make_fill("F2F2F2"), left_cols=[2])
    ws.append([])

rank_fills = ["FFD700", "C0C0C0", "CD7F32", "DDEBF7", "EBF3FB"]

# Device line colors
device_colors = {
    "C2":    "4A235A",   # purple
    "C-100": "1A5276",   # dark blue
}
division_colors = {
    "Support - Americas": "1F4E79",
    "Support - EMEA":     "375623",
    "Support - APAC":     "7B2C2C",
}
bucket_colors = {
    '0-30 days':  'E2EFDA',
    '31-60 days': 'FFF2CC',
    '61-90 days': 'FFD9B3',
    '90+ days':   'FFB3B3'
}
bucket_order = ['0-30 days', '31-60 days', '61-90 days', '90+ days']

try:
    print("Connecting to Odoo...")
    common = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/common")
    uid    = common.authenticate(cfg.DB, cfg.USERNAME, cfg.API_KEY, {})
    models = xmlrpc.client.ServerProxy(f"{cfg.URL}/xmlrpc/2/object")
    print("Connected!\n")

    now  = datetime.now(timezone.utc)
    d30  = now - timedelta(days=30)
    d60  = now - timedelta(days=60)
    d90  = now - timedelta(days=90)

    # --- Get teams ---
    all_teams = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.team', 'search_read',
        [[]], {'fields': ['id', 'name']}
    )
    team_map = {t['name']: t['id'] for t in all_teams}

    # --- Get tag definitions ---
    all_tags = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.tag', 'search_read',
        [[]], {'fields': ['id', 'name']}
    )
    tag_name_map = {t['id']: t['name'] for t in all_tags}

    # --- Resolve C2 product IDs from internal refs ---
    c2_products = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'product.product', 'search_read',
        [[['default_code', 'in', cfg.C2_PRODUCT_REFS]]],
        {'fields': ['id', 'name', 'default_code']}
    )
    c2_product_ids = {p['id'] for p in c2_products}
    print(f"C2 products: {[(p['default_code'], p['name']) for p in c2_products]}")

    # --- Fetch all tickets ---
    all_team_ids = [team_map[d] for d in cfg.DIVISIONS if d in team_map]
    print("\nFetching all tickets...")
    all_tickets = models.execute_kw(cfg.DB, uid, cfg.API_KEY, 'helpdesk.ticket', 'search_read',
        [[['team_id', 'in', all_team_ids]]],
        {'fields': ['id', 'tag_ids', 'create_date', 'team_id', 'stage_id', 'product_id']}
    )
    print(f"  → {len(all_tickets)} total tickets fetched\n")

    # --- Annotate each ticket ---
    for t in all_tickets:
        cd = t.get('create_date', '')
        t['_created']  = datetime.strptime(cd, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone.utc) if cd else None
        t['_division'] = next((d for d in cfg.DIVISIONS if team_map.get(d) == t['team_id'][0]), 'Unknown') if t.get('team_id') else 'Unknown'
        t['_stage']    = t['stage_id'][1] if t.get('stage_id') else 'Unknown'
        t['_is_open']  = t['_stage'] not in cfg.CLOSED_STAGES

        # Device line
        prod_id        = t['product_id'][0] if t.get('product_id') else None
        t['_device']   = 'C2' if prod_id in c2_product_ids else 'C-100'

        # Age bucket
        if t['_created']:
            age = (now - t['_created']).days
            t['_age'] = age
            if age <= 30:   t['_bucket'] = '0-30 days'
            elif age <= 60: t['_bucket'] = '31-60 days'
            elif age <= 90: t['_bucket'] = '61-90 days'
            else:           t['_bucket'] = '90+ days'
        else:
            t['_age']    = 0
            t['_bucket'] = 'Unknown'

    total_all    = len(all_tickets)
    open_tickets = [t for t in all_tickets if t['_is_open']]
    total_open   = len(open_tickets)
    tickets_30   = [t for t in open_tickets if t['_created'] and t['_created'] >= d30]
    tickets_90   = [t for t in open_tickets if t['_created'] and t['_created'] >= d90]
    tickets_p30  = [t for t in open_tickets if t['_created'] and d60 <= t['_created'] < d30]

    tags_all = count_tags(all_tickets)
    tags_90d = count_tags(tickets_90)
    tags_30d = count_tags(tickets_30)

    # --- Trending logic ---
    tags_prev30 = count_tags(tickets_p30)
    trending = {}
    for tag, cnt_now in tags_30d.items():
        if cnt_now < MIN_TREND:
            continue
        cnt_prev = tags_prev30.get(tag, 0)
        if cnt_prev == 0:
            continue
        pct = round((cnt_now - cnt_prev) / cnt_prev * 100, 1)
        if pct > 0:
            trending[tag] = {'now': cnt_now, 'prev': cnt_prev, 'change': cnt_now - cnt_prev, 'pct_change': pct}
    top_trending = sorted(trending.items(), key=lambda x: x[1]['pct_change'], reverse=True)[:TOP_N]

    # --- Build slice helper ---
    # slice_tickets(tickets, division=None, device=None)
    def sl(tickets, division=None, device=None):
        r = tickets
        if division: r = [t for t in r if t['_division'] == division]
        if device:   r = [t for t in r if t['_device']   == device]
        return r

    # ================================================================
    # BUILD EXCEL
    # ================================================================
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    MAIN_COLOR = "2E4057"

    # ================================================================
    # SHEET 1 — EXECUTIVE SUMMARY
    # ================================================================
    ws = wb.create_sheet("Executive Summary")
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["E"].width = 24

    ws.append(["Helpdesk Tag & Ticket Analysis — Executive Summary"])
    ws.merge_cells("A1:G1")
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = make_fill(MAIN_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:G2")
    ws["A2"] = (f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   "
                f"All Tickets: {total_all}   |   Open: {total_open}   |   "
                f"Open Last 90d: {len(tickets_90)}   |   Open Last 30d: {len(tickets_30)}")
    ws["A2"].font      = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])

    # --- TABLE 1: Top 5 Tags All Time ---
    write_section_title(ws, "  📊  Top 5 Tags — All Time (all tickets incl. closed)", MAIN_COLOR, 5)
    top_tags_table(ws, all_tickets, MAIN_COLOR, "All Time")

    # --- TABLE 2: Top 5 Tags Last 90d (open only) ---
    write_section_title(ws, "  📊  Top 5 Tags — Last 90 Days (open tickets only)", "375623", 5)
    top_tags_table(ws, tickets_90, "375623", "90d")

    # --- TABLE 3: Top 5 Tags Last 30d (open only) ---
    write_section_title(ws, "  📊  Top 5 Tags — Last 30 Days (open tickets only)", "7B2C2C", 5)
    top_tags_table(ws, tickets_30, "7B2C2C", "30d")

    # --- TABLE 4: Age Distribution (open only) ---
    write_section_title(ws, "  ⏱  Ticket Age Distribution — Open Tickets Only", "4A4A4A", 8)
    write_headers(ws, ["Age Bracket", "Total Open", "% of Open", "Americas", "EMEA", "APAC", "C2", "C-100"], "4A4A4A")
    for bucket in bucket_order:
        bt = [t for t in open_tickets if t['_bucket'] == bucket]
        cnt = len(bt)
        pct = round(cnt / total_open * 100, 1) if total_open else 0
        div_cnts = [len(sl(bt, division=d)) for d in cfg.DIVISIONS]
        dev_cnts = [len(sl(bt, device=dv)) for dv in cfg.DEVICE_LINES]
        ws.append([bucket, cnt, f"{pct}%"] + div_cnts + dev_cnts)
        style_row(ws, ws.max_row, 8, make_fill(bucket_colors[bucket]), left_cols=[1])
    ws.append([])

    # --- TABLE 5: Trending Tags ---
    write_section_title(ws, f"  📈  Trending Tags — Last 30d vs Prior 30d (min {MIN_TREND} tickets, rising only)", "5C4033", 6)
    write_headers(ws, ["Rank", "Tag", "Last 30d", "Prior 30d", "Change", "% Increase"], "5C4033")
    if top_trending:
        for rank, (tag, data) in enumerate(top_trending, 1):
            ws.append([rank, tag, data['now'], data['prev'], f"+{data['change']}", f"+{data['pct_change']}%"])
            style_row(ws, ws.max_row, 6, make_fill(rank_fills[rank-1]), left_cols=[2])
    else:
        ws.append(["", "No qualifying trending tags found", "", "", "", ""])
        style_row(ws, ws.max_row, 6, make_fill("F2F2F2"), left_cols=[2])
    ws.append([f"  Note: Min {MIN_TREND} tickets in last 30d required. Tags new in last 30d excluded."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="808080")
    ws.append([])

    # --- TABLE 6: Volume by Division × Device ---
    write_section_title(ws, "  🌍  Ticket Volume — Division × Device Line", MAIN_COLOR, 8)
    write_headers(ws, ["Segment", "Device", "All Time", "Open", "Open Last 90d", "Open Last 30d", "30d % of Open", "Trending Tags"], MAIN_COLOR)
    for div in cfg.DIVISIONS:
        short = div.replace("Support - ", "")
        for device in cfg.DEVICE_LINES:
            d_all  = len(sl(all_tickets,  division=div, device=device))
            d_open = len(sl(open_tickets, division=div, device=device))
            d_90   = len(sl(tickets_90,   division=div, device=device))
            d_30   = len(sl(tickets_30,   division=div, device=device))
            pct30  = round(d_30 / d_open * 100, 1) if d_open else 0
            # Top trending tag for this segment
            seg_tags = count_tags(sl(tickets_30, division=div, device=device))
            top_tag  = seg_tags.most_common(1)[0][0] if seg_tags else "—"
            color_key = div
            row_fill  = make_fill("EBF3FB") if device == "C2" else make_fill("FFF9E6")
            ws.append([short, device, d_all, d_open, d_90, d_30, f"{pct30}%", top_tag])
            style_row(ws, ws.max_row, 8, row_fill, left_cols=[1, 2, 8])
    ws.append([])

    # ================================================================
    # SHEET 1b — PIE CHARTS (inserted into Executive Summary)
    # ================================================================
    ws_charts = wb.create_sheet("Tag Charts")
    ws_charts.column_dimensions["A"].width = 3

    ws_charts.append(["Tag Distribution — Pie Charts"])
    ws_charts.merge_cells("A1:L1")
    ws_charts["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws_charts["A1"].fill      = make_fill(MAIN_COLOR)
    ws_charts["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_charts.row_dimensions[1].height = 26

    ws_charts.merge_cells("A2:L2")
    ws_charts["A2"] = f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   Open tickets only"
    ws_charts["A2"].font      = Font(italic=True)
    ws_charts["A2"].alignment = Alignment(horizontal="center")

    # Chart 1 — Last 90 days
    if tags_90d:
        chart_90 = add_pie_chart(wb, f"Top 5 Tags — Last 90 Days ({len(tickets_90)} open tickets)",
                                 tags_90d, len(tickets_90), "90d", "1F4E79")
        ws_charts.add_chart(chart_90, "B4")

    # Chart 2 — Last 30 days
    if tags_30d:
        chart_30 = add_pie_chart(wb, f"Top 5 Tags — Last 30 Days ({len(tickets_30)} open tickets)",
                                 tags_30d, len(tickets_30), "30d", "7B2C2C")
        ws_charts.add_chart(chart_30, "K4")

    # ================================================================
    # SHEET 2 — BY DIVISION × DEVICE (top 5 tags per segment)
    # ================================================================
    ws2 = wb.create_sheet("By Division × Device")
    ws2.column_dimensions["A"].width = 38
    for i in range(2, 8):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    ws2.append(["Top 5 Tags by Division × Device Line"])
    ws2.merge_cells("A1:G1")
    ws2["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws2["A1"].fill      = make_fill(MAIN_COLOR)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    ws2.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC   |   Open tickets only"])
    ws2.merge_cells("A2:G2")
    ws2["A2"].font      = Font(italic=True)
    ws2["A2"].alignment = Alignment(horizontal="center")
    ws2.append([])

    for div in cfg.DIVISIONS:
        short = div.replace("Support - ", "")
        div_color = division_colors.get(div, MAIN_COLOR)
        for device in cfg.DEVICE_LINES:
            dev_color   = device_colors.get(device, MAIN_COLOR)
            seg_open    = sl(open_tickets,  division=div, device=device)
            seg_30      = sl(tickets_30,    division=div, device=device)
            seg_90      = sl(tickets_90,    division=div, device=device)
            seg_all     = sl(all_tickets,   division=div, device=device)

            write_section_title(ws2, f"  {short}  ▸  {device}  ({len(seg_open)} open tickets)", dev_color, 6)

            # All time
            write_section_title(ws2, f"    All Time ({len(seg_all)} tickets incl. closed)", div_color, 6)
            top_tags_table(ws2, seg_all, div_color, "All Time", num_cols=5)

            # Last 90d
            write_section_title(ws2, f"    Last 90 Days ({len(seg_90)} open tickets)", div_color, 6)
            top_tags_table(ws2, seg_90, div_color, "90d", num_cols=5)

            # Last 30d
            write_section_title(ws2, f"    Last 30 Days ({len(seg_30)} open tickets)", div_color, 6)
            top_tags_table(ws2, seg_30, div_color, "30d", num_cols=5)

            # Age buckets for this segment
            write_section_title(ws2, f"    Age Distribution ({len(seg_open)} open tickets)", div_color, 5)
            write_headers(ws2, ["Age Bracket", "# Tickets", "% of Segment", "Visual Bar", ""], div_color)
            for bucket in bucket_order:
                bt  = [t for t in seg_open if t['_bucket'] == bucket]
                cnt = len(bt)
                pct = round(cnt / len(seg_open) * 100, 1) if seg_open else 0
                ws2.append([bucket, cnt, f"{pct}%", make_bar(pct), ""])
                style_row(ws2, ws2.max_row, 5, make_fill(bucket_colors[bucket]), left_cols=[1])
            ws2.append([])

    # ================================================================
    # SHEET 3 — CROSS-DIVISION COMPARISON (top 10 global tags)
    # ================================================================
    ws3 = wb.create_sheet("Cross-Division")
    ws3.column_dimensions["A"].width = 35
    for i in range(2, 10):
        ws3.column_dimensions[get_column_letter(i)].width = 14

    ws3.append(["Cross-Division × Device Tag Comparison — Top 10 Tags (All Time)"])
    ws3.merge_cells("A1:I1")
    ws3["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill      = make_fill(MAIN_COLOR)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26
    ws3.append([f"Generated: {now.strftime('%Y-%m-%d %H:%M')} UTC"])
    ws3.merge_cells("A2:I2")
    ws3["A2"].alignment = Alignment(horizontal="center")
    ws3["A2"].font      = Font(italic=True)
    ws3.append([])

    write_headers(ws3, ["Tag", "Americas C2", "Americas C-100", "EMEA C2", "EMEA C-100", "APAC C2", "APAC C-100", "Global Total", "Trend (30d)"], MAIN_COLOR)
    alt = False
    for tag, global_cnt in tags_all.most_common(10):
        row_fill = make_fill("EBF3FB") if alt else make_fill("FFFFFF")
        alt = not alt
        trend_data = trending.get(tag, {})
        t_str = f"🔺 +{trend_data['pct_change']}%" if trend_data else "—"
        row = [tag]
        for div in cfg.DIVISIONS:
            for device in cfg.DEVICE_LINES:
                row.append(count_tags(sl(all_tickets, division=div, device=device)).get(tag, 0))
        row += [global_cnt, t_str]
        ws3.append(row)
        style_row(ws3, ws3.max_row, 9, row_fill, left_cols=[1])

    # Save
    output_file = f"tag_analysis_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(output_file)
    print(f"\nExcel report saved: {output_file}")
    print("=== Done ===")

except Exception as e:
    import traceback
    print(f"\nERROR: {e}")
    traceback.print_exc()

input("\nPress Enter to close...")
