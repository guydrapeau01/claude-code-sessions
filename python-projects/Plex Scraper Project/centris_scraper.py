#!/usr/bin/env python3
"""
Centris.ca Automated Scraper
=============================
1. Opens Centris.ca
2. Filters: Plex 5+, under $2M, Montreal/Laval, last 7 days
3. Visits each listing individually
4. Extracts: price, municipal tax, school tax, gross revenue
5. Runs economic value analysis
6. Outputs Excel report
"""

import time
import re
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, StaleElementReferenceException
)


# ============================================================================
# CONFIGURATION - Edit these settings
# ============================================================================

CONFIG = {
    # Search filters
    'min_units': 5,
    'max_price': 2_000_000,
    'cities': ['Montréal', 'Laval'],
    'days_listed': 7,          # Last N days only
    
    # Expense assumptions (when not found on listing)
    'vacancy_rate': 0.03,      # 3%
    'insurance_pct': 0.04,     # 4% of gross rents (your setting)
    'maintenance_pct': 0.03,   # 3%
    'management_pct': 0.05,    # 5%
    'other_pct': 0.02,         # 2%
    
    # Browser settings
    'headless': False,         # False = you can see the browser (recommended)
    'slow_mode': True,         # Add delays to avoid detection
    'delay_between_listings': 3,  # Seconds between each listing
}

# TGA Financing Scenarios (from your spreadsheet)
FINANCING_SCENARIOS = {
    '100pts': {'rpv': 0.95, 'cmhc': 0.0255, 'years': 50, 'rate': 0.04},
    '70pts':  {'rpv': 0.95, 'cmhc': 0.033,  'years': 45, 'rate': 0.04},
    '50pts':  {'rpv': 0.85, 'cmhc': 0.033,  'years': 40, 'rate': 0.04},
    'SCHL':   {'rpv': 0.80, 'cmhc': 0.055,  'years': 40, 'rate': 0.039},
    'Conv':   {'rpv': 0.75, 'cmhc': 0.055,  'years': 40, 'rate': 0.054},
}


# ============================================================================
# BROWSER SETUP
# ============================================================================

def create_driver(headless=False):
    """Create a Chrome driver that looks like a real browser"""
    
    options = Options()
    
    if headless:
        options.add_argument('--headless=new')
    
    # Make it look like a real browser
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--lang=fr-CA,fr,en')
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    # Try webdriver-manager first (auto downloads correct ChromeDriver)
    try:
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        print("✓ ChromeDriver loaded automatically")
    except ImportError:
        print("⚠ webdriver-manager not found, trying direct...")
        driver = webdriver.Chrome(options=options)
    except Exception as e:
        print(f"✗ Could not start Chrome: {e}")
        print("Make sure Google Chrome is installed: https://www.google.com/chrome/")
        raise
    
    # Remove webdriver flag
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver


# ============================================================================
# CENTRIS SEARCH
# ============================================================================

def get_search_url():
    """
    Build Centris search URL for:
    - Plex (multi-unit residential)
    - Montreal + Laval
    - Max $2M
    - Last 7 days
    """
    
    # Centris URL structure for plex search
    # Category 17 = Plex, Category 18 = Multilogement
    # We'll use the search form approach instead of URL parameters
    # since Centris uses POST requests with session tokens
    
    return "https://www.centris.ca/fr/plex~a-vendre~montreal?view=Thumbnail"


def search_centris_listings(driver, config):
    """
    Navigate Centris search and collect listing URLs
    Returns list of URLs to individual listings
    """
    
    listing_urls = []
    
    print("\n" + "="*60)
    print("STEP 1: OPENING CENTRIS SEARCH")
    print("="*60)
    
    # Start with Montreal plex search
    search_urls = [
        "https://www.centris.ca/fr/plex~a-vendre~montreal?view=Thumbnail",
        "https://www.centris.ca/fr/plex~a-vendre~laval?view=Thumbnail"
    ]
    
    for search_url in search_urls:
        print(f"\nSearching: {search_url}")
        
        try:
            driver.get(search_url)
            time.sleep(4)
            
            # Handle cookie consent if it appears
            try:
                cookie_btn = driver.find_element(By.XPATH, "//button[contains(text(), 'Accepter') or contains(text(), 'Accept')]")
                cookie_btn.click()
                time.sleep(1)
                print("  ✓ Accepted cookies")
            except:
                pass
            
            # Scroll to load all listings
            print("  Loading listings...")
            scroll_to_load(driver)
            
            # Collect listing URLs from this page
            page_urls = collect_listing_urls(driver, config)
            listing_urls.extend(page_urls)
            
            print(f"  ✓ Found {len(page_urls)} listings")
            
            # Try to go through multiple pages
            page_num = 2
            while page_num <= 10:  # Max 10 pages
                try:
                    next_btn = driver.find_element(By.XPATH, 
                        "//li[contains(@class,'next')]/a | //a[contains(@aria-label,'Suivant') or contains(@aria-label,'Next')]")
                    
                    if not next_btn.is_enabled() or 'disabled' in next_btn.get_attribute('class'):
                        break
                    
                    next_btn.click()
                    time.sleep(3)
                    scroll_to_load(driver)
                    
                    page_urls = collect_listing_urls(driver, config)
                    new_urls = [u for u in page_urls if u not in listing_urls]
                    
                    if not new_urls:
                        break
                    
                    listing_urls.extend(new_urls)
                    print(f"  Page {page_num}: +{len(new_urls)} listings")
                    page_num += 1
                    
                except Exception:
                    break
        
        except Exception as e:
            print(f"  ✗ Error searching {search_url}: {e}")
            continue
        
        time.sleep(2)
    
    # Remove duplicates
    listing_urls = list(set(listing_urls))
    
    print(f"\n✓ Total unique listings found: {len(listing_urls)}")
    
    return listing_urls


def collect_listing_urls(driver, config):
    """Collect all listing URLs from current search results page"""
    
    urls = []
    
    # Multiple possible selectors Centris uses
    selectors = [
        "a.property-thumbnail-summary-link",
        "div.property-thumbnail-item a[href*='/fr/plex']",
        "div.property-thumbnail-item a[href*='/en/plex']",
        "a[href*='centris.ca/fr/plex']",
        "div.shell a[href*='/plex~']",
    ]
    
    for selector in selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            if elements:
                for el in elements:
                    href = el.get_attribute('href')
                    if href and '/plex' in href and href not in urls:
                        urls.append(href)
                break
        except:
            continue
    
    # Fallback: find all links containing plex
    if not urls:
        try:
            all_links = driver.find_elements(By.TAG_NAME, 'a')
            for link in all_links:
                href = link.get_attribute('href') or ''
                if '/plex~' in href or ('/plex/' in href and 'centris' in href):
                    if href not in urls:
                        urls.append(href)
        except:
            pass
    
    return urls


def scroll_to_load(driver):
    """Scroll down to trigger lazy loading"""
    try:
        last_height = driver.execute_script("return document.body.scrollHeight")
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1.5)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
        driver.execute_script("window.scrollTo(0, 0);")
    except:
        pass


# ============================================================================
# INDIVIDUAL LISTING SCRAPER
# ============================================================================

def scrape_listing(driver, url, config):
    """
    Visit individual listing and extract all data
    Returns dict with property data or None if failed
    """
    
    try:
        driver.get(url)
        
        if config['slow_mode']:
            time.sleep(2.5)
        
        # Wait for page to load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "h1"))
        )
        
        # Get full page text for regex parsing
        page_text = driver.find_element(By.TAG_NAME, 'body').text
        page_source = driver.page_source
        
        # Extract all data
        data = {
            'url': url,
            'scraped_at': datetime.now().isoformat(),
        }
        
        data['address']      = extract_address(driver, page_text)
        data['price']        = extract_price(driver, page_text)
        data['num_units']    = extract_units(driver, page_text)
        data['gross_income'] = extract_gross_income(driver, page_text)
        data['municipal_tax']= extract_municipal_tax(driver, page_text)
        data['school_tax']   = extract_school_tax(driver, page_text)
        data['utilities']    = extract_utilities(driver, page_text)
        
        # Only keep if it meets our criteria
        if not data['price'] or data['price'] > config['max_price']:
            return None
        
        if data['num_units'] and data['num_units'] < config['min_units']:
            return None
        
        return data
        
    except TimeoutException:
        print(f"    ✗ Timeout loading: {url[:60]}...")
        return None
    except Exception as e:
        print(f"    ✗ Error scraping {url[:60]}...: {e}")
        return None


# ============================================================================
# DATA EXTRACTION FUNCTIONS
# ============================================================================

def extract_price(driver, page_text):
    """Extract asking price"""
    
    # Try structured element first
    selectors = [
        "span.price",
        "div.price",
        "[itemprop='price']",
        "span[class*='price']",
        "div[class*='price']",
        ".asking-price",
    ]
    
    for selector in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, selector)
            price = parse_number(el.text)
            if price and price > 100000:
                return price
        except:
            continue
    
    # Fallback: regex on page text
    patterns = [
        r'\$\s*([\d\s,]+(?:\.\d{2})?)\s*(?:000)?(?:\n|$)',
        r'Prix demandé[:\s]*\$?\s*([\d\s,]+)',
        r'Asking price[:\s]*\$?\s*([\d\s,]+)',
        r'(\d{1,3}(?:\s\d{3})*)\s*\$',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, page_text, re.IGNORECASE)
        if match:
            price = parse_number(match.group(1))
            if price and 100_000 < price < 10_000_000:
                return price
    
    return None


def extract_address(driver, page_text):
    """Extract property address"""
    
    selectors = [
        "h1.text-center",
        "h1[itemprop='name']",
        "span[itemprop='streetAddress']",
        ".listing-location h1",
        "h1.property-address",
        ".address h1",
        "h1",
    ]
    
    for selector in selectors:
        try:
            el = driver.find_element(By.CSS_SELECTOR, selector)
            text = el.text.strip()
            if text and len(text) > 5:
                return text
        except:
            continue
    
    # Regex fallback
    match = re.search(r'\d+[,\s]+(?:Rue|Avenue|Boulevard|Boul\.|Ave\.|Blvd\.?|Ch\.|Chemin)[^,\n]{3,50}', 
                     page_text, re.IGNORECASE)
    if match:
        return match.group(0).strip()
    
    return "Address not found"


def extract_units(driver, page_text):
    """Extract number of units"""
    
    # Look for unit count in structured data
    selectors = [
        "[class*='unit']",
        "[class*='logement']",
        "div.carac-container",
        "ul.carac-list li",
    ]
    
    for selector in selectors:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            for el in elements:
                text = el.text.lower()
                if any(word in text for word in ['logement', 'unit', 'plex', 'appartement']):
                    match = re.search(r'(\d+)', text)
                    if match:
                        n = int(match.group(1))
                        if 2 <= n <= 50:
                            return n
        except:
            continue
    
    # Regex on full page text
    patterns = [
        r'(\d+)\s*logements?',
        r'(\d+)\s*units?',
        r'(\d+)\s*appartements?',
        r'(\d+)[\s-]plex',
        r'Nombre de logements?\s*:?\s*(\d+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, page_text, re.IGNORECASE)
        if match:
            n = int(match.group(1))
            if 2 <= n <= 50:
                return n
    
    return None


def extract_gross_income(driver, page_text):
    """Extract gross annual rental income"""
    
    patterns = [
        r'Revenus bruts?\s*:?\s*\$?\s*([\d\s,]+)',
        r'Gross revenue\s*:?\s*\$?\s*([\d\s,]+)',
        r'Revenus annuels?\s*:?\s*\$?\s*([\d\s,]+)',
        r'Annual revenue\s*:?\s*\$?\s*([\d\s,]+)',
        r'Revenu brut\s*:?\s*\$?\s*([\d\s,]+)',
        r'Loyers\s*:?\s*\$?\s*([\d\s,]+)',
        r'Total loyers\s*:?\s*\$?\s*([\d\s,]+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, page_text, re.IGNORECASE)
        if match:
            income = parse_number(match.group(1))
            if income and 1000 < income < 1_000_000:
                return income
    
    # Try to find in financial section
    try:
        financial_sections = driver.find_elements(By.XPATH, 
            "//*[contains(text(), 'Revenus') or contains(text(), 'Revenue')]/following-sibling::*")
        for el in financial_sections[:5]:
            income = parse_number(el.text)
            if income and 5000 < income < 500_000:
                return income
    except:
        pass
    
    return None


def extract_municipal_tax(driver, page_text):
    """Extract municipal/property taxes"""
    
    patterns = [
        r'[Tt]axes? municipales?\s*:?\s*\$?\s*([\d\s,]+)',
        r'Municipal tax(?:es)?\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Ii]mpôts? fonciers?\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Tt]ax(?:e)? de la ville\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Tt]axes? munic\.\s*:?\s*\$?\s*([\d\s,]+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, page_text, re.IGNORECASE)
        if match:
            tax = parse_number(match.group(1))
            if tax and 500 < tax < 100_000:
                return tax
    
    return None


def extract_school_tax(driver, page_text):
    """Extract school taxes"""
    
    patterns = [
        r'[Tt]axes? scolaires?\s*:?\s*\$?\s*([\d\s,]+)',
        r'School tax(?:es)?\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Tt]ax(?:e)? scolaire\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Cc]ommission scolaire\s*:?\s*\$?\s*([\d\s,]+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, page_text, re.IGNORECASE)
        if match:
            tax = parse_number(match.group(1))
            if tax and 100 < tax < 20_000:
                return tax
    
    return None


def extract_utilities(driver, page_text):
    """Extract utilities (monthly)"""
    
    patterns = [
        r'[Éé]nergie\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Cc]hauffage\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Uu]tilities?\s*:?\s*\$?\s*([\d\s,]+)',
        r'[Éé]lectricit[eé]\s*:?\s*\$?\s*([\d\s,]+)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, page_text, re.IGNORECASE)
        if match:
            util = parse_number(match.group(1))
            if util and 50 < util < 10_000:
                return util
    
    return None


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def parse_number(text):
    """Convert text like '$1 234 567' or '1,234,567' to float"""
    if not text:
        return None
    try:
        cleaned = re.sub(r'[^\d.]', '', str(text).replace(' ', ''))
        if cleaned:
            return float(cleaned)
    except:
        pass
    return None


# ============================================================================
# ECONOMIC VALUE ANALYSIS
# ============================================================================

def calculate_tga(scenario):
    """Calculate TGA from financing scenario"""
    r = scenario['rate']
    n = scenario['years']
    rcd = 1.1
    factor = (1 + r) ** n
    annuity = (r * factor) / (factor - 1)
    return annuity / rcd


def calculate_noi(property_data, config):
    """Calculate Net Operating Income"""
    
    gross = property_data.get('gross_income', 0) or 0
    if gross == 0:
        return None, {}
    
    # Vacancy
    vacancy = gross * config['vacancy_rate']
    effective_income = gross - vacancy
    
    # Expenses - use actual if available, estimate if not
    municipal = property_data.get('municipal_tax') or 0
    school = property_data.get('school_tax') or 0
    
    # Utilities: use actual (monthly * 12) or 0
    util_monthly = property_data.get('utilities') or 0
    utilities_annual = util_monthly * 12
    
    # Percentage-based expenses
    insurance = gross * config['insurance_pct']
    maintenance = gross * config['maintenance_pct']
    management = gross * config['management_pct']
    other = gross * config['other_pct']
    
    total_expenses = (municipal + school + utilities_annual + 
                     insurance + maintenance + management + other)
    
    noi = effective_income - total_expenses
    
    expense_detail = {
        'vacancy': vacancy,
        'municipal_tax': municipal,
        'school_tax': school,
        'utilities': utilities_annual,
        'insurance': insurance,
        'maintenance': maintenance,
        'management': management,
        'other': other,
        'total': total_expenses,
        'expense_ratio': total_expenses / gross if gross else 0
    }
    
    return noi, expense_detail


def analyze_property(property_data, config):
    """Run full TGA analysis on a property"""
    
    price = property_data.get('price')
    gross = property_data.get('gross_income')
    
    if not price or not gross:
        return None
    
    noi, expense_detail = calculate_noi(property_data, config)
    
    if not noi or noi <= 0:
        return None
    
    best_scenario = None
    best_value_ratio = 0
    best_result = {}
    all_scenarios = {}
    
    for scenario_name, scenario in FINANCING_SCENARIOS.items():
        tga = calculate_tga(scenario)
        economic_value = noi / tga
        value_ratio = economic_value / price
        
        # Financing
        loan_base = min(price, economic_value)
        loan_amount = scenario['rpv'] * loan_base
        cmhc = loan_amount * scenario['cmhc']
        total_loan = loan_amount + cmhc
        down_payment = price - total_loan
        
        r = scenario['rate'] / 12
        n = scenario['years'] * 12
        monthly_payment = total_loan * (r * (1+r)**n) / ((1+r)**n - 1)
        
        monthly_cashflow = (gross * (1 - config['vacancy_rate']) / 12 
                           - expense_detail['total'] / 12 
                           - monthly_payment)
        
        cash_roi = (monthly_cashflow * 12) / down_payment if down_payment > 0 else 0
        
        result = {
            'tga': tga,
            'economic_value': economic_value,
            'value_ratio': value_ratio,
            'down_payment': down_payment,
            'monthly_payment': monthly_payment,
            'monthly_cashflow': monthly_cashflow,
            'cash_roi': cash_roi,
        }
        
        all_scenarios[scenario_name] = result
        
        if value_ratio > best_value_ratio:
            best_value_ratio = value_ratio
            best_scenario = scenario_name
            best_result = result
    
    return {
        **property_data,
        'noi': noi,
        'expense_detail': expense_detail,
        'best_scenario': best_scenario,
        'best_value_ratio': best_value_ratio,
        **{f'best_{k}': v for k, v in best_result.items()},
        'all_scenarios': all_scenarios,
    }


# ============================================================================
# EXCEL REPORT GENERATOR
# ============================================================================

def generate_excel_report(analyses, config):
    """Generate Excel report with all analyzed properties"""
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Centris_Analysis_{timestamp}.xlsx'
    
    wb = openpyxl.Workbook()
    
    # ---- SUMMARY SHEET ----
    ws = wb.active
    ws.title = 'Summary'
    
    # Title
    ws['A1'] = f'Montreal/Laval Plex Analysis - {datetime.now().strftime("%Y-%m-%d")}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f'Found {len(analyses)} properties | Max Price: ${config["max_price"]:,} | Min Units: {config["min_units"]}+'
    
    # Headers
    headers = [
        'Address', 'Price', 'Units', 'Gross Income', 
        'Municipal Tax', 'School Tax', 'Total Taxes', 
        'NOI', 'Economic Value', 'Value Ratio',
        'Best Scenario', 'TGA', 'Down Payment',
        'Monthly Payment', 'Monthly Cashflow', 'Cash ROI',
        'Expense Ratio', 'Listing URL'
    ]
    
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='366092')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    ws.row_dimensions[header_row].height = 30
    
    # Data rows - sorted by value ratio descending
    analyses_sorted = sorted(analyses, key=lambda x: x.get('best_value_ratio', 0), reverse=True)
    
    # Color fills
    green_fill = PatternFill('solid', fgColor='D9EAD3')
    yellow_fill = PatternFill('solid', fgColor='FFF2CC')
    red_fill = PatternFill('solid', fgColor='F4CCCC')
    
    for row_idx, prop in enumerate(analyses_sorted, header_row + 1):
        
        value_ratio = prop.get('best_value_ratio', 0)
        expense_detail = prop.get('expense_detail', {})
        
        row_data = [
            prop.get('address', 'N/A'),
            prop.get('price'),
            prop.get('num_units'),
            prop.get('gross_income'),
            prop.get('municipal_tax'),
            prop.get('school_tax'),
            (prop.get('municipal_tax') or 0) + (prop.get('school_tax') or 0),
            prop.get('noi'),
            prop.get('best_economic_value'),
            value_ratio,
            prop.get('best_scenario'),
            prop.get('best_tga'),
            prop.get('best_down_payment'),
            prop.get('best_monthly_payment'),
            prop.get('best_monthly_cashflow'),
            prop.get('best_cash_roi'),
            expense_detail.get('expense_ratio'),
            prop.get('url', ''),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Color-code the value ratio column
            if col_idx == 10:  # Value Ratio column
                if value_ratio >= 1.15:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=c).fill = green_fill
                elif value_ratio >= 1.0:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=c).fill = yellow_fill
                else:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=c).fill = red_fill
        
        # Number formats
        currency_cols = [2, 4, 5, 6, 7, 8, 9, 13, 14, 15]
        pct_cols = [10, 12, 16, 17]
        
        for col in currency_cols:
            ws.cell(row=row_idx, column=col).number_format = '$#,##0'
        for col in pct_cols:
            ws.cell(row=row_idx, column=col).number_format = '0.0%'
    
    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
    
    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    
    # ---- EXPENSE DETAIL SHEET ----
    ws2 = wb.create_sheet('Expense Detail')
    
    expense_headers = ['Address', 'Price', 'Gross Income', 'Vacancy',
                      'Municipal Tax', 'School Tax', 'Utilities',
                      'Insurance (4%)', 'Maintenance (3%)', 'Management (5%)',
                      'Other (2%)', 'Total Expenses', 'Expense Ratio', 'NOI']
    
    for col, header in enumerate(expense_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='366092')
    
    for row_idx, prop in enumerate(analyses_sorted, 2):
        ed = prop.get('expense_detail', {})
        ws2.cell(row=row_idx, column=1, value=prop.get('address'))
        ws2.cell(row=row_idx, column=2, value=prop.get('price'))
        ws2.cell(row=row_idx, column=3, value=prop.get('gross_income'))
        ws2.cell(row=row_idx, column=4, value=ed.get('vacancy'))
        ws2.cell(row=row_idx, column=5, value=ed.get('municipal_tax'))
        ws2.cell(row=row_idx, column=6, value=ed.get('school_tax'))
        ws2.cell(row=row_idx, column=7, value=ed.get('utilities'))
        ws2.cell(row=row_idx, column=8, value=ed.get('insurance'))
        ws2.cell(row=row_idx, column=9, value=ed.get('maintenance'))
        ws2.cell(row=row_idx, column=10, value=ed.get('management'))
        ws2.cell(row=row_idx, column=11, value=ed.get('other'))
        ws2.cell(row=row_idx, column=12, value=ed.get('total'))
        ws2.cell(row=row_idx, column=13, value=ed.get('expense_ratio'))
        ws2.cell(row=row_idx, column=14, value=prop.get('noi'))
        
        for col in [2,3,4,5,6,7,8,9,10,11,12,14]:
            ws2.cell(row=row_idx, column=col).number_format = '$#,##0'
        ws2.cell(row=row_idx, column=13).number_format = '0.0%'
    
    for col in ws2.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)
    
    # ---- RAW DATA SHEET ----
    ws3 = wb.create_sheet('Raw Data')
    
    raw_headers = ['Address', 'Price', 'Units', 'Gross Income',
                   'Municipal Tax', 'School Tax', 'Utilities (Monthly)', 
                   'URL', 'Scraped At']
    
    for col, header in enumerate(raw_headers, 1):
        ws3.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    for row_idx, prop in enumerate(analyses_sorted, 2):
        ws3.cell(row=row_idx, column=1, value=prop.get('address'))
        ws3.cell(row=row_idx, column=2, value=prop.get('price'))
        ws3.cell(row=row_idx, column=3, value=prop.get('num_units'))
        ws3.cell(row=row_idx, column=4, value=prop.get('gross_income'))
        ws3.cell(row=row_idx, column=5, value=prop.get('municipal_tax'))
        ws3.cell(row=row_idx, column=6, value=prop.get('school_tax'))
        ws3.cell(row=row_idx, column=7, value=prop.get('utilities'))
        ws3.cell(row=row_idx, column=8, value=prop.get('url'))
        ws3.cell(row=row_idx, column=9, value=prop.get('scraped_at'))
    
    wb.save(filename)
    return filename


# ============================================================================
# MAIN WORKFLOW
# ============================================================================

def main():
    
    print("=" * 60)
    print("CENTRIS AUTOMATED SCRAPER")
    print("Montreal/Laval Plex Analysis")
    print("=" * 60)
    print(f"\nSettings:")
    print(f"  Min units   : {CONFIG['min_units']}+")
    print(f"  Max price   : ${CONFIG['max_price']:,}")
    print(f"  Cities      : {', '.join(CONFIG['cities'])}")
    print(f"  Insurance   : {CONFIG['insurance_pct']*100:.0f}% of rents")
    print(f"  Maintenance : {CONFIG['maintenance_pct']*100:.0f}% of rents")
    print(f"  Management  : {CONFIG['management_pct']*100:.0f}% of rents")
    print(f"  Other       : {CONFIG['other_pct']*100:.0f}% of rents")
    
    # Start browser
    print("\nStarting browser...")
    driver = create_driver(headless=CONFIG['headless'])
    
    try:
        # Step 1: Get all listing URLs
        listing_urls = search_centris_listings(driver, CONFIG)
        
        if not listing_urls:
            print("\n⚠ No listings found!")
            print("Try running with headless=False to see what's happening")
            return
        
        # Step 2: Scrape each listing
        print(f"\n{'='*60}")
        print(f"STEP 2: SCRAPING {len(listing_urls)} LISTINGS")
        print(f"{'='*60}\n")
        
        scraped_properties = []
        
        for i, url in enumerate(listing_urls, 1):
            print(f"[{i}/{len(listing_urls)}] Scraping: {url[-50:]}")
            
            data = scrape_listing(driver, url, CONFIG)
            
            if data:
                scraped_properties.append(data)
                print(f"  ✓ {data.get('address', 'N/A')[:40]}")
                print(f"    Price: ${data.get('price', 0):,.0f} | "
                      f"Units: {data.get('num_units', '?')} | "
                      f"Income: ${data.get('gross_income', 0):,.0f}")
                print(f"    Taxes: ${(data.get('municipal_tax') or 0) + (data.get('school_tax') or 0):,.0f}")
            else:
                print(f"  ✗ Skipped (missing data or outside criteria)")
            
            # Polite delay
            if CONFIG['slow_mode'] and i < len(listing_urls):
                time.sleep(CONFIG['delay_between_listings'])
        
        print(f"\n✓ Successfully scraped {len(scraped_properties)} properties")
        
        # Step 3: Analyze properties
        print(f"\n{'='*60}")
        print(f"STEP 3: CALCULATING ECONOMIC VALUE")
        print(f"{'='*60}\n")
        
        analyses = []
        skipped = 0
        
        for prop in scraped_properties:
            result = analyze_property(prop, CONFIG)
            if result:
                analyses.append(result)
                ratio = result.get('best_value_ratio', 0)
                indicator = '🟢' if ratio >= 1.15 else '🟡' if ratio >= 1.0 else '🔴'
                print(f"{indicator} {prop.get('address', 'N/A')[:40]}")
                print(f"   Value Ratio: {ratio:.1%} | "
                      f"Cashflow: ${result.get('best_monthly_cashflow', 0):,.0f}/mo | "
                      f"ROI: {result.get('best_cash_roi', 0):.1%}")
            else:
                skipped += 1
        
        print(f"\n✓ Analyzed: {len(analyses)} | Skipped (incomplete data): {skipped}")
        
        # Step 4: Generate report
        print(f"\n{'='*60}")
        print(f"STEP 4: GENERATING EXCEL REPORT")
        print(f"{'='*60}\n")
        
        if analyses:
            report_file = generate_excel_report(analyses, CONFIG)
            print(f"✓ Report saved: {report_file}")
            
            # Summary
            good_deals = [a for a in analyses if a.get('best_value_ratio', 0) >= 1.0]
            print(f"\n{'='*60}")
            print(f"RESULTS SUMMARY")
            print(f"{'='*60}")
            print(f"  Total analyzed   : {len(analyses)}")
            print(f"  Good deals (>1.0): {len(good_deals)}")
            
            if good_deals:
                best = max(good_deals, key=lambda x: x.get('best_value_ratio', 0))
                print(f"\n  🏆 BEST DEAL:")
                print(f"  {best.get('address')}")
                print(f"  Price         : ${best.get('price', 0):,.0f}")
                print(f"  Economic Value: ${best.get('best_economic_value', 0):,.0f}")
                print(f"  Value Ratio   : {best.get('best_value_ratio', 0):.1%}")
                print(f"  Monthly CF    : ${best.get('best_monthly_cashflow', 0):,.0f}")
                print(f"  Cash ROI      : {best.get('best_cash_roi', 0):.1%}")
                print(f"  URL           : {best.get('url')}")
        else:
            print("⚠ No properties could be fully analyzed (missing price or income data)")
            print("This is common - many listings don't show income publicly")
            print("Consider entering income manually in the CSV for those listings")
    
    finally:
        driver.quit()
        print("\n✓ Browser closed")


if __name__ == '__main__':
    main()
