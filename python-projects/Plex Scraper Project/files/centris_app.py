#!/usr/bin/env python3
"""
Centris Commercial Multifamilial Scraper
=========================================
Automated: opens browser, applies your filters, paginates,
visits each listing, extracts financials, runs TGA analysis,
saves Excel report.

Usage: py centris_app.py
"""

import re, time, os, math
from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── TGA CONFIG ────────────────────────────────────────────────────────────────
EXPENSE_CONFIG = {
    'vacancy_rate'    : 0.03,
    'insurance_pct'   : 0.04,
    'maintenance_pct' : 0.03,
    'management_pct'  : 0.05,
    'other_pct'       : 0.02,
}

SCENARIOS = {
    '100pts': {'rpv': 0.95, 'cmhc': 0.0255, 'years': 50, 'rate': 0.040},
    '70pts' : {'rpv': 0.95, 'cmhc': 0.0330, 'years': 45, 'rate': 0.040},
    '50pts' : {'rpv': 0.85, 'cmhc': 0.0330, 'years': 40, 'rate': 0.040},
    'SCHL'  : {'rpv': 0.80, 'cmhc': 0.0550, 'years': 40, 'rate': 0.039},
    'Conv'  : {'rpv': 0.75, 'cmhc': 0.0550, 'years': 40, 'rate': 0.054},
}

CITY_SLUGS = {
    '1': ('Montréal',  'montreal'),
    '2': ('Laval',     'laval'),
    '3': ('Longueuil', 'longueuil'),
    '4': ('Québec',    'quebec'),
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def parse_num(text):
    if not text: return None
    try:
        cleaned = re.sub(r'[^\d.]', '', str(text).replace('\xa0','').replace(' ',''))
        return float(cleaned) if cleaned else None
    except: return None

def find_num(text, patterns, lo=None, hi=None):
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            v = parse_num(m.group(1))
            if v is not None:
                if lo and v < lo: continue
                if hi and v > hi: continue
                return v
    return None

# ── CLI FILTERS ───────────────────────────────────────────────────────────────
def ask_filters():
    print("\n" + "═"*55)
    print("   CENTRIS MULTIFAMILIAL SCRAPER")
    print("═"*55)

    print("\nCities:")
    for k,(label,_) in CITY_SLUGS.items():
        print(f"  {k}. {label}")
    raw = input("\nCity numbers (e.g. 1 or 1,2): ").strip()
    cities = []
    for c in raw.split(','):
        c = c.strip()
        if c in CITY_SLUGS:
            cities.append(CITY_SLUGS[c])
    if not cities:
        cities = [CITY_SLUGS['1']]

    raw_price = input("\nMax price (default 2000000): ").strip()
    max_price = int(parse_num(raw_price) or 2_000_000)

    raw_units = input("Min units (default 5): ").strip()
    min_units = int(parse_num(raw_units) or 5)

    raw_days = input("Listed within last N days (default 7, 0 = skip filter): ").strip()
    days = int(parse_num(raw_days) or 7)

    raw_max = input("Max properties to scrape per city (default 50): ").strip()
    max_props = int(parse_num(raw_max) or 50)

    print(f"\n{'─'*55}")
    print(f"  Cities    : {', '.join(l for l,_ in cities)}")
    print(f"  Max price : ${max_price:,}")
    print(f"  Min units : {min_units}+")
    print(f"  Last days : {days if days else 'any'}")
    print(f"  Max props : {max_props} per city")
    print(f"{'─'*55}")
    input("\nPress ENTER to start...")

    return {'cities': cities, 'max_price': max_price,
            'min_units': min_units, 'days': days, 'max_props': max_props}

# ── SCRAPER ───────────────────────────────────────────────────────────────────
def scrape_all(filters):
    UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
          "AppleWebKit/537.36 (KHTML, like Gecko) "
          "Chrome/122.0.0.0 Safari/537.36")

    all_props = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=False,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled",
                  "--start-maximized"]
        )
        ctx = browser.new_context(
            user_agent=UA, locale="fr-CA",
            timezone_id="America/Toronto",
            no_viewport=True,
        )
        page = ctx.new_page()

        for city_label, city_slug in filters['cities']:
            print(f"\n{'═'*55}")
            print(f"  City: {city_label}")
            print(f"{'═'*55}")

            urls = collect_urls(page, city_slug, filters)
            print(f"\n  → {len(urls)} listings found")

            for i, url in enumerate(urls, 1):
                print(f"  [{i}/{len(urls)}] Scraping...", end=" ", flush=True)
                prop = scrape_listing(page, url, filters)
                if prop:
                    prop['city'] = city_label
                    price = prop.get('price', 0) or 0
                    max_p = filters.get('max_price', 99999999)
                    if price > max_p:
                        print(f"✗ over price limit (${price:,.0f} > ${max_p:,.0f})")
                    else:
                        all_props.append(prop)
                        print(f"✓ ${price:,.0f} | "
                              f"{prop.get('num_units','?')} units | "
                              f"Income: ${prop.get('gross_income') or 0:,.0f}")
                else:
                    print("✗ skipped")
                time.sleep(1)

        browser.close()

    return all_props


def collect_urls(page, city_slug, filters):
    """Navigate to search page with filters pre-baked into URL, collect all listing URLs"""

    url = _build_search_url(city_slug, filters)
    print(f"\n  Loading: {url[:100]}...")
    page.goto(url, wait_until="domcontentloaded", timeout=30000)
    page.wait_for_timeout(3000)

    # Dismiss any cookie/login popup
    for sel in ["button:has-text('Fermer')", "button:has-text('Accepter')",
                ".modal .close", "[aria-label='Close']"]:
        try:
            b = page.locator(sel).first
            if b.is_visible(timeout=800):
                b.click()
                page.wait_for_timeout(500)
        except: pass

    # ── Collect all listing URLs across pages ─────────────────────────────
    urls = []
    page_num = 1
    while True:
        print(f"    Page {page_num}...", end=" ", flush=True)

        # Scroll to load lazy content
        for _ in range(3):
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(600)
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(400)

        # Extract listing URLs — match /fr/<anything>~a-vendre~<city>/<8-digit-id>
        new_urls = page.evaluate("""
            () => {
                const seen = new Set();
                const results = [];
                document.querySelectorAll('a[href]').forEach(a => {
                    const h = a.href;
                    // Must contain ~a-vendre~ and end with /XXXXXXXX (7+ digits)
                    if (h.includes('centris.ca/fr/') &&
                        h.includes('~a-vendre~') &&
                        /[/][0-9]{7,}$/.test(h) &&
                        !seen.has(h)) {
                        seen.add(h);
                        results.push(h);
                    }
                });
                return results;
            }
        """)
        # Debug: show sample URL if found
        if new_urls:
            print(f"    sample: {new_urls[0][-60:]}")

        added = [u for u in new_urls if u not in urls]
        urls.extend(added)
        print(f"{len(added)} new ({len(urls)} total)")

        # Stop if we've hit the max
        if len(urls) >= filters.get('max_props', 50):
            urls = urls[:filters.get('max_props', 50)]
            print(f"  → Reached max {filters.get('max_props',50)} properties, stopping pagination")
            break

        # Next page — use JS to find and click the next pagination button
        went_next = page.evaluate("""
            () => {
                // Centris pagination: look for the > arrow link that's not disabled
                const candidates = Array.from(document.querySelectorAll(
                    'li.next:not(.disabled) a, ' +
                    'li:not(.disabled) a[aria-label], ' +
                    '.pagination a'
                ));
                for (const a of candidates) {
                    const txt = a.textContent.trim();
                    const aria = a.getAttribute('aria-label') || '';
                    if (txt === '›' || txt === '»' || txt === '>' ||
                        aria.includes('suivant') || aria.includes('Suivant') ||
                        aria.includes('next') || aria.includes('Next')) {
                        const li = a.closest('li');
                        if (!li || !li.className.includes('disabled')) {
                            a.click();
                            return true;
                        }
                    }
                }
                return false;
            }
        """)
        if went_next:
            page.wait_for_timeout(2500)
            page_num += 1
        else:
            break

    return urls


def _build_search_url(city_slug, filters):
    """
    Build Centris search URL by directly encoding filters as gzipped JSON in q= param.
    No UI interaction needed. Decoded by reverse-engineering the q= parameter format.
    
    Filter JSON structure (confirmed by decoding live URLs):
    {
      "fieldsValues": [
        {"fieldId":"Category","value":"Commercial"},
        {"fieldId":"SellingType","value":"Sale"},
        {"fieldId":"PropertyType","value":"MultiFamily"},
        {"fieldId":"CityDistrictAll","value":"5"},  // 5=Montreal
        {"fieldId":"SalePrice","value":"0","value2":"2000000"},
        {"fieldId":"NumberUnits","value":"5"},
        {"fieldId":"LastModifiedDate","value":"2026-02-06"}
      ]
    }
    """
    import gzip, base64, json
    from datetime import timedelta

    # City config - confirmed by decoding live Centris URLs
    # Format: each range filter needs TWO entries (min + max="999999999999")
    # filters[] array needs city entry; fieldsValues[] has the actual filter values
    INF = "999999999999"
    city_config = {
        'montreal':  {"id": "5",        "matchType": "CityDistrictAll", "text": "Montréal (Tous les arrondissements)", "fieldId": "CityDistrictAll"},
        'longueuil': {"id": "4",        "matchType": "CityDistrictAll", "text": "Longueuil",                           "fieldId": "CityDistrictAll"},
        'laval':     {"id": "GSGS4622", "matchType": "GeographicArea",  "text": "Laval",                               "fieldId": "GeographicArea"},
    }
    city = city_config.get(city_slug, city_config['montreal'])

    fields = [
        {"fieldId": city["fieldId"],  "value": city["id"]},
        {"fieldId": "PropertyType",   "value": "MultiFamily"},
        {"fieldId": "Category",       "value": "Commercial"},
        {"fieldId": "SellingType",    "value": "Sale"},
    ]

    # Units: two entries (min, max)
    min_units = filters.get('min_units', 0)
    fields.append({"fieldId": "NumberUnits", "value": str(max(min_units, 0))})
    fields.append({"fieldId": "NumberUnits", "value": INF})

    # Price: two entries (min=0, max)
    max_price = filters.get('max_price', 0)
    fields.append({"fieldId": "SalePrice", "value": "0"})
    fields.append({"fieldId": "SalePrice", "value": str(max_price) if max_price else INF})

    # Date
    if filters.get('days', 0) > 0:
        since = (datetime.now() - timedelta(days=filters['days'])).strftime('%Y-%m-%d')
        fields.append({"fieldId": "LastModifiedDate", "value": since})

    payload = {
        "mls": "0", "brokerCode": "", "officeKey": "",
        "useGeographyShapes": 0, "shapeViews": [], "searchName": "",
        "filters": [{"matchType": city["matchType"], "text": city["text"], "id": city["id"]}],
        "fieldsValues": fields
    }

    json_bytes = json.dumps(payload, separators=(',', ':')).encode('utf-8')
    compressed = gzip.compress(json_bytes)
    q = base64.urlsafe_b64encode(compressed).rstrip(b'=').decode('ascii')

    return f"https://www.centris.ca/fr/multifamilial~a-vendre~{city_slug}?sort=DateDesc&q={q}&pageSize=20"


def _apply_filters(page, filters):
    """Filters are baked into the URL via _build_search_url - nothing to do here."""
    pass


def scrape_listing(page, url, filters):
    """Visit one listing page and extract financial data"""
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2000)

        text = page.evaluate("() => document.body.innerText")

        # Price
        price = None
        try:
            pel = page.locator("span.price, [class*='asking'] span, h2.price").first
            price = parse_num(pel.text_content(timeout=1500))
        except: pass
        if not price:
            price = find_num(text, [
                r'(\d[\d\s\xa0]{4,})\s*\$',
                r'\$\s*([\d\s\xa0]{5,})',
            ], lo=80_000, hi=50_000_000)
        if not price or price > filters['max_price']:
            return None

        # Address
        address = "?"
        try:
            address = page.locator("[itemprop='address']").first.text_content(timeout=1500).strip()
        except: pass

        # Units
        num_units = find_num(text, [
            r"Nombre d[''`]unités\s*:?\s*Résidentiel\s*\((\d+)\)",
            r"Résidentiel\s*\((\d+)\)",
            r"(\d+)\s*logements?",
            r"(\d+)\s*appartements?",
            r"(\d+)[- ]plex",
        ], lo=2, hi=500)
        if num_units: num_units = int(num_units)
        if num_units and num_units < filters['min_units']:
            return None

        # Gross income
        gross_income = find_num(text, [
            r'Revenus?\s*bruts?\s*potentiels\s*:?\s*\$?\s*([\d\s\xa0,]+)',
            r'Revenus?\s*bruts?\s*:?\s*\$?\s*([\d\s\xa0,]+)',
            r'Rev\.\s*bruts?\s*pot\.\s*:?\s*\$?\s*([\d\s\xa0,]+)',
            r'Revenus?\s*annuels?\s*:?\s*\$?\s*([\d\s\xa0,]+)',
        ], lo=1000, hi=5_000_000)

        # Taxes
        municipal_tax = find_num(text, [
            r'Municipales?\s*\(202[4-9]\)\s*:?\s*\$?\s*([\d\s\xa0,]+)',
            r'Taxes?\s*municipales?\s*:?\s*\$?\s*([\d\s\xa0,]+)',
            r'Impôts?\s*fonciers?\s*:?\s*\$?\s*([\d\s\xa0,]+)',
        ], lo=500, hi=500_000)

        school_tax = find_num(text, [
            r'Scolaires?\s*\(202[4-9]\)\s*:?\s*\$?\s*([\d\s\xa0,]+)',
            r'Taxes?\s*scolaires?\s*:?\s*\$?\s*([\d\s\xa0,]+)',
        ], lo=100, hi=100_000)

        return {
            'address'      : address,
            'price'        : price,
            'num_units'    : num_units,
            'gross_income' : gross_income,
            'municipal_tax': municipal_tax,
            'school_tax'   : school_tax,
            'url'          : url,
            'scraped_at'   : datetime.now().strftime('%Y-%m-%d %H:%M'),
        }
    except Exception as e:
        return None


# ── TGA ANALYSIS ──────────────────────────────────────────────────────────────
def calculate_tga(sc):
    r, n = sc['rate'], sc['years']
    f = (1+r)**n
    return (r*f/(f-1)) / 1.1

def analyze(prop):
    price = prop['price']
    gross = prop.get('gross_income')
    if not gross or gross <= 0: return None

    cfg = EXPENSE_CONFIG
    vacancy     = gross * cfg['vacancy_rate']
    eff         = gross - vacancy
    municipal   = prop.get('municipal_tax') or 0
    school      = prop.get('school_tax') or 0
    insurance   = gross * cfg['insurance_pct']
    maintenance = gross * cfg['maintenance_pct']
    management  = gross * cfg['management_pct']
    other       = gross * cfg['other_pct']
    total_exp   = municipal + school + insurance + maintenance + management + other
    noi         = eff - total_exp
    if noi <= 0: return None

    best_ratio, best_name, best = 0, None, {}
    all_sc = {}
    for name, sc in SCENARIOS.items():
        tga      = calculate_tga(sc)
        econ     = noi / tga
        ratio    = econ / price
        loan     = sc['rpv'] * min(price, econ)
        loan_tot = loan * (1 + sc['cmhc'])
        down     = price - loan_tot
        r, n_    = sc['rate']/12, sc['years']*12
        pmt      = loan_tot * (r*(1+r)**n_) / ((1+r)**n_-1)
        cf       = eff/12 - total_exp/12 - pmt
        roi      = (cf*12)/down if down > 0 else 0
        all_sc[name] = {'tga':tga,'econ':econ,'ratio':ratio,
                        'down':down,'pmt':pmt,'cf':cf,'roi':roi}
        if ratio > best_ratio:
            best_ratio, best_name, best = ratio, name, all_sc[name]

    return {**prop, 'noi':noi, 'total_exp':total_exp, 'exp_ratio':total_exp/gross,
            'best_scenario':best_name, 'best_ratio':best_ratio,
            'best_econ':best.get('econ'), 'best_down':best.get('down'),
            'best_pmt':best.get('pmt'), 'best_cf':best.get('cf'),
            'best_roi':best.get('roi'), 'all_sc':all_sc,
            'expenses':{'vacancy':vacancy,'municipal':municipal,'school':school,
                        'insurance':insurance,'maintenance':maintenance,
                        'management':management,'other':other}}


# ── EXCEL REPORT ──────────────────────────────────────────────────────────────
def generate_excel(analyses, filters):
    analyses = sorted(analyses, key=lambda x: x.get('best_ratio') or 0, reverse=True)
    wb = Workbook()

    HDR   = PatternFill('solid', fgColor='1F3864')
    GREEN = PatternFill('solid', fgColor='D9EAD3')
    YELL  = PatternFill('solid', fgColor='FFF2CC')
    RED   = PatternFill('solid', fgColor='F4CCCC')
    thin  = Side(style='thin', color='CCCCCC')
    BDR   = Border(left=thin,right=thin,top=thin,bottom=thin)

    def hdr_row(ws, cols, row=1):
        for c,h in enumerate(cols,1):
            cell = ws.cell(row=row,column=c,value=h)
            cell.font      = Font(name='Arial',bold=True,color='FFFFFF',size=9)
            cell.fill      = HDR
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            cell.border    = BDR
        ws.row_dimensions[row].height = 30

    def put(ws, row, col, val, fmt=None, fill=None, align='right', bold=False):
        c = ws.cell(row=row,column=col,value=val)
        c.font      = Font(name='Arial',size=9,bold=bold)
        c.alignment = Alignment(horizontal=align,vertical='center')
        c.border    = BDR
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill

    def rfill(ratio):
        if not ratio: return None
        return GREEN if ratio>=1.15 else YELL if ratio>=1.0 else RED

    def autosize(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2,8),50)

    scored = [a for a in analyses if a.get('noi')]

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Summary'
    ws.freeze_panes = 'A6'

    ws['A1'] = 'CENTRIS MULTIFAMILIAL — TGA ANALYSIS'
    ws['A1'].font = Font(name='Arial',bold=True,size=13)
    ws.merge_cells('A1:Q1')

    ws['A2'] = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                f"Cities: {', '.join(l for l,_ in filters['cities'])}  |  "
                f"Max Price: ${filters['max_price']:,}  |  "
                f"Min Units: {filters['min_units']}+  |  "
                f"Last {filters['days']} days")
    ws['A2'].font = Font(name='Arial',size=9,italic=True)
    ws.merge_cells('A2:Q2')

    good = sum(1 for a in scored if a['best_ratio']>=1.15)
    ok   = sum(1 for a in scored if 1.0<=a['best_ratio']<1.15)
    bad  = sum(1 for a in scored if a['best_ratio']<1.0)
    ws['A3'] = (f"🟢 Excellent (≥115%): {good}   "
                f"🟡 Good (100-115%): {ok}   "
                f"🔴 Overpriced (<100%): {bad}   "
                f"({len(analyses)-len(scored)} missing income data)")
    ws['A3'].font = Font(name='Arial',bold=True,size=10)
    ws.merge_cells('A3:Q3')

    COLS = ['Address','City','Price','Units','Gross Income',
            'Muni Tax','School Tax','NOI',
            'Economic Value','Value Ratio','Best Scenario',
            'TGA','Down Payment','Monthly Pmt','Monthly CF',
            'Cash ROI','URL']
    hdr_row(ws, COLS, row=5)

    for r, a in enumerate(analyses, 6):
        fill = rfill(a.get('best_ratio'))
        sc   = a.get('all_sc',{}).get(a.get('best_scenario') or '',{})
        vals = [
            a.get('address',''), a.get('city',''),
            a.get('price'), a.get('num_units'),
            a.get('gross_income') or 0,
            a.get('municipal_tax') or 0, a.get('school_tax') or 0,
            a.get('noi'), a.get('best_econ'), a.get('best_ratio'),
            a.get('best_scenario'), sc.get('tga'),
            a.get('best_down'), a.get('best_pmt'), a.get('best_cf'),
            a.get('best_roi'), a.get('url','')
        ]
        fmts = [None,None,'$#,##0','0','$#,##0',
                '$#,##0','$#,##0','$#,##0',
                '$#,##0','0.0%',None,'0.00%',
                '$#,##0','$#,##0','$#,##0_);($#,##0)','0.0%',None]
        als  = ['left','left']+['right']*13+['right','right','left']
        for c,(v,f,al) in enumerate(zip(vals,fmts,als),1):
            put(ws, r, c, v, fmt=f, fill=fill, align=al)

    autosize(ws)
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['Q'].width = 60

    # ── Sheet 2: Expense Detail ───────────────────────────────────────────
    ws2 = wb.create_sheet('Expense Detail')
    ws2.freeze_panes = 'A2'
    H2 = ['Address','Gross Income','Vacancy (3%)','Muni Tax','School Tax',
          'Insurance (4%)','Maintenance (3%)','Management (5%)','Other (2%)',
          'Total Expenses','Expense Ratio','NOI']
    hdr_row(ws2, H2)
    for r,a in enumerate(scored, 2):
        e = a.get('expenses',{})
        fill = rfill(a['best_ratio'])
        vals = [a.get('address',''), a.get('gross_income',0), e.get('vacancy',0),
                e.get('municipal',0), e.get('school',0), e.get('insurance',0),
                e.get('maintenance',0), e.get('management',0), e.get('other',0),
                a.get('total_exp',0), a.get('exp_ratio'), a.get('noi',0)]
        fmts = ['@']+['$#,##0']*9+['0.0%','$#,##0']
        als  = ['left']+['right']*11
        for c,(v,f,al) in enumerate(zip(vals,fmts,als),1):
            put(ws2, r, c, v, fmt=f, fill=fill, align=al)
    autosize(ws2)
    ws2.column_dimensions['A'].width = 45

    # ── Sheet 3: All Scenarios ────────────────────────────────────────────
    ws3 = wb.create_sheet('All Scenarios')
    ws3.freeze_panes = 'A2'
    H3 = ['Address','Scenario','Rate','Years','TGA',
          'Economic Value','Value Ratio','Down Payment',
          'Monthly Pmt','Monthly CF','Cash ROI']
    hdr_row(ws3, H3)
    row = 2
    for a in scored:
        for sn, sc in a.get('all_sc',{}).items():
            fill = rfill(sc['ratio'])
            vals = [a.get('address',''), sn, sc['tga']*100, sc.get('years',0) if 'years' not in sc else SCENARIOS.get(sn,{}).get('years'),
                    sc['tga'], sc['econ'], sc['ratio'],
                    sc['down'], sc['pmt'], sc['cf'], sc['roi']]
            fmts = ['@','@','0.0%','0','0.00%',
                    '$#,##0','0.0%','$#,##0',
                    '$#,##0','$#,##0_);($#,##0)','0.0%']
            als  = ['left','left']+['right']*9
            for c,(v,f,al) in enumerate(zip(vals,fmts,als),1):
                put(ws3, row, c, v, fmt=f, fill=fill, align=al)
            row += 1
        row += 1  # blank separator
    autosize(ws3)
    ws3.column_dimensions['A'].width = 45

    # ── Sheet 4: Raw Data ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('Raw Data')
    RCOLS = ['address','city','price','num_units','gross_income',
             'municipal_tax','school_tax','url','scraped_at']
    hdr_row(ws4, RCOLS)
    for r,a in enumerate(analyses, 2):
        for c,col in enumerate(RCOLS, 1):
            cell = ws4.cell(row=r,column=c,value=a.get(col,''))
            cell.font   = Font(name='Arial',size=9)
            cell.border = BDR
            cell.alignment = Alignment(horizontal='right' if c>2 else 'left')
    autosize(ws4)
    ws4.column_dimensions['A'].width = 45
    ws4.column_dimensions['H'].width = 60

    fname = f"Centris_Plex_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    return fname


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    filters = ask_filters()

    print(f"\n{'═'*55}")
    print("  SCRAPING...")
    print(f"{'═'*55}")
    properties = scrape_all(filters)

    print(f"\n{'═'*55}")
    print(f"  Scraped {len(properties)} properties")
    print(f"{'═'*55}")

    if not properties:
        print("\n⚠ No properties found. Try adjusting your filters.")
        input("\nPress ENTER to exit...")
        return

    print(f"\n{'═'*55}")
    print("  RUNNING TGA ANALYSIS...")
    print(f"{'═'*55}\n")

    analyses, no_income = [], []
    for prop in properties:
        if prop.get('gross_income'):
            result = analyze(prop)
            if result:
                analyses.append(result)
                r    = result['best_ratio']
                icon = '🟢' if r>=1.15 else '🟡' if r>=1.0 else '🔴'
                print(f"  {icon} {prop.get('address','?')[:48]}")
                print(f"     Ratio: {r:.1%}  CF: ${result['best_cf']:,.0f}/mo  ROI: {result['best_roi']:.1%}")
            else:
                no_income.append(prop)
        else:
            no_income.append(prop)
            print(f"  ⚪ {prop.get('address','?')[:48]} — no income data")

    # Add no-income props for raw data sheet
    for p in no_income:
        analyses.append({**p, 'best_ratio':0, 'noi':None, 'total_exp':None,
                         'exp_ratio':None, 'best_scenario':None,
                         'best_econ':None, 'best_down':None,
                         'best_pmt':None, 'best_cf':None, 'best_roi':None,
                         'all_sc':{}, 'expenses':{}})

    print(f"\n{'═'*55}")
    print("  GENERATING EXCEL...")
    print(f"{'═'*55}")
    fname = generate_excel(analyses, filters)

    print(f"\n  ✓ Saved: {fname}")

    scored = [a for a in analyses if a.get('noi')]
    if scored:
        best = max(scored, key=lambda x: x['best_ratio'])
        print(f"\n  🏆 BEST DEAL:")
        print(f"  {best.get('address','?')}")
        print(f"  Price:    ${best['price']:,.0f}")
        print(f"  Value:    ${best['best_econ']:,.0f}")
        print(f"  Ratio:    {best['best_ratio']:.1%}")
        print(f"  CF:       ${best['best_cf']:,.0f}/mo")
        print(f"  URL: {best.get('url','')}")

    print()
    input("Press ENTER to exit...")

if __name__ == '__main__':
    main()
