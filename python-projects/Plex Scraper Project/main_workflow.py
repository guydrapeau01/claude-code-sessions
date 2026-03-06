#!/usr/bin/env python3
"""
Complete Workflow for Montreal/Laval Plex Analysis
Integrates scraping, analysis, and reporting
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

from property_analyzer import PropertyAnalyzer, FINANCING_SCENARIOS
from property_scraper import ManualPropertyLoader, PropertyScraper


class ExcelReportGenerator:
    """Generate detailed Excel reports with multiple sheets"""
    
    def __init__(self, filename: str = None):
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Property_Analysis_Report_{timestamp}.xlsx'
        
        self.filename = filename
        self.wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
    
    def create_summary_sheet(self, summary_df: pd.DataFrame):
        """Create summary overview sheet"""
        
        ws = self.wb.create_sheet('Summary', 0)
        
        # Add title
        ws['A1'] = 'PROPERTY ANALYSIS SUMMARY'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        
        # Add summary data
        start_row = 4
        
        # Headers
        headers = list(summary_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data
        for row_idx, row in enumerate(summary_df.values, start=start_row+1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Format numbers
                col_name = headers[col_idx-1]
                if any(x in col_name for x in ['Price', 'Value', 'Income', 'Payment', 'Cashflow', 'Down']):
                    cell.number_format = '$#,##0'
                elif 'Ratio' in col_name or 'ROI' in col_name or 'TGA' in col_name:
                    cell.number_format = '0.00%'
                elif 'MRB' in col_name or 'Per Unit' in col_name:
                    cell.number_format = '#,##0.00'
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def create_detailed_sheet(self, property_analysis: dict):
        """Create detailed analysis sheet for a single property"""
        
        address = property_analysis['address'][:30]  # Truncate for sheet name
        ws = self.wb.create_sheet(address)
        
        # Property info
        ws['A1'] = 'DETAILED PROPERTY ANALYSIS'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        ws[f'A{row}'] = 'Address:'
        ws[f'B{row}'] = property_analysis['address']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Market Price:'
        ws[f'B{row}'] = property_analysis['market_price']
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Number of Units:'
        ws[f'B{row}'] = property_analysis['num_units']
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Gross Annual Income:'
        ws[f'B{row}'] = property_analysis['gross_annual_income']
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'A{row}'].font = Font(bold=True)
        
        # Scenario comparison
        row += 3
        ws[f'A{row}'] = 'FINANCING SCENARIO COMPARISON'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 2
        
        # Create comparison table
        scenario_headers = ['Metric'] + list(property_analysis['scenarios'].keys())
        for col_idx, header in enumerate(scenario_headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Metrics to display
        metrics = [
            ('Economic Value', 'economic_value', '$#,##0'),
            ('Value Ratio', 'value_ratio', '0.00%'),
            ('TGA', 'tga', '0.00%'),
            ('Down Payment', 'down_payment', '$#,##0'),
            ('Down Payment %', 'down_payment_pct', '0.00%'),
            ('Monthly Payment', 'monthly_payment', '$#,##0'),
            ('Monthly Cashflow', 'monthly_cashflow', '$#,##0'),
            ('Annual Cashflow', 'annual_cashflow', '$#,##0'),
            ('Cash ROI', 'cash_roi', '0.00%'),
            ('NOI', 'noi', '$#,##0'),
            ('Value per Unit', 'value_per_unit', '$#,##0'),
            ('MRB', 'mrb', '0.00'),
            ('MRN', 'mrn', '0.00')
        ]
        
        row += 1
        for metric_name, metric_key, number_format in metrics:
            ws.cell(row=row, column=1).value = metric_name
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for col_idx, scenario_name in enumerate(property_analysis['scenarios'].keys(), start=2):
                scenario = property_analysis['scenarios'][scenario_name]
                cell = ws.cell(row=row, column=col_idx)
                cell.value = scenario.get(metric_key, 0)
                cell.number_format = number_format
            
            row += 1
        
        # Highlight best scenario
        best_scenario = property_analysis['best_scenario']
        best_col = list(property_analysis['scenarios'].keys()).index(best_scenario) + 2
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def create_parameters_sheet(self):
        """Create a sheet showing analysis parameters"""
        
        ws = self.wb.create_sheet('Parameters')
        
        ws['A1'] = 'ANALYSIS PARAMETERS'
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        ws[f'A{row}'] = 'Vacancy Rate:'
        ws[f'B{row}'] = 0.03
        ws[f'B{row}'].number_format = '0.0%'
        
        row += 1
        ws[f'A{row}'] = 'Operating Expense Ratio:'
        ws[f'B{row}'] = 0.25
        ws[f'B{row}'].number_format = '0.0%'
        
        row += 3
        ws[f'A{row}'] = 'FINANCING SCENARIOS'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        row += 2
        
        # Scenario table
        headers = ['Scenario', 'RPV', 'CMHC Premium', 'Amortization', 'Interest Rate', 'RCD']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
        
        row += 1
        for scenario_name, scenario in FINANCING_SCENARIOS.items():
            ws.cell(row=row, column=1).value = scenario.name
            ws.cell(row=row, column=2).value = scenario.rpv
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=3).value = scenario.prime_schl_pct
            ws.cell(row=row, column=3).number_format = '0.00%'
            ws.cell(row=row, column=4).value = f"{scenario.amortization_years} years"
            ws.cell(row=row, column=5).value = scenario.interest_rate
            ws.cell(row=row, column=5).number_format = '0.00%'
            ws.cell(row=row, column=6).value = scenario.rcd
            row += 1
    
    def save(self):
        """Save the workbook"""
        self.wb.save(self.filename)
        print(f"\nReport saved to: {self.filename}")
        return self.filename


def run_complete_analysis(properties: list, output_filename: str = None):
    """
    Run complete analysis workflow
    
    Args:
        properties: List of property dicts with keys:
                   - address
                   - market_price
                   - num_units
                   - gross_annual_income (optional, will estimate if missing)
    """
    
    print("="*80)
    print("MONTREAL/LAVAL PLEX ANALYZER")
    print("="*80)
    
    # Initialize analyzer
    analyzer = PropertyAnalyzer(vacancy_rate=0.03, expense_ratio=0.25)
    scraper = PropertyScraper()
    
    # Process properties
    results = []
    
    for prop in properties:
        # Estimate income if not provided
        if not prop.get('gross_annual_income') or prop['gross_annual_income'] == 0:
            estimated_income = scraper.estimate_income(
                prop['num_units'],
                prop.get('address', 'montreal')
            )
            prop['gross_annual_income'] = estimated_income
            print(f"\nEstimated income for {prop['address']}: ${estimated_income:,.0f}/year")
        
        # Analyze property
        analysis = analyzer.analyze_property(**prop)
        results.append(analysis)
    
    # Create summary report
    summary_df = analyzer.create_summary_report(results)
    
    # Generate Excel report
    report = ExcelReportGenerator(output_filename)
    report.create_summary_sheet(summary_df)
    
    for prop_analysis in results:
        report.create_detailed_sheet(prop_analysis)
    
    report.create_parameters_sheet()
    
    output_file = report.save()
    
    # Print summary to console
    print("\n" + "="*80)
    print("ANALYSIS SUMMARY")
    print("="*80 + "\n")
    print(summary_df.to_string(index=False))
    
    # Print top opportunities
    print("\n" + "="*80)
    print("TOP OPPORTUNITIES (by Value Ratio)")
    print("="*80)
    
    top_5 = summary_df.head(5)
    for idx, row in top_5.iterrows():
        print(f"\n{row['Address']}")
        print(f"  Market Price: ${row['Market Price']:,.0f}")
        print(f"  Economic Value: ${row['Economic Value']:,.0f}")
        print(f"  Value Ratio: {row['Value Ratio']:.1%} ({'GOOD DEAL' if row['Value Ratio'] > 1.0 else 'OVERPRICED'})")
        print(f"  Monthly Cashflow: ${row['Monthly Cashflow']:,.0f}")
        print(f"  Cash ROI: {row['Cash ROI']:.1%}")
    
    return results, summary_df, output_file


def main():
    """Main entry point with example usage"""
    
    # Example 1: Load from CSV
    print("OPTION 1: Load properties from CSV file")
    print("-" * 40)
    
    # Check if template exists
    if not os.path.exists('property_input_template.csv'):
        ManualPropertyLoader.save_template('property_input_template.csv')
    
    # Example 2: Define properties programmatically
    print("\nOPTION 2: Analyze sample properties")
    print("-" * 40)
    
    sample_properties = [
        {
            'address': '1825, Rue Sainte-Hélène, Longueuil',
            'market_price': 1199999,
            'num_units': 6,
            'gross_annual_income': 84660
        },
        {
            'address': '1, Rue des Pins, Chambly',
            'market_price': 839000,
            'num_units': 6,
            'gross_annual_income': 57672
        },
        {
            'address': '69-71A, Rue Saint-Charles, Saint-Jean-sur-Richelieu',
            'market_price': 947900,
            'num_units': 6,
            'gross_annual_income': 68220
        },
        {
            'address': '263-265, Rue Saint-Charles, La Prairie',
            'market_price': 938000,
            'num_units': 6,
            'gross_annual_income': 72600
        }
    ]
    
    # Run analysis
    results, summary, output_file = run_complete_analysis(
        sample_properties,
        'Montreal_Plex_Analysis.xlsx'
    )
    
    print(f"\n{'='*80}")
    print(f"Analysis complete! Report saved to: {output_file}")
    print(f"{'='*80}\n")


if __name__ == '__main__':
    main()
