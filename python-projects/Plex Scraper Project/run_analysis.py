#!/usr/bin/env python3
"""
run_analysis.py
===============
ONE SCRIPT TO RULE THEM ALL

What this does:
  1. Scrapes Centris for 5+ unit plexes in Montreal/Laval under $2M
  2. Saves scraped properties to scraped_properties.csv
  3. Pauses and lets you add/edit properties manually if needed
  4. Runs full TGA economic value analysis
  5. Generates one Excel report

Usage:
  python run_analysis.py
"""

import time
import re
import os
import sys
import pandas as pd
from datetime import datetime


# ============================================================================
# CONFIGURATION - Edit these settings
# ============================================================================

CONFIG = {
    # Search filters
    'min_units'   : 5,
    'max_price'   : 2_000_000,
    'cities'      : ['montreal', 'laval'],

    # Expense assumptions
    'vacancy_rate'    : 0.03,   # 3%
    'insurance_pct'   : 0.04,   # 4% of gross rents
    'maintenance_pct' : 0.03,   # 3%
    'management_pct'  : 0.05,   # 5%
    'other_pct'       : 0.02,   # 2%

    # Browser
    'headless'               : False,  # False = see browser (recommended)
    'delay_between_listings' : 3,      # seconds between each listing
}

# TGA Financing Scenarios
SCENARIOS = {
    '100pts': {'rpv': 0.95, 'cmhc': 0.0255, 'years': 50, 'rate': 0.04},
    '70pts' : {'rpv': 0.95, 'cmhc': 0.033,  'years': 45, 'rate': 0.04},
    '50pts' : {'rpv': 0.85, 'cmhc': 0.033,  'years': 40, 'rate': 0.04},
    'SCHL'  : {'rpv': 0.80, 'cmhc': 0.055,  'years': 40, 'rate': 0.039},
    'Conv'  : {'rpv': 0.75, 'cmhc': 0.055,  'years': 40, 'rate': 0.054},
}

CSV_FILE = 'scraped_properties.csv'


# ============================================================================
# STEP 1 - SCRAPE CENTRIS
# ============================================================================

def scrape_centris():
    """Open Centris, search for plexes, return list of property dicts"""

    print("\n" + "="*60)
    print("STEP 1: SCRAPING CENTRIS")
    print("="*60)

    try:
        from selenium import webdriver
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.chrome.options import Options
        from selenium.common.exceptions import TimeoutException
    except ImportError:
        print("ERROR: selenium not installed.")
        print("  Run: python -m pip install selenium")
        return []

    driver = _create_driver()
    if not driver:
        return []

    scraped = []

    try:
        search_urls = [
            "https://www.centris.ca/fr/plex~a-vendre~montreal?view=Thumbnail",
            "https://www.centris.ca/fr/plex~a-vendre~laval?view=Thumbnail",
        ]

        listing_urls = []
        for search_url in search_urls:
            city_urls = _collect_listing_urls(driver, search_url)
            listing_urls.extend(city_urls)

        listing_urls = list(set(listing_urls))
        print(f"\nFound {len(listing_urls)} listings to visit")

        if not listing_urls:
            print("No listings found - Centris may have blocked the search.")
            print("Script will continue with any manually added properties.")
            return []

        print(f"\nVisiting each listing individually...")
        print("-" * 60)

        for i, url in enumerate(listing_urls, 1):
            print(f"[{i}/{len(listing_urls)}] {url[-55:]}")
            data = _scrape_listing(driver, url)

            if data:
                price = data.get('price') or 0
                units = data.get('num_units') or 0

                if price > CONFIG['max_price']:
                    print(f"  -> Skipped: price ${price:,.0f} over limit")
                    continue
                if units and units < CONFIG['min_units']:
                    print(f"  -> Skipped: only {units} units")
                    continue

                scraped.append(data)
                print(f"  OK  {data.get('address','?')[:45]}")
                print(f"      Price: ${price:>10,.0f} | "
                      f"Units: {units or '?'} | "
                      f"Income: ${data.get('gross_income') or 0:>8,.0f} | "
                      f"Taxes: ${(data.get('municipal_tax') or 0) + (data.get('school_tax') or 0):>7,.0f}")
            else:
                print(f"  SKIP  Could not extract data")

            time.sleep(CONFIG['delay_between_listings'])

    finally:
        driver.quit()
        print("\nBrowser closed")

    return scraped


def _create_driver():
    """Start Chrome with anti-detection settings"""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options

        options = Options()
        if CONFIG['headless']:
            options.add_argument('--headless=new')

        options.add_argument('--window-size=1920,1080')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--lang=fr-CA,fr,en')
        options.add_argument(
            'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/120.0.0.0 Safari/537.36'
        )
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)

        try:
            from selenium.webdriver.chrome.service import Service
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            print("Chrome started (ChromeDriver auto-installed)")
        except ImportError:
            driver = webdriver.Chrome(options=options)
            print("Chrome started")

        driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        return driver

    except Exception as e:
        print(f"Could not start Chrome: {e}")
        print("Make sure Google Chrome is installed: https://www.google.com/chrome/")
        return None


def _collect_listing_urls(driver, search_url):
    """Go to search results page and collect all listing URLs"""

    print(f"\nSearching: {search_url}")
    urls = []

    try:
        driver.get(search_url)
        time.sleep(4)

        # Accept cookies if prompted
        try:
            btn = driver.find_element(
                'xpath',
                "//button[contains(text(),'Accepter') or contains(text(),'Accept')]"
            )
            btn.click()
            time.sleep(1)
        except Exception:
            pass

        page = 1
        while page <= 15:
            _scroll_page(driver)
            page_urls = _extract_urls_from_page(driver)
            new = [u for u in page_urls if u not in urls]
            if not new:
                break
            urls.extend(new)
            print(f"  Page {page}: {len(new)} listings ({len(urls)} total)")

            try:
                next_btn = driver.find_element(
                    'xpath',
                    "//li[contains(@class,'next') and not(contains(@class,'disabled'))]/a"
                    " | //a[@aria-label='Suivant' or @aria-label='Next']"
                )
                next_btn.click()
                time.sleep(3)
                page += 1
            except Exception:
                break

    except Exception as e:
        print(f"  Error: {e}")

    return urls


def _extract_urls_from_page(driver):
    """Extract listing URLs from current search results page"""

    urls = []
    selectors = [
        "a.property-thumbnail-summary-link",
        "div.property-thumbnail-item a",
        ".shell a",
    ]

    for sel in selectors:
        try:
            elements = driver.find_elements('css selector', sel)
            for el in elements:
                href = el.get_attribute('href') or ''
                if '/plex~' in href or '/plex/' in href:
                    if href not in urls:
                        urls.append(href)
            if urls:
                break
        except Exception:
            continue

    if not urls:
        try:
            for link in driver.find_elements('tag name', 'a'):
                href = link.get_attribute('href') or ''
                if 'centris' in href and '/plex' in href:
                    if href not in urls:
                        urls.append(href)
        except Exception:
            pass

    return urls


def _scroll_page(driver):
    """Scroll to trigger lazy loading"""
    try:
        last = driver.execute_script("return document.body.scrollHeight")
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1.5)
            new = driver.execute_script("return document.body.scrollHeight")
            if new == last:
                break
            last = new
        driver.execute_script("window.scrollTo(0, 0);")
    except Exception:
        pass


def _scrape_listing(driver, url):
    """Visit one listing page and extract all fields"""

    try:
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC

        driver.get(url)
        time.sleep(2)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(('tag name', 'h1'))
        )

        text = driver.find_element('tag name', 'body').text

        return {
            'address'      : _get_address(driver, text),
            'price'        : _get_price(driver, text),
            'num_units'    : _get_units(text),
            'gross_income' : _get_income(text),
            'municipal_tax': _get_municipal_tax(text),
            'school_tax'   : _get_school_tax(text),
            'utilities'    : _get_utilities(text),
            'url'          : url,
            'scraped_at'   : datetime.now().strftime('%Y-%m-%d %H:%M'),
        }

    except Exception:
        return None


# ---- Field extractors ----

def _parse_num(text):
    if not text:
        return None
    try:
        cleaned = re.sub(r'[^\d.]', '', str(text).replace('\xa0', '').replace(' ', ''))
        return float(cleaned) if cleaned else None
    except Exception:
        return None


def _get_price(driver, text):
    for sel in ["span.price", "div.price", "[class*='price']"]:
        try:
            v = _parse_num(driver.find_element('css selector', sel).text)
            if v and v > 100_000:
                return v
        except Exception:
            pass
    for pat in [r'Prix demandé[:\s]*\$?\s*([\d\s,\xa0]+)',
                r'Asking price[:\s]*\$?\s*([\d\s,\xa0]+)',
                r'(\d[\d\s]{4,})\s*\$']:
        m = re.search(pat, text, re.I)
        if m:
            v = _parse_num(m.group(1))
            if v and 100_000 < v < 10_000_000:
                return v
    return None


def _get_address(driver, text):
    for sel in ["h1", "span[itemprop='streetAddress']", ".address"]:
        try:
            t = driver.find_element('css selector', sel).text.strip()
            if t and len(t) > 5:
                return t
        except Exception:
            pass
    m = re.search(
        r'\d+[,\s]+(?:Rue|Avenue|Boulevard|Boul\.|Ave\.|Ch\.)[^\n]{3,50}',
        text, re.I
    )
    return m.group(0).strip() if m else "Address not found"


def _get_units(text):
    for pat in [r'(\d+)\s*logements?', r'(\d+)\s*units?',
                r'(\d+)\s*appartements?', r'(\d+)[\s-]plex',
                r'Nombre de logements?\s*:?\s*(\d+)']:
        m = re.search(pat, text, re.I)
        if m:
            n = int(m.group(1))
            if 2 <= n <= 50:
                return n
    return None


def _get_income(text):
    for pat in [r'Revenus? bruts?\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'Gross revenue\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'Revenus? annuels?\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'Total loyers\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'Loyers\s*:?\s*\$?\s*([\d\s,\xa0]+)']:
        m = re.search(pat, text, re.I)
        if m:
            v = _parse_num(m.group(1))
            if v and 1_000 < v < 1_000_000:
                return v
    return None


def _get_municipal_tax(text):
    for pat in [r'[Tt]axes? municipales?\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'Municipal tax\w*\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'[Ii]mpôts? fonciers?\s*:?\s*\$?\s*([\d\s,\xa0]+)']:
        m = re.search(pat, text, re.I)
        if m:
            v = _parse_num(m.group(1))
            if v and 500 < v < 100_000:
                return v
    return None


def _get_school_tax(text):
    for pat in [r'[Tt]axes? scolaires?\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'School tax\w*\s*:?\s*\$?\s*([\d\s,\xa0]+)']:
        m = re.search(pat, text, re.I)
        if m:
            v = _parse_num(m.group(1))
            if v and 100 < v < 20_000:
                return v
    return None


def _get_utilities(text):
    for pat in [r'[Ee]nergie\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'[Cc]hauffage\s*:?\s*\$?\s*([\d\s,\xa0]+)',
                r'[Uu]tilities?\s*:?\s*\$?\s*([\d\s,\xa0]+)']:
        m = re.search(pat, text, re.I)
        if m:
            v = _parse_num(m.group(1))
            if v and 50 < v < 10_000:
                return v
    return None


# ============================================================================
# STEP 2 - SAVE TO CSV AND PAUSE FOR MANUAL EDITS
# ============================================================================

def save_to_csv(properties):
    """Save scraped properties to CSV"""

    columns = ['address','price','num_units','gross_income',
               'municipal_tax','school_tax','utilities','url','scraped_at']

    if os.path.exists(CSV_FILE):
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                existing = pd.read_csv(CSV_FILE, encoding=enc)
                break
            except UnicodeDecodeError:
                continue

        existing_urls = set(existing['url'].dropna().tolist())
        new_props = [p for p in properties if p.get('url') not in existing_urls]

        if new_props:
            new_df   = pd.DataFrame(new_props, columns=columns)
            combined = pd.concat([existing, new_df], ignore_index=True)
            combined.to_csv(CSV_FILE, index=False, encoding='utf-8-sig')
            print(f"Added {len(new_props)} new properties to {CSV_FILE}")
        else:
            print(f"No new properties (CSV already up to date)")
    else:
        df = pd.DataFrame(properties, columns=columns)
        df.to_csv(CSV_FILE, index=False, encoding='utf-8-sig')
        print(f"Saved {len(properties)} properties to {CSV_FILE}")


def pause_for_manual_edits():
    """Pause and let user fill in missing data"""

    print("\n" + "="*60)
    print("STEP 2: REVIEW & EDIT CSV  (optional but important!)")
    print("="*60)
    print(f"""
Open '{CSV_FILE}' in Excel to:

  1. Fill in missing 'gross_income' for any listings
     where Centris did not show it publicly
     (estimate: nb_units x avg_rent x 12)

  2. Add any properties manually (DuProprio, off-market, etc.)

  3. Delete rows you're not interested in

Column reference:
  address        Property full address
  price          Asking price ($)
  num_units      Number of units
  gross_income   Total annual rent ($)  <-- most important
  municipal_tax  Annual municipal tax ($)
  school_tax     Annual school tax ($)
  utilities      Monthly utilities ($, 0 if tenant-paid)
  url            Centris listing link
""")
    input("  Press ENTER when you are done editing the CSV...")
    print()


# ============================================================================
# STEP 3 - LOAD CSV
# ============================================================================

def load_csv():
    """Load and validate properties from CSV"""

    if not os.path.exists(CSV_FILE):
        print(f"ERROR: {CSV_FILE} not found")
        return []

    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        try:
            df = pd.read_csv(CSV_FILE, encoding=enc)
            break
        except UnicodeDecodeError:
            continue

    properties = []
    skipped    = []

    for _, row in df.iterrows():
        addr  = row.get('address')
        price = row.get('price')
        inc   = row.get('gross_income')

        if pd.isna(addr) or pd.isna(price):
            continue

        if pd.isna(inc) or float(inc) == 0:
            skipped.append(str(addr)[:50])
            continue

        def safe(col):
            v = row.get(col, 0)
            return float(v) if not pd.isna(v) else 0.0

        properties.append({
            'address'      : addr,
            'price'        : float(price),
            'num_units'    : int(safe('num_units')),
            'gross_income' : float(inc),
            'municipal_tax': safe('municipal_tax'),
            'school_tax'   : safe('school_tax'),
            'utilities'    : safe('utilities'),
            'url'          : row.get('url', ''),
        })

    print(f"Loaded   : {len(properties)} properties with income data")
    if skipped:
        print(f"Skipped  : {len(skipped)} properties missing gross_income:")
        for s in skipped:
            print(f"           - {s}")

    return properties


# ============================================================================
# STEP 4 - ECONOMIC VALUE ANALYSIS
# ============================================================================

def calculate_tga(scenario):
    r, n = scenario['rate'], scenario['years']
    f = (1 + r) ** n
    return (r * f / (f - 1)) / 1.1   # 1.1 = RCD


def analyze_property(prop):
    """Run full TGA analysis on one property"""

    price = prop['price']
    gross = prop['gross_income']

    # Expenses
    vacancy      = gross * CONFIG['vacancy_rate']
    eff_income   = gross - vacancy
    municipal    = prop.get('municipal_tax', 0) or 0
    school       = prop.get('school_tax', 0)    or 0
    utilities    = (prop.get('utilities', 0) or 0) * 12
    insurance    = gross * CONFIG['insurance_pct']
    maintenance  = gross * CONFIG['maintenance_pct']
    management   = gross * CONFIG['management_pct']
    other        = gross * CONFIG['other_pct']
    total_exp    = municipal + school + utilities + insurance + maintenance + management + other
    noi          = eff_income - total_exp

    if noi <= 0:
        return None

    best_name, best_ratio, best = None, 0, {}
    all_scenarios = {}

    for name, sc in SCENARIOS.items():
        tga    = calculate_tga(sc)
        econ   = noi / tga
        ratio  = econ / price

        loan_base   = min(price, econ)
        loan        = sc['rpv'] * loan_base
        cmhc        = loan * sc['cmhc']
        total_loan  = loan + cmhc
        down        = price - total_loan

        r  = sc['rate'] / 12
        n  = sc['years'] * 12
        pmt = total_loan * (r * (1+r)**n) / ((1+r)**n - 1)
        cf  = eff_income / 12 - total_exp / 12 - pmt
        roi = (cf * 12) / down if down > 0 else 0

        all_scenarios[name] = {
            'tga': tga, 'economic_value': econ, 'value_ratio': ratio,
            'down_payment': down, 'monthly_payment': pmt,
            'monthly_cashflow': cf, 'cash_roi': roi,
        }

        if ratio > best_ratio:
            best_ratio, best_name, best = ratio, name, all_scenarios[name]

    return {
        **prop,
        'noi'             : noi,
        'total_expenses'  : total_exp,
        'expense_ratio'   : total_exp / gross,
        'taxes_total'     : municipal + school,
        'best_scenario'   : best_name,
        'best_value_ratio': best_ratio,
        'best_econ_value' : best['economic_value'],
        'best_tga'        : best['tga'],
        'best_down'       : best['down_payment'],
        'best_monthly_pmt': best['monthly_payment'],
        'best_monthly_cf' : best['monthly_cashflow'],
        'best_cash_roi'   : best['cash_roi'],
        'all_scenarios'   : all_scenarios,
        'expense_detail'  : {
            'vacancy': vacancy, 'municipal': municipal, 'school': school,
            'utilities': utilities, 'insurance': insurance,
            'maintenance': maintenance, 'management': management, 'other': other,
        }
    }


# ============================================================================
# STEP 5 - GENERATE EXCEL REPORT
# ============================================================================

def generate_report(analyses):
    """Create formatted Excel workbook with 4 sheets"""

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    filename = f"Plex_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb       = openpyxl.Workbook()

    GREEN  = PatternFill('solid', fgColor='D9EAD3')
    YELLOW = PatternFill('solid', fgColor='FFF2CC')
    RED    = PatternFill('solid', fgColor='F4CCCC')
    BLUE   = PatternFill('solid', fgColor='366092')

    def hdr(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = BLUE
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[row].height = 28

    def autosize(ws):
        for col in ws.columns:
            w = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 45)

    def color_row(ws, row_idx, ratio, num_cols):
        fill = GREEN if ratio >= 1.15 else YELLOW if ratio >= 1.0 else RED
        for c in range(1, num_cols + 1):
            ws.cell(row=row_idx, column=c).fill = fill

    analyses = sorted(analyses, key=lambda x: x['best_value_ratio'], reverse=True)

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = 'Summary'
    good = sum(1 for a in analyses if a['best_value_ratio'] >= 1.0)
    ws1['A1'] = f"Montreal / Laval Plex Analysis  -  {datetime.now().strftime('%Y-%m-%d')}"
    ws1['A1'].font = Font(size=13, bold=True)
    ws1['A2'] = (f"{len(analyses)} properties  |  {good} good deals (>=1.0)  |  "
                 f"Sorted best to worst")

    H1 = ['Address','Price','Units','Gross Income',
          'Municipal Tax','School Tax','Total Taxes',
          'NOI','Economic Value','Value Ratio',
          'Best Scenario','TGA','Down Payment',
          'Monthly Payment','Monthly Cashflow','Cash ROI',
          'Expense Ratio','Listing URL']
    hdr(ws1, H1, row=4)

    for r, a in enumerate(analyses, 5):
        ratio = a['best_value_ratio']
        vals  = [
            a['address'], a['price'], a['num_units'], a['gross_income'],
            a.get('municipal_tax') or 0, a.get('school_tax') or 0, a['taxes_total'],
            a['noi'], a['best_econ_value'], ratio,
            a['best_scenario'], a['best_tga'], a['best_down'],
            a['best_monthly_pmt'], a['best_monthly_cf'], a['best_cash_roi'],
            a['expense_ratio'], a.get('url', '')
        ]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=r, column=c, value=v)
        color_row(ws1, r, ratio, len(H1))
        for c in [2,4,5,6,7,8,9,13,14,15]:
            ws1.cell(row=r, column=c).number_format = '$#,##0'
        for c in [10,12,16,17]:
            ws1.cell(row=r, column=c).number_format = '0.0%'

    ws1.freeze_panes = 'A5'
    autosize(ws1)

    # ---- Sheet 2: Expense Detail ----
    ws2 = wb.create_sheet('Expense Detail')
    H2  = ['Address','Gross Income','Vacancy (3%)','Municipal Tax',
           'School Tax','Utilities','Insurance (4%)','Maintenance (3%)',
           'Management (5%)','Other (2%)','Total Expenses','Expense Ratio','NOI']
    hdr(ws2, H2)

    for r, a in enumerate(analyses, 2):
        ed   = a['expense_detail']
        vals = [a['address'], a['gross_income'], ed['vacancy'],
                ed['municipal'], ed['school'], ed['utilities'],
                ed['insurance'], ed['maintenance'], ed['management'], ed['other'],
                a['total_expenses'], a['expense_ratio'], a['noi']]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=v)
        for c in [2,3,4,5,6,7,8,9,10,11,13]:
            ws2.cell(row=r, column=c).number_format = '$#,##0'
        ws2.cell(row=r, column=12).number_format = '0.0%'
    autosize(ws2)

    # ---- Sheet 3: All Scenarios ----
    ws3 = wb.create_sheet('All Scenarios')
    H3  = ['Address','Scenario','TGA','Economic Value','Value Ratio',
           'Down Payment','Monthly Payment','Monthly Cashflow','Cash ROI']
    hdr(ws3, H3)
    row = 2
    for a in analyses:
        for sc_name, sc in a['all_scenarios'].items():
            ratio = sc['value_ratio']
            vals  = [a['address'], sc_name, sc['tga'], sc['economic_value'], ratio,
                     sc['down_payment'], sc['monthly_payment'],
                     sc['monthly_cashflow'], sc['cash_roi']]
            for c, v in enumerate(vals, 1):
                ws3.cell(row=row, column=c, value=v).fill = (
                    GREEN if ratio >= 1.15 else YELLOW if ratio >= 1.0 else RED
                )
            for c in [4,6,7,8]:
                ws3.cell(row=row, column=c).number_format = '$#,##0'
            for c in [3,5,9]:
                ws3.cell(row=row, column=c).number_format = '0.0%'
            row += 1
        row += 1
    ws3.freeze_panes = 'A2'
    autosize(ws3)

    # ---- Sheet 4: Raw Data ----
    ws4 = wb.create_sheet('Raw Data (all scraped)')
    if os.path.exists(CSV_FILE):
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                raw_df = pd.read_csv(CSV_FILE, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        hdr(ws4, list(raw_df.columns))
        for r, row_data in enumerate(raw_df.values.tolist(), 2):
            for c, v in enumerate(row_data, 1):
                ws4.cell(row=r, column=c, value=v)
        autosize(ws4)

    wb.save(filename)
    return filename


# ============================================================================
# MAIN
# ============================================================================

def main():

    print()
    print("=" * 60)
    print("  MONTREAL / LAVAL PLEX ANALYZER")
    print("=" * 60)
    print(f"  Min units : {CONFIG['min_units']}+")
    print(f"  Max price : ${CONFIG['max_price']:,}")
    print(f"  Cities    : Montreal, Laval")
    print(f"  Insurance : {CONFIG['insurance_pct']*100:.0f}% of rents")
    print()

    # STEP 1: Scrape
    scraped = scrape_centris()

    # STEP 2: Save CSV + pause
    print("\n" + "="*60)
    print("STEP 2: SAVING TO CSV")
    print("="*60)

    if scraped:
        save_to_csv(scraped)
    elif not os.path.exists(CSV_FILE):
        df = pd.DataFrame(columns=[
            'address','price','num_units','gross_income',
            'municipal_tax','school_tax','utilities','url','scraped_at'
        ])
        df.to_csv(CSV_FILE, index=False, encoding='utf-8-sig')
        print(f"Created empty template: {CSV_FILE}")

    pause_for_manual_edits()

    # STEP 3: Load
    print("="*60)
    print("STEP 3: LOADING PROPERTIES")
    print("="*60)
    properties = load_csv()

    if not properties:
        print(f"\nNo properties with income data.")
        print(f"Open '{CSV_FILE}' and fill in the 'gross_income' column.")
        return

    # STEP 4: Analyze
    print("\n" + "="*60)
    print("STEP 4: CALCULATING ECONOMIC VALUE")
    print("="*60 + "\n")

    analyses = []
    for prop in properties:
        result = analyze_property(prop)
        if result:
            analyses.append(result)
            r    = result['best_value_ratio']
            icon = 'GREAT' if r >= 1.15 else 'GOOD' if r >= 1.0 else 'SKIP'
            print(f"  [{icon}]  {prop['address'][:45]}")
            print(f"          Value Ratio: {r:.1%}  |  "
                  f"Cashflow: ${result['best_monthly_cf']:,.0f}/mo  |  "
                  f"ROI: {result['best_cash_roi']:.1%}")
        else:
            print(f"  [SKIP]  {prop['address'][:45]}  (NOI negative)")

    # STEP 5: Report
    print("\n" + "="*60)
    print("STEP 5: GENERATING EXCEL REPORT")
    print("="*60)

    if not analyses:
        print("Nothing to report.")
        return

    report_file = generate_report(analyses)

    # Summary
    good = [a for a in analyses if a['best_value_ratio'] >= 1.15]
    ok   = [a for a in analyses if 1.0 <= a['best_value_ratio'] < 1.15]
    bad  = [a for a in analyses if a['best_value_ratio'] < 1.0]
    best = analyses[0]

    print()
    print("=" * 60)
    print("  RESULTS")
    print("=" * 60)
    print(f"  Excellent deals (>115%) : {len(good)}")
    print(f"  Good deals (100-115%)   : {len(ok)}")
    print(f"  Overpriced (<100%)      : {len(bad)}")
    print()
    print(f"  BEST DEAL:")
    print(f"  {best['address']}")
    print(f"  Price          : ${best['price']:,.0f}")
    print(f"  Economic Value : ${best['best_econ_value']:,.0f}")
    print(f"  Value Ratio    : {best['best_value_ratio']:.1%}")
    print(f"  Monthly CF     : ${best['best_monthly_cf']:,.0f}/month")
    print(f"  Cash ROI       : {best['best_cash_roi']:.1%}")
    if best.get('url'):
        print(f"  URL            : {best['url']}")
    print()
    print(f"  Report saved: {report_file}")
    print()


if __name__ == '__main__':
    main()
